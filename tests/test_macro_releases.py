from __future__ import annotations

import hashlib
import uuid
from copy import deepcopy
from datetime import UTC, datetime, timedelta
from decimal import Decimal
from io import BytesIO
from pathlib import Path
from types import SimpleNamespace

import httpx
import pytest
from django.db import connection
from django.db.models.query import QuerySet
from django.test.utils import CaptureQueriesContext
from openpyxl import Workbook

from research.consumer_credit import (
    G19_SERIES,
    HHDC_BALANCE_SERIES,
    HHDC_DELINQUENCY_SERIES,
)
from research.data_catalog import DATA_REQUIREMENTS
from research.macro_contract import (
    GDP_CONTRACT_VERSION,
    GDP_REQUIRED_CHART_KEYS,
    GDP_REQUIRED_METRIC_KEYS,
    GDP_REQUIRED_SECTION_KEYS,
    _validate_gdp_semantic_alignment,
    coordinate_gdp_dashboard,
    gdp_snapshot_is_publicly_displayable,
    publish_gdp_revision,
    select_public_gdp_snapshot,
)
from research.macro_releases import (
    BEA_GDP_PAGE,
    BEA_PIO_PAGE,
    BEA_PIO_SECTION2_WORKBOOK,
    BEA_VINTAGE_WORKBOOK,
    CENSUS_MARTS_CURRENT_WORKBOOK,
    CENSUS_MARTS_INDEX,
    XLSX_CONTENT_TYPE,
    BEAGDPReleaseProvider,
    BEAPIOReleaseProvider,
    CensusMARTSReleaseProvider,
)
from research.models import (
    DashboardSnapshot,
    IngestionRun,
    MetricSnapshot,
    Observation,
    RawArtifact,
    ReleaseVintageObservation,
    SeriesDefinition,
)
from research.official_data import (
    ECONOMY_COMPONENTS,
    MACRO_PUBLICATION_GROUPS,
    _economy_component_payload,
    _keys_with_current_required_batches,
    _mark_latest_dashboards_stale,
    _publish_dashboard,
    _publishable_keys_for_source_groups,
    _record_and_coordinate_gdp_result,
    _record_census_revision_witness,
    _store_bea_release_observations_v2,
    _store_census_marts_observations_v2,
    _store_release_workbook_observations,
    publish_official_dashboards,
)
from research.providers import ProviderResult
from research.raw_evidence import parse_evidence_bundle
from research.services import (
    begin_ingestion,
    ensure_source,
    finish_ingestion,
    record_provider_result,
    store_series_observations,
)


def _workbook_bytes(workbook: Workbook) -> bytes:
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def _consumer_credit_results() -> tuple[ProviderResult, ProviderResult]:
    g19_records = []
    for period, offset in (("2026-04-01", Decimal("0")), ("2026-05-01", Decimal("1"))):
        for index, (_, (series_id, _)) in enumerate(G19_SERIES.items(), start=1):
            g19_records.append(
                {
                    "series_id": series_id,
                    "date": period,
                    "value": Decimal(index) + offset,
                }
            )
    household_records = []
    household_series = [
        *HHDC_BALANCE_SERIES.values(),
        *HHDC_DELINQUENCY_SERIES.values(),
    ]
    for period, offset in (("2025-12-31", Decimal("0")), ("2026-03-31", Decimal("1"))):
        for index, series_id in enumerate(household_series, start=1):
            household_records.append(
                {
                    "series_id": series_id,
                    "date": period,
                    "value": Decimal(index) + offset,
                }
            )
    return (
        ProviderResult(
            provider="federal-reserve-g19",
            dataset="g19-fixture",
            records=g19_records,
        ),
        ProviderResult(
            provider="ny-fed-household-credit",
            dataset="hhdc-fixture",
            records=household_records,
        ),
    )


def _census_api_result() -> ProviderResult:
    records = []
    values = {
        "2026-03-01": ("754013", "1.7", "4.2"),
        "2026-04-01": ("757036", "0.4", "4.8"),
        "2026-05-01": ("763705", "0.9", "6.9"),
    }
    for period, (level, mom, yoy) in values.items():
        records.extend(
            [
                {
                    "series_id": "CENSUS-MRTS-44X72-SM-SA",
                    "date": period,
                    "value": level,
                },
                {
                    "series_id": "CENSUS-MRTS-44X72-SM-SA-MOM",
                    "date": period,
                    "value": mom,
                },
                {
                    "series_id": "CENSUS-MRTS-44X72-SM-SA-YOY",
                    "date": period,
                    "value": yoy,
                },
            ]
        )
    return ProviderResult(
        provider="census",
        dataset="marts:44X72:SM:yes",
        records=records,
    )


def _bea_vintage_workbook() -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Vintage History"
    sheet.append(["Last Updated June 25, 2026"])
    sheet.append(["2026Q1"])
    sheet.append([None, "Vintage", "GDP", "GDI", "Real GDP", "Real GDI", "Release Date"])
    sheet.append([None, "Third", "31,865.7", "31,574.2", "2.1", "1.2", "Jun 25, 2026"])
    sheet.append([None, "Second", "31,819.5", "31,539.4", "1.6", "0.9", "May 28, 2026"])
    sheet.append(["2025Q4"])
    sheet.append([None, "Vintage", "GDP", "GDI", "Real GDP", "Real GDI", "Release Date"])
    sheet.append(
        [
            None,
            "Revised",
            "31,422.5",
            "31,199.9",
            "0.5",
            "1.6",
            "May 28, 2026 GDP not open for revision",
        ]
    )
    return _workbook_bytes(workbook)


def _bea_comparison_workbook(
    *,
    quarter: str = "2026Q1",
    release_date: str = "June 25, 2026",
    estimate_round: str = "Third",
) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "GDPhistQ"
    sheet["A1"] = release_date
    sheet["A2"] = (
        f"{quarter} ({estimate_round} Estimate) Comparisons -- Percent Change from Preceding Period "
        "in Real Gross Domestic Product and Related Measures"
    )
    rows = [
        ("Gross domestic product (GDP)", 2.1, "2025Q4", 0.5),
        ("Personal consumption expenditures", 0.5, "2025Q4", 1.9),
        ("Goods", 0.5, "2025Q4", 0.3),
        ("Services", 0.5, "2025Q4", 2.7),
        ("Gross private domestic investment", 7.9, "2025Q4", 2.3),
        ("Fixed investment", 6.5, "2025Q4", 1.5),
        ("Exports", 5.4, "2025Q4", -3.1),
        ("Imports", 1.8, "2025Q4", -1.2),
        ("Government consumption expenditures and gross investment", 1.2, "2025Q4", 0.4),
    ]
    for row_number, (label, current, previous_period, previous) in enumerate(rows, start=6):
        sheet.cell(row_number, 1, label)
        sheet.cell(row_number, 2, current)
        sheet.cell(row_number, 5, previous_period)
        sheet.cell(row_number, 6, previous)
    sheet.cell(
        20,
        1,
        f"{quarter} ({estimate_round} Estimate) Comparisons -- Contributions to Percent Change "
        "in Real Gross Domestic Product",
    )
    contribution_rows = [
        ("Personal consumption expenditures", 0.37, "2025Q4", 1.30),
        ("Gross private domestic investment", 1.35, "2025Q4", 0.40),
        ("Net exports of goods and services", -0.37, "2025Q4", 0.46),
        ("Government consumption expenditures and gross investment", 0.74, "2025Q4", -0.99),
    ]
    for row_number, (label, current, previous_period, previous) in enumerate(
        contribution_rows, start=24
    ):
        sheet.cell(row_number, 1, label)
        sheet.cell(row_number, 2, current)
        sheet.cell(row_number, 5, previous_period)
        sheet.cell(row_number, 6, previous)
    return _workbook_bytes(workbook)


def _census_workbook(
    *,
    sales_declaration: str = (
        "(Total sales estimates are shown in millions of dollars and are based on official survey data.)"
    ),
    change_declaration: str = (
        "(Estimates are shown as percents and are based on official survey data.)"
    ),
) -> bytes:
    workbook = Workbook()
    sales = workbook.active
    sales.title = "Table 1."
    sales.cell(
        4,
        1,
        sales_declaration,
    )
    sales.cell(6, 10, "Adjusted2")
    sales.cell(7, 10, 2026)
    sales.cell(7, 13, 2025)
    for column, label in enumerate(("Apr.3", "Mar.", "Feb.", "Apr.", "Mar."), start=10):
        sales.cell(8, column, label)
    for column, status in enumerate(("(a)", "(p)", "(r)", "(r)", "(r)"), start=10):
        sales.cell(9, column, status)
    sales.cell(11, 2, "Retail & food services, ")
    sales.cell(12, 2, "  total")
    for column, value in enumerate((757085, 753370, 741278, 721903, 723339), start=10):
        sales.cell(12, column, value)

    changes = workbook.create_sheet("Table 2.")
    changes.cell(
        3,
        1,
        change_declaration,
    )
    changes.cell(8, 3, "Apr. 2026 Advance")
    changes.cell(8, 5, "Mar. 2026 Preliminary")
    changes.cell(11, 3, "Mar. 2026")
    changes.cell(11, 4, "Apr. 2025")
    changes.cell(11, 5, "Feb. 2026")
    changes.cell(11, 6, "Mar. 2025")
    changes.cell(14, 2, "Retail & food services, ")
    changes.cell(15, 2, "  total")
    changes.cell(15, 3, 0.5)
    changes.cell(15, 4, 4.9)
    changes.cell(15, 5, 1.6)
    changes.cell(15, 6, 4.2)
    return _workbook_bytes(workbook)


def _census_current_workbook(
    *,
    sales_declaration: str = (
        "(Total sales estimates are shown in millions of dollars and are based on official survey data.)"
    ),
    change_declaration: str = (
        "(Estimates are shown as percents and are based on official survey data.)"
    ),
) -> bytes:
    workbook = Workbook()
    sales = workbook.active
    sales.title = "Table 1."
    sales.cell(
        4,
        1,
        sales_declaration,
    )
    sales.cell(6, 10, "Adjusted2")
    sales.cell(7, 10, 2026)
    sales.cell(7, 13, 2025)
    for column, label in enumerate(("May.3", "Apr.", "Mar.", "May.", "Apr."), start=10):
        sales.cell(8, column, label)
    for column, status in enumerate(("(a)", "(r)", "(r)", "(r)", "(r)"), start=10):
        sales.cell(9, column, status)
    sales.cell(11, 2, "Retail & food services, ")
    sales.cell(12, 2, "  total")
    for column, value in enumerate((763705, 757036, 754013, 714568, 721903), start=10):
        sales.cell(12, column, value)

    changes = workbook.create_sheet("Table 2.")
    changes.cell(
        3,
        1,
        change_declaration,
    )
    changes.cell(8, 3, "May. 2026 Advance")
    changes.cell(8, 5, "Apr. 2026 Revised")
    changes.cell(11, 3, "Apr. 2026")
    changes.cell(11, 4, "May. 2025")
    changes.cell(11, 5, "Mar. 2026")
    changes.cell(11, 6, "Apr. 2025")
    changes.cell(14, 2, "Retail & food services, ")
    changes.cell(15, 2, "  total")
    changes.cell(15, 3, 0.9)
    changes.cell(15, 4, 6.9)
    changes.cell(15, 5, 0.4)
    changes.cell(15, 6, 4.8)
    return _workbook_bytes(workbook)


def _bea_pio_summary_workbook(*, real_pce: float = 0.3) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "PIOhist_M"
    sheet["G1"] = datetime(2026, 6, 25)
    sheet["A2"] = "May 2026 Personal Income and Outlays"
    sheet["A3"] = "Historical Comparisons"
    sheet["B5"] = datetime(2026, 5, 1)
    sheet["A16"] = "Chained dollars"
    sheet["A20"] = "Percent change from preceding month:"
    sheet["A21"] = "DPI"
    sheet["B21"] = 0.3
    sheet["A22"] = "PCE"
    sheet["B22"] = real_pce
    sheet["A30"] = "Personal saving as a percentage of DPI"
    sheet["A31"] = "Personal saving rate"
    sheet["B31"] = 3.0
    return _workbook_bytes(workbook)


def _bea_pio_section2_workbook(
    *, duplicate_code: bool = False, missing_middle: bool = False
) -> bytes:
    workbook = Workbook()
    section_206 = workbook.active
    section_206.title = "T20600-M"
    section_20801 = workbook.create_sheet("T20801-M")
    section_20806 = workbook.create_sheet("T20806-M")
    section_20804 = workbook.create_sheet("T20804-M")
    for sheet, title in (
        (section_206, "Table 2.6. Personal Income and Its Disposition, Monthly"),
        (
            section_20801,
            "Table 2.8.1. Percent Change From Preceding Period in Real "
            "Personal Consumption Expenditures by Major Type of Product, Monthly",
        ),
        (
            section_20804,
            "Table 2.8.4. Price Indexes for Personal Consumption Expenditures "
            "by Major Type of Product, Monthly",
        ),
        (
            section_20806,
            "Table 2.8.6. Real Personal Consumption Expenditures by Major "
            "Type of Product, Monthly, Chained Dollars",
        ),
    ):
        sheet["A1"] = title
        sheet["A5"] = "Data published June 25, 2026"
    section_206["A2"] = (
        "[Millions of dollars; months are seasonally adjusted at annual rates]"
    )
    section_20801["A2"] = "[Percent]"
    section_20806["A2"] = (
        "[Millions of chained (2017) dollars; seasonally adjusted at annual rates]"
    )
    section_20804["A2"] = "[Index numbers, 2017=100; seasonally adjusted]"

    def monthly_periods(start_year: int, start_month: int) -> list[str]:
        periods = []
        year, month = start_year, start_month
        while (year, month) <= (2026, 5):
            periods.append(f"{year:04d}M{month:02d}")
            if month == 12:
                year, month = year + 1, 1
            else:
                month += 1
        return periods

    periods_206 = monthly_periods(1959, 1)
    periods_20801 = monthly_periods(1959, 2)
    periods_20804 = monthly_periods(1959, 1)
    periods_20806 = monthly_periods(2007, 1)
    for sheet, periods in (
        (section_206, periods_206),
        (section_20801, periods_20801),
        (section_20804, periods_20804),
        (section_20806, periods_20806),
    ):
        sheet["A3"] = f"Monthly data from {periods[0]} to {periods[-1]}"
        for offset, period in enumerate(periods, start=4):
            sheet.cell(8, offset, period)

    rows_206 = (
        (35, "Disposable personal income", "A067RC", 20000000, 23486851, 23651714, False),
        (43, "Personal saving rate", "A072RC", 5.0, 3.0, 3.0, False),
        (
            47,
            "Real disposable personal income, chained (2017) dollars",
            "A067RX",
            15000000,
            17938761,
            17983827,
            False,
        ),
        (53, "Disposable personal income MoM", "A067RCM", 0.1, -0.1, 0.7, True),
        (
            55,
            "Real disposable personal income, chained (2017) dollars, MoM",
            "A067RM",
            0.1,
            -0.5,
            0.3,
            True,
        ),
    )
    for row_number, label, code, default, previous, current, leading_blank in rows_206:
        section_206.cell(row_number, 2, label)
        section_206.cell(row_number, 3, code)
        for offset, _period in enumerate(periods_206, start=4):
            if leading_blank and offset == 4:
                continue
            section_206.cell(row_number, offset, default)
        section_206.cell(row_number, 3 + len(periods_206) - 1, previous)
        section_206.cell(row_number, 3 + len(periods_206), current)
    if duplicate_code:
        section_206.cell(56, 2, "Duplicate real DPI")
        section_206.cell(56, 3, "A067RM")
        section_206.cell(56, 4, -0.5)
        section_206.cell(56, 5, 0.3)

    section_20801.cell(9, 2, "Personal consumption expenditures")
    section_20801.cell(9, 3, "DPCERAM")
    for offset, _period in enumerate(periods_20801, start=4):
        section_20801.cell(9, offset, 0.1)
    section_20801.cell(9, 3 + len(periods_20801) - 1, 0.0)
    section_20801.cell(9, 3 + len(periods_20801), 0.3)
    if missing_middle:
        section_20801.cell(9, 100, ".....")
    section_20806.cell(9, 2, "Personal consumption expenditures")
    section_20806.cell(9, 3, "DPCERX")
    for offset, _period in enumerate(periods_20806, start=4):
        section_20806.cell(9, offset, 15000000)
    section_20806.cell(9, 3 + len(periods_20806) - 1, 16729609)
    section_20806.cell(9, 3 + len(periods_20806), 16773429)
    section_20804.cell(9, 2, "Personal consumption expenditures (PCE)")
    section_20804.cell(9, 3, "DPCERG")
    section_20804.cell(33, 2, "PCE excluding food and energy")
    section_20804.cell(33, 3, "DPCCRG")
    for offset, _period in enumerate(periods_20804, start=4):
        section_20804.cell(9, offset, 100.0)
        section_20804.cell(33, offset, 200.0)
    section_20804.cell(9, 3 + len(periods_20804) - 12, 120.0)
    section_20804.cell(9, 3 + len(periods_20804) - 6, 124.0)
    section_20804.cell(9, 3 + len(periods_20804) - 3, 125.0)
    section_20804.cell(9, 3 + len(periods_20804) - 1, 129.0)
    section_20804.cell(9, 3 + len(periods_20804), 130.0)
    section_20804.cell(33, 3 + len(periods_20804) - 12, 250.0)
    section_20804.cell(33, 3 + len(periods_20804) - 6, 254.0)
    section_20804.cell(33, 3 + len(periods_20804) - 3, 255.0)
    section_20804.cell(33, 3 + len(periods_20804) - 1, 259.0)
    section_20804.cell(33, 3 + len(periods_20804), 260.0)
    return _workbook_bytes(workbook)


def _bea_client(vintage: bytes, comparisons: bytes) -> httpx.Client:
    comparison_url = "https://www.bea.gov/sites/default/files/2026-06/hist1q26-3rd.xlsx"

    def handler(request: httpx.Request) -> httpx.Response:
        url = str(request.url)
        if url == BEA_GDP_PAGE:
            return httpx.Response(
                200,
                text=(
                    '<html><a href="/sites/default/files/2026-06/hist1q26-3rd.xlsx">'
                    "Historical Comparisons</a></html>"
                ),
                headers={"content-type": "text/html"},
            )
        if url == BEA_VINTAGE_WORKBOOK:
            return httpx.Response(
                200,
                content=vintage,
                headers={"content-type": XLSX_CONTENT_TYPE, "last-modified": "June 25, 2026"},
            )
        if url == comparison_url:
            return httpx.Response(
                200,
                content=comparisons,
                headers={"content-type": XLSX_CONTENT_TYPE, "last-modified": "June 25, 2026"},
            )
        raise AssertionError(f"unexpected URL: {url}")

    return httpx.Client(transport=httpx.MockTransport(handler), follow_redirects=True)


def _census_client(workbook: bytes, *, current_status: int = 403) -> httpx.Client:
    latest = f"{CENSUS_MARTS_INDEX}rs2604.xlsx"

    def handler(request: httpx.Request) -> httpx.Response:
        url = str(request.url)
        if url == CENSUS_MARTS_CURRENT_WORKBOOK:
            if current_status == 200:
                return httpx.Response(
                    200,
                    content=workbook,
                    headers={
                        "content-type": XLSX_CONTENT_TYPE,
                        "last-modified": "June 17, 2026",
                    },
                )
            return httpx.Response(current_status, text="current workbook unavailable")
        if url == CENSUS_MARTS_INDEX:
            return httpx.Response(
                200,
                text=(
                    '<html><a href="rs9912.xlsx">old</a>'
                    '<a href="rs2603.xlsx">prior</a>'
                    '<a href="rs2604.xlsx">latest</a></html>'
                ),
                headers={"content-type": "text/html"},
            )
        if url == latest:
            return httpx.Response(
                200,
                content=workbook,
                headers={"content-type": XLSX_CONTENT_TYPE, "last-modified": "May 15, 2026"},
            )
        raise AssertionError(f"unexpected URL: {url}")

    return httpx.Client(transport=httpx.MockTransport(handler), follow_redirects=True)


def _bea_pio_client(summary: bytes, section2: bytes) -> httpx.Client:
    summary_url = "https://www.bea.gov/sites/default/files/2026-06/pi0526-hist.xlsx"

    def handler(request: httpx.Request) -> httpx.Response:
        url = str(request.url)
        if url == BEA_PIO_PAGE:
            return httpx.Response(
                200,
                text=(
                    '<html><a href="/sites/default/files/2026-06/pi0526-hist.xlsx">'
                    "Historical Comparisons</a></html>"
                ),
                headers={"content-type": "text/html"},
            )
        if url == summary_url:
            return httpx.Response(
                200,
                content=summary,
                headers={"content-type": XLSX_CONTENT_TYPE, "last-modified": "June 25, 2026"},
            )
        if url == BEA_PIO_SECTION2_WORKBOOK:
            return httpx.Response(
                200,
                content=section2,
                headers={"content-type": XLSX_CONTENT_TYPE, "last-modified": "June 25, 2026"},
            )
        raise AssertionError(f"unexpected URL: {url}")

    return httpx.Client(transport=httpx.MockTransport(handler), follow_redirects=True)


def test_bea_release_provider_preserves_vintages_components_and_artifact_hashes():
    vintage = _bea_vintage_workbook()
    comparisons = _bea_comparison_workbook()
    result = BEAGDPReleaseProvider(client=_bea_client(vintage, comparisons)).gdp_pce()

    assert result.ok
    assert result.metadata["quarter_count"] == 2
    assert result.metadata["comparison_quarter"] == "2026Q1"
    assert result.metadata["comparison_release_date"] == "2026-06-25"
    assert result.metadata["comparison_estimate_round"] == "Third"
    assert result.metadata["vintage_release_count"] == 3
    assert result.metadata["vintage_observation_count"] == 12
    assert len(result.metadata["artifacts"]) == 3
    assert result.metadata["artifacts"][1]["sha256"] == hashlib.sha256(vintage).hexdigest()
    evidence = parse_evidence_bundle(
        result.raw_bytes,
        expected_provider="bea-release",
        expected_dataset="gdp-release-workbooks",
    )
    assert set(evidence.responses) == {
        "release-page",
        "vintage-workbook",
        "comparison-workbook",
    }
    assert evidence.responses["vintage-workbook"] == vintage
    assert evidence.responses["comparison-workbook"] == comparisons
    replay_records, replay_supplemental, replay_metadata = (
        BEAGDPReleaseProvider.replay_evidence_bundle(result.raw_bytes)
    )
    assert replay_records == result.records
    assert replay_supplemental == result.supplemental_records
    assert replay_metadata["comparison_quarter"] == "2026Q1"

    by_series_and_date = {
        (item["series_id"], item["date"]): item for item in result.records
    }
    latest_gdp = by_series_and_date[("BEA-A191RL", "2026-01-01")]
    assert latest_gdp["value"] == Decimal("2.1")
    assert latest_gdp["metadata"]["vintage_label"] == "Third"
    assert latest_gdp["metadata"]["estimate_round"] == "Third"
    assert latest_gdp["metadata"]["source_revision_date"] == "2026-06-25"
    gdp_vintages = [
        item
        for item in result.supplemental_records["release_vintages"]
        if item["series_id"] == "BEA-A191RL" and item["date"] == "2026-01-01"
    ]
    assert [item["estimate_round"] for item in gdp_vintages] == ["Third", "Second"]
    assert [item["value"] for item in gdp_vintages] == [Decimal("2.1"), Decimal("1.6")]
    assert [item["release_date"] for item in gdp_vintages] == [
        "2026-06-25",
        "2026-05-28",
    ]
    assert by_series_and_date[("BEA-DPCERL", "2026-01-01")]["value"] == Decimal("0.5")
    assert by_series_and_date[("BEA-DPCERL", "2025-10-01")]["value"] == Decimal("1.9")
    assert by_series_and_date[("BEA-GPDI-GROWTH", "2026-01-01")]["value"] == Decimal("7.9")
    assert by_series_and_date[("BEA-PCE-CONTRIBUTION", "2026-01-01")][
        "value"
    ] == Decimal("0.37")
    assert by_series_and_date[("BEA-NET-EXPORTS-CONTRIBUTION", "2026-01-01")][
        "metadata"
    ]["unit"] == "percentage points contribution to real GDP growth"


def test_bea_gdp_replay_rechecks_xlsx_expanded_size(monkeypatch):
    result = BEAGDPReleaseProvider(
        client=_bea_client(_bea_vintage_workbook(), _bea_comparison_workbook())
    ).gdp_pce()
    assert result.ok
    monkeypatch.setattr(BEAGDPReleaseProvider, "max_expanded_bytes", 1)

    with pytest.raises(ValueError, match="expanded-size limit"):
        BEAGDPReleaseProvider.replay_evidence_bundle(result.raw_bytes)


def test_bea_pio_provider_parses_full_history_codes_and_cross_checks_summary():
    summary = _bea_pio_summary_workbook()
    section2 = _bea_pio_section2_workbook()
    result = BEAPIOReleaseProvider(
        client=_bea_pio_client(summary, section2)
    ).personal_income_outlays()

    assert result.ok
    assert result.row_count == 6702
    assert result.metadata["latest_value_date"] == "2026-05-01"
    assert result.metadata["source_revision_date"] == "2026-06-25"
    assert result.metadata["summary_cross_check"] == "passed"
    artifacts = {item["url"]: item for item in result.metadata["artifacts"]}
    assert set(artifacts) == {
        BEA_PIO_PAGE,
        "https://www.bea.gov/sites/default/files/2026-06/pi0526-hist.xlsx",
        BEA_PIO_SECTION2_WORKBOOK,
    }
    assert artifacts[BEA_PIO_SECTION2_WORKBOOK]["sha256"] == hashlib.sha256(
        section2
    ).hexdigest()
    evidence = parse_evidence_bundle(
        result.raw_bytes,
        expected_provider="bea-pio-release",
        expected_dataset="personal-income-outlays-release",
    )
    assert set(evidence.responses) == {
        "release-page",
        "summary-workbook",
        "section2-workbook",
    }
    assert evidence.responses["summary-workbook"] == summary
    assert evidence.responses["section2-workbook"] == section2
    replay_records, replay_metadata = BEAPIOReleaseProvider.replay_evidence_bundle(
        result.raw_bytes
    )
    assert replay_records == result.records
    assert replay_metadata["latest_value_date"] == "2026-05-01"

    by_series_and_date = {
        (item["series_id"], item["date"]): item for item in result.records
    }
    assert len(by_series_and_date) == result.row_count
    assert by_series_and_date[("BEA-REAL-PCE-MOM", "2026-05-01")][
        "value"
    ] == Decimal("0.3")
    assert by_series_and_date[("BEA-REAL-DPI-MOM", "2026-04-01")][
        "value"
    ] == Decimal("-0.5")
    assert by_series_and_date[("BEA-PERSONAL-SAVING-RATE", "2026-05-01")][
        "value"
    ] == Decimal("3")
    assert by_series_and_date[("BEA-DPI-NOMINAL-SAAR", "2026-05-01")][
        "value"
    ] == Decimal("23651714")
    assert by_series_and_date[("BEA-DPI-REAL-SAAR", "2026-05-01")][
        "value"
    ] == Decimal("17983827")
    assert by_series_and_date[("BEA-REAL-PCE-SAAR", "2026-05-01")][
        "value"
    ] == Decimal("16773429")
    assert by_series_and_date[("BEA-PCE-PRICE-INDEX", "2026-05-01")][
        "value"
    ] == Decimal("130")
    assert by_series_and_date[("BEA-CORE-PCE-PRICE-INDEX", "2026-05-01")][
        "value"
    ] == Decimal("260")
    assert by_series_and_date[("BEA-PCE-PRICE-INDEX", "2026-05-01")]["metadata"][
        "official_series_code"
    ] == "DPCERG"
    assert by_series_and_date[("BEA-CORE-PCE-PRICE-INDEX", "2026-05-01")]["metadata"][
        "official_series_code"
    ] == "DPCCRG"
    real_pce = by_series_and_date[("BEA-REAL-PCE-MOM", "2026-05-01")]
    assert real_pce["metadata"]["official_series_code"] == "DPCERAM"
    assert real_pce["metadata"]["vintage_status"] == "current_release_vintage"
    assert by_series_and_date[("BEA-DPI-REAL-SAAR", "2026-05-01")]["metadata"][
        "reference_year"
    ] == 2017


def test_bea_pio_replay_rechecks_xlsx_expanded_size(monkeypatch):
    result = BEAPIOReleaseProvider(
        client=_bea_pio_client(
            _bea_pio_summary_workbook(),
            _bea_pio_section2_workbook(),
        )
    ).personal_income_outlays()
    assert result.ok
    monkeypatch.setattr(BEAPIOReleaseProvider, "max_expanded_bytes", 1)

    with pytest.raises(ValueError, match="expanded-size limit"):
        BEAPIOReleaseProvider.replay_evidence_bundle(result.raw_bytes)


@pytest.mark.parametrize(
    ("summary", "section2", "message"),
    [
        (
            _bea_pio_summary_workbook(real_pce=0.4),
            _bea_pio_section2_workbook(),
            "disagree for BEA-REAL-PCE-MOM",
        ),
        (
            _bea_pio_summary_workbook(),
            _bea_pio_section2_workbook(duplicate_code=True),
            "duplicated NIPA code A067RM",
        ),
        (
            _bea_pio_summary_workbook(),
            _bea_pio_section2_workbook(missing_middle=True),
            "missing/non-numeric value after history began",
        ),
    ],
)
def test_bea_pio_provider_fails_closed_on_inconsistent_workbooks(
    summary, section2, message
):
    result = BEAPIOReleaseProvider(
        client=_bea_pio_client(summary, section2)
    ).personal_income_outlays()

    assert not result.ok
    assert result.records == []
    assert message in result.error


@pytest.mark.django_db
def test_bea_gdp_v2_keeps_private_bundle_and_append_only_release_vintages(
    settings,
    tmp_path,
):
    settings.RAW_ARTIFACT_ROOT = tmp_path / "raw-artifacts"
    first_result = BEAGDPReleaseProvider(
        client=_bea_client(_bea_vintage_workbook(), _bea_comparison_workbook())
    ).gdp_pce()
    first_raw = bytes(first_result.raw_bytes)
    first = record_provider_result(
        first_result,
        persist=_store_bea_release_observations_v2,
    )

    assert first.status == IngestionRun.Status.SUCCESS
    assert first.row_count == first_result.row_count + len(
        first_result.supplemental_records["release_vintages"]
    )
    assert Observation.objects.filter(batch_id=first.batch_id).count() == len(
        first_result.records
    )
    assert ReleaseVintageObservation.objects.filter(
        batch_id=first.batch_id
    ).count() == len(first_result.supplemental_records["release_vintages"])
    first_artifact = RawArtifact.objects.get(run=first)
    assert first_artifact.uri.startswith("private://bea-release/")
    first_path = (
        Path(settings.RAW_ARTIFACT_ROOT)
        / first_artifact.sha256[:2]
        / f"{first_artifact.sha256}.bin"
    )
    assert first_path.read_bytes() == first_raw
    first_observations = list(
        Observation.objects.filter(batch_id=first.batch_id)
        .order_by("pk")
        .values_list("pk", "value", "updated_at")
    )
    first_vintages = list(
        ReleaseVintageObservation.objects.filter(batch_id=first.batch_id)
        .order_by("pk")
        .values_list("pk", "value", "updated_at")
    )

    second_result = BEAGDPReleaseProvider(
        client=_bea_client(_bea_vintage_workbook(), _bea_comparison_workbook())
    ).gdp_pce()
    second = record_provider_result(
        second_result,
        persist=_store_bea_release_observations_v2,
    )

    assert second.status == IngestionRun.Status.SUCCESS
    assert second.batch_id != first.batch_id
    assert RawArtifact.objects.filter(run__in=(first, second)).count() == 2
    assert RawArtifact.objects.get(run=second).sha256 == first_artifact.sha256
    assert Observation.objects.filter(batch_id=second.batch_id).count() == len(
        second_result.records
    )
    assert ReleaseVintageObservation.objects.filter(
        batch_id=second.batch_id
    ).count() == len(second_result.supplemental_records["release_vintages"])
    assert list(
        Observation.objects.filter(batch_id=first.batch_id)
        .order_by("pk")
        .values_list("pk", "value", "updated_at")
    ) == first_observations
    assert list(
        ReleaseVintageObservation.objects.filter(batch_id=first.batch_id)
        .order_by("pk")
        .values_list("pk", "value", "updated_at")
    ) == first_vintages


def _strict_gdp_run(*, vintage: bytes, comparisons: bytes) -> IngestionRun:
    result = BEAGDPReleaseProvider(
        client=_bea_client(vintage, comparisons)
    ).gdp_pce()
    assert result.ok, result.error
    run = record_provider_result(
        result,
        persist=_store_bea_release_observations_v2,
    )
    assert run.status == IngestionRun.Status.SUCCESS, run.error
    return run


@pytest.mark.django_db
def test_gdp_v2_dedicated_publisher_selector_and_route(
    client,
    settings,
    tmp_path,
):
    settings.RAW_ARTIFACT_ROOT = tmp_path / "raw-artifacts"
    run = _strict_gdp_run(
        vintage=_bea_vintage_workbook(),
        comparisons=_bea_comparison_workbook(),
    )

    snapshot = publish_gdp_revision(run=run)

    assert snapshot is not None
    assert snapshot.data["contract_version"] == GDP_CONTRACT_VERSION
    assert {item["key"] for item in snapshot.data["metrics"]} == (
        GDP_REQUIRED_METRIC_KEYS
    )
    assert {item["key"] for item in snapshot.data["charts"]} == (
        GDP_REQUIRED_CHART_KEYS
    )
    assert {item["key"] for item in snapshot.data["sections"]} == (
        GDP_REQUIRED_SECTION_KEYS
    )
    assert snapshot.data["component_batches"] == [str(run.batch_id)]
    assert snapshot.data["input_run"]["ingestion_run_id"] == run.pk
    assert MetricSnapshot.objects.filter(
        batch_id=snapshot.batch_id,
        key__startswith="gdp-",
    ).count() == len(GDP_REQUIRED_METRIC_KEYS)
    selected = select_public_gdp_snapshot()
    assert selected is not None
    assert selected.pk == snapshot.pk
    assert selected.gdp_publication_state == "current_candidate"
    assert gdp_snapshot_is_publicly_displayable(snapshot)
    economy_component = _economy_component_payload(
        "gdp",
        ECONOMY_COMPONENTS["gdp"],
        now=datetime.fromisoformat(snapshot.data["fresh_until"])
        - timedelta(seconds=1),
    )
    assert isinstance(economy_component, tuple)

    response = client.get("/economy/gdp/")
    body = response.content.decode()
    assert response.status_code == 200
    assert body.count(" data-chart ") == 2
    assert "GDP 发布轮次与修订路径" in body
    assert "1.60% → 2.10%" in body

    assert publish_gdp_revision(run=run) is None
    assert DashboardSnapshot.objects.filter(
        key="gdp",
        data__contract_version=GDP_CONTRACT_VERSION,
    ).count() == 1
    assert publish_official_dashboards(keys={"gdp"}) == []
    with pytest.raises(ValueError, match="dedicated macro v2"):
        _publish_dashboard(
            key="gdp",
            title=snapshot.title,
            summary=snapshot.summary,
            metrics=snapshot.data["metrics"],
            charts=snapshot.data["charts"],
            sections=snapshot.data["sections"],
            batch_id=uuid.uuid4(),
        )


@pytest.mark.django_db
def test_gdp_publisher_locks_sources_before_runs_and_nullable_rows(
    monkeypatch,
    settings,
    tmp_path,
):
    settings.RAW_ARTIFACT_ROOT = tmp_path / "raw-artifacts"
    run = _strict_gdp_run(
        vintage=_bea_vintage_workbook(),
        comparisons=_bea_comparison_workbook(),
    )
    original = QuerySet.select_for_update
    lock_models = []
    joined_lock_calls = []
    nullable_lock_shapes = []

    def recording_select_for_update(queryset, *args, **kwargs):
        lock_models.append(queryset.model)
        if queryset.model in {IngestionRun, DashboardSnapshot}:
            joined_lock_calls.append((queryset.model, kwargs.get("of")))
        if queryset.model in {Observation, ReleaseVintageObservation}:
            nullable_lock_shapes.append(
                (queryset.model, queryset.query.select_related)
            )
        return original(queryset, *args, **kwargs)

    monkeypatch.setattr(QuerySet, "select_for_update", recording_select_for_update)
    published = publish_gdp_revision(run=run)
    assert published is not None
    failure_at = datetime.now(UTC)
    IngestionRun.objects.create(
        source=run.source,
        dataset=run.dataset,
        started_at=failure_at,
        completed_at=failure_at,
        status=IngestionRun.Status.FAILED,
        error="lock-path fixture",
    )
    coordinate_gdp_dashboard()

    assert lock_models.index(run.source.__class__) < lock_models.index(IngestionRun)
    assert joined_lock_calls
    assert all(of == ("self",) for _model, of in joined_lock_calls)
    assert nullable_lock_shapes == [
        (Observation, False),
        (ReleaseVintageObservation, False),
    ]


@pytest.mark.django_db
def test_gdp_v2_same_values_new_run_is_new_revision_and_rogue_is_ignored(
    settings,
    tmp_path,
):
    settings.RAW_ARTIFACT_ROOT = tmp_path / "raw-artifacts"
    vintage = _bea_vintage_workbook()
    comparisons = _bea_comparison_workbook()
    first_run = _strict_gdp_run(vintage=vintage, comparisons=comparisons)
    first = publish_gdp_revision(run=first_run)
    assert first is not None

    second_run = _strict_gdp_run(vintage=vintage, comparisons=comparisons)
    second = publish_gdp_revision(run=second_run)

    assert second is not None
    assert second.pk != first.pk
    assert second.batch_id != first.batch_id
    assert second.data["fingerprint"] != first.data["fingerprint"]
    assert DashboardSnapshot.objects.filter(
        key="gdp",
        data__contract_version=GDP_CONTRACT_VERSION,
    ).count() == 2
    selected = select_public_gdp_snapshot()
    assert selected is not None and selected.pk == second.pk

    rogue_batch = uuid.uuid4()
    rogue_data = deepcopy(second.data)
    rogue_data["publication_batch_id"] = str(rogue_batch)
    DashboardSnapshot.objects.create(
        key="gdp",
        title=second.title,
        summary=second.summary,
        as_of=second.as_of,
        batch_id=rogue_batch,
        quality_status=second.quality_status,
        data=rogue_data,
        source=second.source,
        is_published=True,
    )
    selected = select_public_gdp_snapshot()
    assert selected is not None and selected.pk == second.pk


@pytest.mark.parametrize("malformed_container", ["metrics", "charts", "sections"])
@pytest.mark.django_db
def test_gdp_selector_skips_malformed_strict_looking_containers(
    client,
    malformed_container,
    settings,
    tmp_path,
):
    settings.RAW_ARTIFACT_ROOT = tmp_path / "raw-artifacts"
    run = _strict_gdp_run(
        vintage=_bea_vintage_workbook(),
        comparisons=_bea_comparison_workbook(),
    )
    published = publish_gdp_revision(run=run)
    assert published is not None
    rogue_batch = uuid.uuid4()
    rogue_data = deepcopy(published.data)
    rogue_data["publication_batch_id"] = str(rogue_batch)
    rogue_data[malformed_container] = [
        None for _item in published.data[malformed_container]
    ]
    DashboardSnapshot.objects.create(
        key="gdp",
        title=published.title,
        summary=published.summary,
        as_of=published.as_of,
        batch_id=rogue_batch,
        quality_status=published.quality_status,
        data=rogue_data,
        source=published.source,
        is_published=True,
    )

    selected = select_public_gdp_snapshot()
    assert selected is not None and selected.pk == published.pk
    assert client.get("/economy/gdp/").status_code == 200


@pytest.mark.django_db
@pytest.mark.parametrize(
    "tamper_kind",
    [
        "raw",
        "observation",
        "observation-source",
        "vintage",
        "vintage-source",
        "frequency",
        "metric",
        "payload",
        "license",
        "derived-license",
        "rogue",
    ],
)
def test_gdp_v2_tamper_fails_closed(
    settings,
    tmp_path,
    tamper_kind,
):
    settings.RAW_ARTIFACT_ROOT = tmp_path / "raw-artifacts"
    run = _strict_gdp_run(
        vintage=_bea_vintage_workbook(),
        comparisons=_bea_comparison_workbook(),
    )
    snapshot = publish_gdp_revision(run=run)
    assert snapshot is not None

    if tamper_kind == "raw":
        artifact = RawArtifact.objects.get(run=run)
        path = (
            Path(settings.RAW_ARTIFACT_ROOT)
            / artifact.sha256[:2]
            / f"{artifact.sha256}.bin"
        )
        path.write_bytes(path.read_bytes() + b"tamper")
    elif tamper_kind in {"observation", "observation-source"}:
        observation = Observation.objects.filter(batch_id=run.batch_id).first()
        assert observation is not None
        if tamper_kind == "observation-source":
            observation.source = ensure_source("bls")
            observation.save(update_fields=["source", "updated_at"])
        else:
            observation.value += Decimal("1")
            observation.save(update_fields=["value", "updated_at"])
    elif tamper_kind in {"vintage", "vintage-source"}:
        vintage = ReleaseVintageObservation.objects.filter(
            batch_id=run.batch_id
        ).first()
        assert vintage is not None
        if tamper_kind == "vintage-source":
            vintage.source = ensure_source("bls")
            vintage.save(update_fields=["source", "updated_at"])
        else:
            vintage.value += Decimal("1")
            vintage.save(update_fields=["value", "updated_at"])
    elif tamper_kind == "frequency":
        series = Observation.objects.filter(batch_id=run.batch_id).first().series
        series.frequency = "annual"
        series.save(update_fields=["frequency", "updated_at"])
    elif tamper_kind == "metric":
        metric = MetricSnapshot.objects.filter(batch_id=snapshot.batch_id).first()
        assert metric is not None and metric.value is not None
        metric.value += Decimal("1")
        metric.save(update_fields=["value", "updated_at"])
    elif tamper_kind == "payload":
        data = deepcopy(snapshot.data)
        data["metrics"][0]["value"] += 1
        snapshot.data = data
        snapshot.save(update_fields=["data", "updated_at"])
    elif tamper_kind in {"license", "derived-license"}:
        licence = run.source.licenses.get(is_current=True)
        field = (
            "derived_display_allowed"
            if tamper_kind == "derived-license"
            else "public_display_allowed"
        )
        setattr(licence, field, False)
        licence.save(update_fields=[field, "updated_at"])
    else:
        original = Observation.objects.filter(batch_id=run.batch_id).first()
        assert original is not None
        Observation.objects.create(
            series=original.series,
            instrument=None,
            value=original.value,
            value_date=original.value_date,
            as_of=original.as_of,
            fetched_at=original.fetched_at,
            batch_id=run.batch_id,
            source=original.source,
            fallback_source=None,
            quality_status=original.quality_status,
            metadata=original.metadata,
        )

    assert select_public_gdp_snapshot() is None


@pytest.mark.django_db
def test_gdp_v2_transition_failure_and_natural_expiry_states(
    monkeypatch,
    settings,
    tmp_path,
):
    settings.RAW_ARTIFACT_ROOT = tmp_path / "raw-artifacts"
    run = _strict_gdp_run(
        vintage=_bea_vintage_workbook(),
        comparisons=_bea_comparison_workbook(),
    )
    snapshot = publish_gdp_revision(run=run)
    assert snapshot is not None

    running = begin_ingestion("bea-release", "gdp-release-workbooks")
    selected = select_public_gdp_snapshot()
    assert selected is not None
    assert selected.gdp_publication_state == "transition_pending"

    failed = finish_ingestion(
        running,
        status=IngestionRun.Status.FAILED,
        error="fixture upstream failure",
    )
    published, stale = coordinate_gdp_dashboard([failed])
    assert published == []
    assert stale == {"gdp"}
    selected = select_public_gdp_snapshot()
    assert selected is not None
    assert selected.gdp_publication_state == "retained_failure"
    assert selected.data["refresh_failure"]["attempt"]["ingestion_run_id"] == failed.pk

    recovery = begin_ingestion("bea-release", "gdp-release-workbooks")
    selected = select_public_gdp_snapshot()
    assert selected is not None
    assert selected.gdp_publication_state == "transition_pending"
    assert "refresh_failure" not in selected.data
    recovery.delete()

    failed.delete()
    stored = DashboardSnapshot.objects.get(pk=snapshot.pk)
    data = deepcopy(stored.data)
    data.pop("refresh_failure", None)
    stored.data = data
    stored.quality_status = Observation.Quality.FRESH
    stored.save(update_fields=["data", "quality_status", "updated_at"])
    expired_at = datetime.fromisoformat(stored.data["fresh_until"]) + timedelta(seconds=1)
    monkeypatch.setattr("research.macro_contract.timezone.now", lambda: expired_at)
    selected = select_public_gdp_snapshot()
    assert selected is not None
    assert selected.gdp_publication_state == "natural_expiry"


@pytest.mark.django_db
def test_gdp_expired_success_rolls_back_and_retains_then_recovers(
    monkeypatch,
    settings,
    tmp_path,
):
    settings.RAW_ARTIFACT_ROOT = tmp_path / "raw-artifacts"
    baseline_run = _strict_gdp_run(
        vintage=_bea_vintage_workbook(),
        comparisons=_bea_comparison_workbook(),
    )
    baseline = publish_gdp_revision(run=baseline_run)
    assert baseline is not None
    second_run = _strict_gdp_run(
        vintage=_bea_vintage_workbook(),
        comparisons=_bea_comparison_workbook(),
    )
    snapshot_count = DashboardSnapshot.objects.filter(key="gdp").count()
    metric_count = MetricSnapshot.objects.filter(key__startswith="gdp-").count()
    expired_at = datetime.fromisoformat(baseline.data["fresh_until"]) + timedelta(seconds=1)
    monkeypatch.setattr("research.macro_contract.timezone.now", lambda: expired_at)

    dashboards, stale = coordinate_gdp_dashboard([second_run])

    assert dashboards == [] and stale == {"gdp"}
    assert DashboardSnapshot.objects.filter(key="gdp").count() == snapshot_count
    assert MetricSnapshot.objects.filter(key__startswith="gdp-").count() == metric_count
    retained = select_public_gdp_snapshot()
    assert retained is not None and retained.pk == baseline.pk
    assert retained.gdp_publication_state == "retained_failure"
    marker = retained.data["refresh_failure"]
    assert marker["reason_code"] == "publication-postcondition"
    assert marker["attempt"]["ingestion_run_id"] == second_run.pk

    current_now = datetime.now(UTC)
    monkeypatch.setattr("research.macro_contract.timezone.now", lambda: current_now)
    recovery_run = _strict_gdp_run(
        vintage=_bea_vintage_workbook(),
        comparisons=_bea_comparison_workbook(),
    )
    transitioning = select_public_gdp_snapshot()
    assert transitioning is not None
    assert transitioning.gdp_publication_state == "transition_pending"
    assert "refresh_failure" not in transitioning.data
    recovered, stale = coordinate_gdp_dashboard([recovery_run])
    assert len(recovered) == 1 and stale == set()
    assert select_public_gdp_snapshot().pk == recovered[0].pk


@pytest.mark.django_db
def test_gdp_same_run_natural_expiry_is_idempotent_and_write_free(
    monkeypatch,
    settings,
    tmp_path,
):
    settings.RAW_ARTIFACT_ROOT = tmp_path / "raw-artifacts"
    run = _strict_gdp_run(
        vintage=_bea_vintage_workbook(),
        comparisons=_bea_comparison_workbook(),
    )
    baseline = publish_gdp_revision(run=run)
    assert baseline is not None
    stored_data = deepcopy(baseline.data)
    stored_updated_at = baseline.updated_at
    snapshot_count = DashboardSnapshot.objects.filter(key="gdp").count()
    metric_count = MetricSnapshot.objects.filter(key__startswith="gdp-").count()
    expired_at = datetime.fromisoformat(baseline.data["fresh_until"]) + timedelta(seconds=1)
    monkeypatch.setattr("research.macro_contract.timezone.now", lambda: expired_at)

    with CaptureQueriesContext(connection) as captured:
        dashboards, stale = coordinate_gdp_dashboard([run])

    assert dashboards == [] and stale == {"gdp"}
    write_queries = [
        query["sql"]
        for query in captured.captured_queries
        if query["sql"].lstrip().upper().startswith(("INSERT", "UPDATE", "DELETE"))
    ]
    assert write_queries == []
    assert DashboardSnapshot.objects.filter(key="gdp").count() == snapshot_count
    assert MetricSnapshot.objects.filter(key__startswith="gdp-").count() == metric_count
    baseline.refresh_from_db()
    assert baseline.data == stored_data
    assert baseline.updated_at == stored_updated_at
    selected = select_public_gdp_snapshot()
    assert selected is not None
    assert selected.gdp_publication_state == "natural_expiry"
    assert "refresh_failure" not in selected.data


@pytest.mark.django_db
def test_gdp_v2_publication_exception_becomes_durable_retained_failure(
    monkeypatch,
    settings,
    tmp_path,
):
    settings.RAW_ARTIFACT_ROOT = tmp_path / "raw-artifacts"
    baseline_run = _strict_gdp_run(
        vintage=_bea_vintage_workbook(),
        comparisons=_bea_comparison_workbook(),
    )
    baseline = publish_gdp_revision(run=baseline_run)
    assert baseline is not None

    replacement = BEAGDPReleaseProvider(
        client=_bea_client(_bea_vintage_workbook(), _bea_comparison_workbook())
    ).gdp_pce()
    from research import macro_contract

    original_builder = macro_contract._build_gdp_payload

    def fail_publication(evidence, **kwargs):
        if evidence.run.pk != baseline_run.pk:
            raise RuntimeError("fixture publication crash")
        return original_builder(evidence, **kwargs)

    monkeypatch.setattr(
        "research.macro_contract._build_gdp_payload",
        fail_publication,
    )
    with pytest.raises(RuntimeError, match="fixture publication crash"):
        _record_and_coordinate_gdp_result(
            replacement,
            _store_bea_release_observations_v2,
        )

    failed = IngestionRun.objects.order_by("-started_at", "-id").first()
    assert failed is not None
    assert failed.status == IngestionRun.Status.FAILED
    assert failed.metadata["publication_failure"] is True
    assert "fixture publication crash" in failed.error
    assert not RawArtifact.objects.filter(run=failed).exists()
    assert IngestionRun.objects.filter(
        source__key="bea-release",
        dataset="gdp-release-workbooks",
        status=IngestionRun.Status.SUCCESS,
    ).count() == 1
    selected = select_public_gdp_snapshot()
    assert selected is not None
    assert selected.pk == baseline.pk
    assert selected.gdp_publication_state == "retained_failure"
    assert selected.data["refresh_failure"]["attempt"]["ingestion_run_id"] == failed.pk


@pytest.mark.django_db
def test_gdp_v2_rejects_comparison_workbook_from_another_release(
    settings,
    tmp_path,
):
    settings.RAW_ARTIFACT_ROOT = tmp_path / "raw-artifacts"
    result = BEAGDPReleaseProvider(
        client=_bea_client(
            _bea_vintage_workbook(),
            _bea_comparison_workbook(quarter="2025Q4"),
        )
    ).gdp_pce()
    assert result.ok, result.error

    with pytest.raises(
        ValueError,
        match="comparison workbook does not match the latest vintage release",
    ):
        _validate_gdp_semantic_alignment(
            result.records,
            result.supplemental_records["release_vintages"],
            result.metadata,
        )

    run = record_provider_result(
        result,
        persist=_store_bea_release_observations_v2,
    )
    assert run.status == IngestionRun.Status.FAILED
    assert "comparison workbook does not match" in run.error


@pytest.mark.django_db
def test_bea_v2_acquisition_rejects_catalog_frequency_drift(
    settings,
    tmp_path,
):
    settings.RAW_ARTIFACT_ROOT = tmp_path / "raw-artifacts"
    gdp_source = ensure_source("bea-release")
    SeriesDefinition.objects.create(
        key="bea-a191rl",
        name="tampered GDP fixture",
        unit="%",
        frequency="annual",
        source=gdp_source,
    )
    gdp_result = BEAGDPReleaseProvider(
        client=_bea_client(_bea_vintage_workbook(), _bea_comparison_workbook())
    ).gdp_pce()
    gdp_run = record_provider_result(
        gdp_result,
        persist=_store_bea_release_observations_v2,
    )
    assert gdp_run.status == IngestionRun.Status.FAILED
    assert "persistence postcondition failed" in gdp_run.error
    assert not Observation.objects.filter(batch_id=gdp_run.batch_id).exists()
    assert not RawArtifact.objects.filter(run=gdp_run).exists()

    pio_source = ensure_source("bea-pio-release")
    SeriesDefinition.objects.create(
        key="bea-real-pce-mom",
        name="tampered PIO fixture",
        unit="%",
        frequency="annual",
        source=pio_source,
    )
    pio_result = BEAPIOReleaseProvider(
        client=_bea_pio_client(
            _bea_pio_summary_workbook(),
            _bea_pio_section2_workbook(),
        )
    ).personal_income_outlays()
    pio_run = record_provider_result(
        pio_result,
        persist=_store_bea_release_observations_v2,
    )
    assert pio_run.status == IngestionRun.Status.FAILED
    assert "persistence postcondition failed" in pio_run.error
    assert not Observation.objects.filter(batch_id=pio_run.batch_id).exists()
    assert not RawArtifact.objects.filter(run=pio_run).exists()


@pytest.mark.django_db
def test_bea_pio_v2_keeps_private_bundle_and_append_only_history(settings, tmp_path):
    settings.RAW_ARTIFACT_ROOT = tmp_path / "raw-artifacts"
    first_result = BEAPIOReleaseProvider(
        client=_bea_pio_client(
            _bea_pio_summary_workbook(),
            _bea_pio_section2_workbook(),
        )
    ).personal_income_outlays()
    first = record_provider_result(
        first_result,
        persist=_store_bea_release_observations_v2,
    )

    assert first.status == IngestionRun.Status.SUCCESS
    assert first.row_count == len(first_result.records)
    assert Observation.objects.filter(batch_id=first.batch_id).count() == len(
        first_result.records
    )
    first_artifact = RawArtifact.objects.get(run=first)
    first_path = (
        Path(settings.RAW_ARTIFACT_ROOT)
        / first_artifact.sha256[:2]
        / f"{first_artifact.sha256}.bin"
    )
    assert first_path.read_bytes() == first_result.raw_bytes
    first_latest = Observation.objects.get(
        batch_id=first.batch_id,
        series__key="bea-real-pce-mom",
        value_date=datetime(2026, 5, 1, tzinfo=UTC),
    )
    first_signature = (first_latest.pk, first_latest.value, first_latest.updated_at)

    second_result = BEAPIOReleaseProvider(
        client=_bea_pio_client(
            _bea_pio_summary_workbook(),
            _bea_pio_section2_workbook(),
        )
    ).personal_income_outlays()
    second = record_provider_result(
        second_result,
        persist=_store_bea_release_observations_v2,
    )

    assert second.status == IngestionRun.Status.SUCCESS
    assert second.batch_id != first.batch_id
    assert RawArtifact.objects.filter(run__in=(first, second)).count() == 2
    assert Observation.objects.filter(batch_id=second.batch_id).count() == len(
        second_result.records
    )
    first_latest.refresh_from_db()
    assert (first_latest.pk, first_latest.value, first_latest.updated_at) == first_signature


@pytest.mark.django_db
def test_bea_gdp_v2_rejects_normalized_supplemental_and_metadata_tamper(
    settings,
    tmp_path,
):
    settings.RAW_ARTIFACT_ROOT = tmp_path / "raw-artifacts"
    normalized = BEAGDPReleaseProvider(
        client=_bea_client(_bea_vintage_workbook(), _bea_comparison_workbook())
    ).gdp_pce()
    normalized.records[0]["value"] = Decimal("999")
    normalized_run = record_provider_result(
        normalized,
        persist=_store_bea_release_observations_v2,
    )
    assert normalized_run.status == IngestionRun.Status.FAILED
    assert "normalized observations" in normalized_run.error

    supplemental = BEAGDPReleaseProvider(
        client=_bea_client(_bea_vintage_workbook(), _bea_comparison_workbook())
    ).gdp_pce()
    next(
        item
        for item in supplemental.supplemental_records["release_vintages"]
        if item["estimate_round"] == "Second"
    )["value"] = Decimal("999")
    supplemental_run = record_provider_result(
        supplemental,
        persist=_store_bea_release_observations_v2,
    )
    assert supplemental_run.status == IngestionRun.Status.FAILED
    assert "supplemental observations" in supplemental_run.error

    metadata = BEAGDPReleaseProvider(
        client=_bea_client(_bea_vintage_workbook(), _bea_comparison_workbook())
    ).gdp_pce()
    metadata.metadata["source_url"] = "https://mirror.invalid/gdp"
    metadata_run = record_provider_result(
        metadata,
        persist=_store_bea_release_observations_v2,
    )
    assert metadata_run.status == IngestionRun.Status.FAILED
    assert "replay metadata" in metadata_run.error

    failed_runs = (normalized_run, supplemental_run, metadata_run)
    assert not Observation.objects.filter(batch_id__in=[run.batch_id for run in failed_runs])
    assert not ReleaseVintageObservation.objects.filter(
        batch_id__in=[run.batch_id for run in failed_runs]
    )
    assert not RawArtifact.objects.filter(run__in=failed_runs)


def test_census_release_provider_prefers_current_workbook_and_preserves_status():
    workbook = _census_current_workbook()
    result = CensusMARTSReleaseProvider(
        client=_census_client(workbook, current_status=200)
    ).monthly_retail_sales()

    assert result.ok
    assert result.metadata["workbook_url"] == CENSUS_MARTS_CURRENT_WORKBOOK
    assert result.metadata["workbook_scope"] == "current"
    assert result.metadata["latest_value_date"] == "2026-05-01"
    assert result.metadata["artifacts"][0]["sha256"] == hashlib.sha256(
        workbook
    ).hexdigest()
    evidence = parse_evidence_bundle(
        result.raw_bytes,
        expected_provider="census-release",
        expected_dataset="marts:retail-food-services",
    )
    assert set(evidence.responses) == {"current-workbook"}
    assert evidence.responses["current-workbook"] == workbook
    replay_records, replay_metadata = CensusMARTSReleaseProvider.replay_evidence_bundle(
        result.raw_bytes
    )
    assert replay_records == result.records
    assert replay_metadata["workbook_scope"] == "current"
    by_series_and_date = {
        (item["series_id"], item["date"]): item for item in result.records
    }
    latest = by_series_and_date[("CENSUS-MRTS-44X72-SM-SA", "2026-05-01")]
    assert latest["value"] == 763705
    assert latest["metadata"]["estimate_status"] == "(a)"
    assert by_series_and_date[("CENSUS-MRTS-44X72-SM-SA", "2026-04-01")][
        "value"
    ] == 757036
    assert by_series_and_date[("CENSUS-MRTS-44X72-SM-SA-MOM", "2026-05-01")][
        "value"
    ] == Decimal("0.9")
    assert by_series_and_date[("CENSUS-MRTS-44X72-SM-SA-YOY", "2026-05-01")][
        "value"
    ] == Decimal("6.9")
    assert by_series_and_date[("CENSUS-MRTS-44X72-SM-SA-MOM", "2026-04-01")][
        "value"
    ] == Decimal("0.4")


def test_census_release_replay_rechecks_xlsx_expanded_size(monkeypatch):
    result = CensusMARTSReleaseProvider(
        client=_census_client(_census_current_workbook(), current_status=200)
    ).monthly_retail_sales()
    assert result.ok
    monkeypatch.setattr(CensusMARTSReleaseProvider, "max_expanded_bytes", 1)

    with pytest.raises(ValueError, match="expanded-size limit"):
        CensusMARTSReleaseProvider.replay_evidence_bundle(result.raw_bytes)


@pytest.mark.parametrize(
    "workbook",
    [
        _census_current_workbook(
            sales_declaration="(Total sales estimates are shown in thousands of dollars.)"
        ),
        _census_current_workbook(
            change_declaration="(Estimates are shown as basis points.)"
        ),
    ],
)
def test_census_release_provider_rejects_wrong_declared_units(workbook):
    result = CensusMARTSReleaseProvider(
        client=_census_client(workbook, current_status=200)
    ).monthly_retail_sales()

    assert not result.ok
    assert "declared unit is invalid" in result.error


def test_census_release_provider_falls_back_to_latest_archive_file():
    workbook = _census_workbook()
    result = CensusMARTSReleaseProvider(client=_census_client(workbook)).monthly_retail_sales()

    assert result.ok
    assert result.metadata["workbook_url"].endswith("rs2604.xlsx")
    assert result.metadata["workbook_scope"] == "historical_archive"
    assert result.metadata["latest_value_date"] == "2026-04-01"
    assert result.metadata["artifacts"][1]["sha256"] == hashlib.sha256(workbook).hexdigest()
    evidence = parse_evidence_bundle(
        result.raw_bytes,
        expected_provider="census-release",
        expected_dataset="marts:retail-food-services",
    )
    assert set(evidence.responses) == {"archive-index", "archive-workbook"}
    assert evidence.responses["archive-workbook"] == workbook
    assert b"rs2604.xlsx" in evidence.responses["archive-index"]
    replay_records, replay_metadata = CensusMARTSReleaseProvider.replay_evidence_bundle(
        result.raw_bytes
    )
    assert replay_records == result.records
    assert replay_metadata["workbook_scope"] == "historical_archive"
    by_series_and_date = {
        (item["series_id"], item["date"]): item for item in result.records
    }
    latest = by_series_and_date[("CENSUS-MRTS-44X72-SM-SA", "2026-04-01")]
    assert latest["value"] == 757085
    assert latest["metadata"]["estimate_status"] == "(a)"
    assert by_series_and_date[("CENSUS-MRTS-44X72-SM-SA-MOM", "2026-04-01")][
        "value"
    ] == Decimal("0.5")
    assert by_series_and_date[("CENSUS-MRTS-44X72-SM-SA-YOY", "2026-03-01")][
        "value"
    ] == Decimal("4.2")
    assert by_series_and_date[("CENSUS-MRTS-44X72-SM-SA-MOM", "2026-04-01")][
        "metadata"
    ]["estimate_status"] == "(a)"


@pytest.mark.django_db
def test_census_release_v2_keeps_private_append_only_workbook_evidence(
    settings,
    tmp_path,
):
    settings.RAW_ARTIFACT_ROOT = tmp_path / "raw-artifacts"
    first_result = CensusMARTSReleaseProvider(
        client=_census_client(_census_current_workbook(), current_status=200)
    ).monthly_retail_sales()
    first = record_provider_result(
        first_result,
        persist=_store_census_marts_observations_v2,
    )
    assert first.status == IngestionRun.Status.SUCCESS
    assert Observation.objects.filter(batch_id=first.batch_id).count() == len(
        first_result.records
    )
    artifact = RawArtifact.objects.get(run=first)
    artifact_path = (
        Path(settings.RAW_ARTIFACT_ROOT)
        / artifact.sha256[:2]
        / f"{artifact.sha256}.bin"
    )
    assert artifact_path.read_bytes() == first_result.raw_bytes
    first_rows = list(
        Observation.objects.filter(batch_id=first.batch_id)
        .order_by("pk")
        .values_list("pk", "value", "updated_at")
    )

    second_result = CensusMARTSReleaseProvider(
        client=_census_client(_census_current_workbook(), current_status=200)
    ).monthly_retail_sales()
    second = record_provider_result(
        second_result,
        persist=_store_census_marts_observations_v2,
    )
    assert second.status == IngestionRun.Status.SUCCESS
    assert second.batch_id != first.batch_id
    assert RawArtifact.objects.filter(run__in=(first, second)).count() == 2
    assert Observation.objects.filter(batch_id=second.batch_id).count() == len(
        second_result.records
    )
    assert list(
        Observation.objects.filter(batch_id=first.batch_id)
        .order_by("pk")
        .values_list("pk", "value", "updated_at")
    ) == first_rows


@pytest.mark.django_db
def test_consumer_dashboard_refuses_partial_metric_set():
    census = CensusMARTSReleaseProvider(
        client=_census_client(_census_current_workbook(), current_status=200)
    ).monthly_retail_sales()
    record_provider_result(census, persist=_store_release_workbook_observations)

    assert publish_official_dashboards(keys={"consumer"}) == []


@pytest.mark.django_db
def test_release_workbooks_persist_lineage_and_publish_gdp_and_consumer_pages(
    client,
    settings,
    tmp_path,
):
    settings.RAW_ARTIFACT_ROOT = tmp_path / "raw-artifacts"
    bea = BEAGDPReleaseProvider(
        client=_bea_client(_bea_vintage_workbook(), _bea_comparison_workbook())
    ).gdp_pce()
    census_api = _census_api_result()
    census = CensusMARTSReleaseProvider(
        client=_census_client(_census_current_workbook(), current_status=200)
    ).monthly_retail_sales()
    pio = BEAPIOReleaseProvider(
        client=_bea_pio_client(
            _bea_pio_summary_workbook(),
            _bea_pio_section2_workbook(),
        )
    ).personal_income_outlays()
    bea_run = record_provider_result(
        bea,
        persist=_store_bea_release_observations_v2,
    )
    census_api_run = record_provider_result(
        census_api, persist=_store_release_workbook_observations
    )
    census_run = record_provider_result(census, persist=_store_release_workbook_observations)
    pio_run = record_provider_result(pio, persist=_store_release_workbook_observations)
    g19, household = _consumer_credit_results()
    g19_run = record_provider_result(g19, persist=_store_release_workbook_observations)
    household_run = record_provider_result(
        household, persist=_store_release_workbook_observations
    )
    _record_census_revision_witness(
        [census_api_run, census_run, pio_run, g19_run, household_run]
    )
    census_api_run.refresh_from_db()

    dashboards = {
        item.key: item for item in publish_official_dashboards(keys={"consumer"})
    }
    gdp_snapshot = publish_gdp_revision(run=bea_run)
    assert gdp_snapshot is not None
    dashboards["gdp"] = gdp_snapshot

    assert bea_run.status == "success"
    assert census_api_run.status == "success"
    assert census_run.status == "success"
    assert pio_run.status == "success"
    assert g19_run.status == "success"
    assert household_run.status == "success"
    witness = census_api_run.metadata["legacy_revision_witness"]
    assert witness["latest_value_date"] == "2026-05-01"
    assert witness["overlap_count"] == 7
    assert witness["differences"] == []
    assert RawArtifact.objects.filter(run=bea_run).count() == 1
    assert RawArtifact.objects.filter(run=census_run).count() == 1
    assert RawArtifact.objects.filter(run=pio_run).count() == 3
    assert ReleaseVintageObservation.objects.filter(batch_id=bea_run.batch_id).count() == 12
    second_estimate = ReleaseVintageObservation.objects.get(
        batch_id=bea_run.batch_id,
        series__key="bea-a191rl",
        value_date=datetime(2026, 1, 1, tzinfo=UTC),
        release_date="2026-05-28",
    )
    assert second_estimate.value == Decimal("1.6")
    assert second_estimate.estimate_round == "Second"
    assert second_estimate.as_of.date().isoformat() == "2026-05-28"
    assert second_estimate.fetched_at == bea.fetched_at
    assert second_estimate.source.key == "bea-release"
    assert second_estimate.license_scope == second_estimate.source.license_scope
    assert second_estimate.fallback_source is None
    assert second_estimate.quality_status == Observation.Quality.FRESH
    assert set(dashboards) == {"gdp", "consumer"}
    gdp = {item["key"]: item for item in dashboards["gdp"].data["metrics"]}
    consumer = {
        item["key"]: item for item in dashboards["consumer"].data["metrics"]
    }
    assert gdp["bea-a191rl"]["display_value"] == "2.10%"
    assert gdp["bea-dpcerl"]["display_value"] == "0.50%"
    assert gdp["bea-pce-contribution"]["display_value"] == "0.37pp"
    gdp_charts = dashboards["gdp"].data["charts"]
    assert [chart["key"] for chart in gdp_charts] == [
        "gdp-growth-history",
        "gdp-vintage-trail",
    ]
    assert [row["实际 GDP"] for row in gdp_charts[1]["data"]] == [1.6, 2.1]
    assert [row["date"] for row in gdp_charts[1]["data"]] == [
        "Second\n05-28",
        "Third\n06-25",
    ]
    assert gdp_charts[1]["panel_class"] == "lg:col-span-2"
    assert gdp_charts[1]["data"][0]["_lineage"]["实际 GDP"][
        "estimate_round"
    ] == "Second"
    revision_section = dashboards["gdp"].data["sections"][0]
    assert revision_section["title"] == "GDP 发布轮次与修订路径"
    assert revision_section["rows"][0]["display_value"] == "1.60% → 2.10%"
    assert "累计修订 +0.50pp" in revision_section["rows"][0]["status"]
    assert consumer["census-mrts-44x72-sm-sa"]["display_value"] == "763,705 USD mn"
    assert consumer["census-mrts-44x72-sm-sa-mom"]["display_value"] == "0.90%"
    assert consumer["census-mrts-44x72-sm-sa-mom"]["change_unit"] == "pp"
    assert consumer["bea-real-pce-mom"]["display_value"] == "0.30%"
    assert consumer["bea-real-dpi-mom"]["display_value"] == "0.30%"
    assert consumer["bea-personal-saving-rate"]["display_value"] == "3.00%"
    assert consumer["census-mrts-44x72-sm-sa"]["source_key"] == "census-release"
    assert consumer["bea-real-pce-mom"]["source_key"] == "bea-pio-release"
    assert consumer["g19-consumer-credit-outstanding-sa"]["source_key"] == (
        "federal-reserve-g19"
    )
    assert consumer["hhdc-total-debt-balance"]["source_key"] == (
        "ny-fed-household-credit"
    )
    charts = dashboards["consumer"].data["charts"]
    assert [chart["key"] for chart in charts] == [
        "retail-sales",
        "real-consumption-income-momentum",
        "personal-saving-rate",
        "consumer-credit-composition",
        "household-debt-composition",
        "household-debt-delinquency",
    ]
    assert dashboards["consumer"].data["chart_data"] == charts[0]["data"]
    assert dashboards["consumer"].data["contract_version"] == 1
    assert dashboards["consumer"].data["retail_batch_id"] == str(
        census_run.batch_id
    )
    assert charts[0]["source_keys"] == ["census-release"]
    assert charts[1]["source_keys"] == ["bea-pio-release"]
    assert charts[0]["data"][0]["_lineage"]["零售与餐饮服务"][
        "source_key"
    ] == "census-release"
    response = client.get("/economy/consumer/")
    body = response.content.decode()
    assert response.status_code == 200
    assert body.count(" data-chart ") == 6
    assert "dashboard-chart-0" in body
    assert "dashboard-chart-1" in body
    assert "dashboard-chart-2" in body
    assert "dashboard-chart-5" in body
    assert "实际 PCE 环比" in body
    assert "3.00%" in body
    assert "U.S. Bureau of Economic Analysis Personal Income and Outlays Releases" in body
    assert "New York Fed Household Debt and Credit" in body
    assert "来源：Atlas Macro Derived Data" not in body
    gdp_response = client.get("/economy/gdp/")
    gdp_body = gdp_response.content.decode()
    assert gdp_response.status_code == 200
    assert gdp_body.count(" data-chart ") == 2
    assert "GDP 发布轮次与修订路径" in gdp_body
    assert "1.60% → 2.10%" in gdp_body

    runs = [
        bea_run,
        census_api_run,
        census_run,
        pio_run,
        g19_run,
        household_run,
    ]
    assert _keys_with_current_required_batches({"gdp", "consumer"}, runs) == {
        "gdp",
        "consumer",
    }
    census_api_run.dataset = "mrts:44X72:SM:yes"
    census_api_run.save(update_fields=["dataset", "updated_at"])
    assert _keys_with_current_required_batches({"gdp", "consumer"}, runs) == {
        "gdp",
        "consumer",
    }
    census_api_run.dataset = "marts:44X72:SM:yes"
    census_api_run.save(update_fields=["dataset", "updated_at"])
    census_run.dataset = "marts:44X72:SM:yes"
    census_run.save(update_fields=["dataset", "updated_at"])
    assert _keys_with_current_required_batches({"gdp", "consumer"}, runs) == {"gdp"}
    census_run.dataset = "marts:retail-food-services"
    census_run.save(update_fields=["dataset", "updated_at"])
    latest_pio = Observation.objects.filter(
        series__key="bea-real-pce-mom",
        source__key="bea-pio-release",
    ).latest("value_date")
    latest_pio.batch_id = uuid.uuid4()
    latest_pio.save(update_fields=["batch_id", "updated_at"])
    assert _keys_with_current_required_batches({"gdp", "consumer"}, runs) == {
        "gdp"
    }
    _mark_latest_dashboards_stale({"consumer"}, runs)
    stale = DashboardSnapshot.objects.filter(key="consumer").latest("created_at")
    assert stale.quality_status == "stale"
    assert "上一版完整快照" in stale.data["refresh_failure"]["reason"]

    latest_pio.batch_id = pio_run.batch_id
    latest_pio.save(update_fields=["batch_id", "updated_at"])
    latest_g19 = Observation.objects.filter(
        series__key="g19-consumer-credit-outstanding-sa",
        source__key="federal-reserve-g19",
    ).latest("value_date")
    latest_g19.batch_id = uuid.uuid4()
    latest_g19.save(update_fields=["batch_id", "updated_at"])
    assert _keys_with_current_required_batches({"gdp", "consumer"}, runs) == {
        "gdp"
    }
    latest_g19.batch_id = g19_run.batch_id
    latest_g19.save(update_fields=["batch_id", "updated_at"])
    assert publish_official_dashboards(keys={"consumer"}) == []
    recovered = DashboardSnapshot.objects.get(pk=stale.pk)
    assert "refresh_failure" not in recovered.data

    census_api_run.status = IngestionRun.Status.PARTIAL
    census_api_run.row_count = 0
    census_api_run.metadata = {"reason": "CENSUS_API_KEY is not configured"}
    census_api_run.save(
        update_fields=["status", "row_count", "metadata", "updated_at"]
    )
    assert _publishable_keys_for_source_groups(runs, MACRO_PUBLICATION_GROUPS) == {
        "gdp",
        "consumer",
    }


@pytest.mark.django_db
def test_gdp_generic_publisher_rejects_legacy_cross_source_rows():
    older_fetch = datetime(2026, 6, 25, tzinfo=UTC)
    newer_fetch = older_fetch + timedelta(days=1)
    record_provider_result(
        ProviderResult(
            provider="bea",
            dataset="api-fixture",
            fetched_at=older_fetch,
            records=[
                {"series_id": "BEA-A191RL", "date": "2026-01-01", "value": "9.9"},
                {"series_id": "BEA-A191RL", "date": "2025-10-01", "value": "1.0"},
            ],
        ),
        persist=store_series_observations,
    )
    record_provider_result(
        ProviderResult(
            provider="bea-release",
            dataset="release-fixture",
            fetched_at=newer_fetch,
            records=[
                {"series_id": "BEA-A191RL", "date": "2026-01-01", "value": "2.1"},
                {"series_id": "BEA-A191RL", "date": "2025-10-01", "value": "0.5"},
            ],
        ),
        persist=store_series_observations,
    )

    assert publish_official_dashboards(keys={"gdp"}) == []
    assert not DashboardSnapshot.objects.filter(key="gdp").exists()


@pytest.mark.django_db
def test_release_persistence_rejects_regressed_latest_month():
    current = ProviderResult(
        provider="census-release",
        dataset="regression-guard-current",
        records=[
            {
                "series_id": "CENSUS-MRTS-44X72-SM-SA",
                "date": "2026-05-01",
                "value": "100",
            }
        ],
    )
    current_run = record_provider_result(
        current, persist=_store_release_workbook_observations
    )
    regressed = ProviderResult(
        provider="census-release",
        dataset="regression-guard-old",
        records=[
            {
                "series_id": "CENSUS-MRTS-44X72-SM-SA",
                "date": "2026-04-01",
                "value": "90",
            }
        ],
    )
    regressed_run = record_provider_result(
        regressed, persist=_store_release_workbook_observations
    )

    assert current_run.status == "success"
    assert regressed_run.status == "failed"
    assert "latest value date regressed" in regressed_run.error
    assert Observation.objects.filter(source__key="census-release").count() == 1


@pytest.mark.django_db
def test_gdp_publication_gate_requires_vintages_from_the_same_release_batch():
    result = BEAGDPReleaseProvider(
        client=_bea_client(_bea_vintage_workbook(), _bea_comparison_workbook())
    ).gdp_pce()
    run = record_provider_result(result, persist=_store_release_workbook_observations)

    assert _keys_with_current_required_batches({"gdp"}, [run]) == {"gdp"}
    ReleaseVintageObservation.objects.filter(batch_id=run.batch_id).delete()
    assert _keys_with_current_required_batches({"gdp"}, [run]) == set()


@pytest.mark.django_db
def test_gdp_vintage_persistence_is_idempotent_and_rebinds_the_refresh_batch():
    first_result = BEAGDPReleaseProvider(
        client=_bea_client(_bea_vintage_workbook(), _bea_comparison_workbook())
    ).gdp_pce()
    first_run = record_provider_result(
        first_result,
        persist=_store_release_workbook_observations,
    )
    second_result = BEAGDPReleaseProvider(
        client=_bea_client(_bea_vintage_workbook(), _bea_comparison_workbook())
    ).gdp_pce()
    second_run = record_provider_result(
        second_result,
        persist=_store_release_workbook_observations,
    )

    assert first_run.status == IngestionRun.Status.SUCCESS
    assert second_run.status == IngestionRun.Status.SUCCESS
    assert ReleaseVintageObservation.objects.count() == 12
    assert set(
        ReleaseVintageObservation.objects.values_list("batch_id", flat=True)
    ) == {second_run.batch_id}


def test_economy_catalog_separates_live_release_data_from_remaining_gaps():
    requirements = {item["key"]: item for item in DATA_REQUIREMENTS}

    assert requirements["bea-gdp-pce"]["status"] == "live"
    assert requirements["bea-gdp-contributions"]["status"] == "live"
    assert requirements["bea-gdp-vintage-trail"]["status"] == "live"
    assert "独立 release-vintage 数据层" in requirements["bea-gdp-vintage-trail"][
        "reason"
    ]
    assert requirements["census-retail"]["status"] == "live"
    assert "发布工作簿" in requirements["census-retail"]["reason"]
    assert requirements["bea-personal-income-outlays"]["status"] == "live"
    assert "Section 2" in requirements["bea-personal-income-outlays"]["reason"]
    assert requirements["bea-pio-vintage-trail"]["status"] == "needs_source"
    assert requirements["census-retail-history"]["status"] == "needs_source"
    assert requirements["consumer-credit-official"]["status"] == "live"
    assert requirements["consumer-credit-vintage-trail"]["status"] == "needs_source"
    assert requirements["consumer-confidence"]["status"] == "purchase_required"


@pytest.mark.parametrize(
    ("statuses", "expected"),
    [
        (
            ("success", "success", "success", "success", "success", "success"),
            {"gdp", "consumer"},
        ),
        (
            ("success", "failed", "success", "success", "success", "success"),
            {"gdp", "consumer"},
        ),
        (("success", "success", "failed", "success", "success", "success"), {"gdp"}),
        (("success", "success", "success", "failed", "success", "success"), {"gdp"}),
        (("failed", "success", "success", "success", "success", "success"), {"consumer"}),
        (("success", "success", "success", "success", "partial", "success"), {"gdp"}),
        (("success", "success", "success", "success", "success", "failed"), {"gdp"}),
    ],
)
def test_macro_publication_groups_isolate_unrelated_page_failures(statuses, expected):
    runs = [
        SimpleNamespace(
            source=SimpleNamespace(key=source_key),
            status=status,
            row_count=1 if status == "success" else 0,
        )
        for source_key, status in zip(
            (
                "bea-release",
                "census",
                "census-release",
                "bea-pio-release",
                "federal-reserve-g19",
                "ny-fed-household-credit",
            ),
            statuses,
            strict=True,
        )
    ]

    assert (
        _publishable_keys_for_source_groups(runs, MACRO_PUBLICATION_GROUPS)
        == expected
    )
