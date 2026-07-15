"""Keyless official BEA and Census release-workbook adapters.

Both agencies publish first-party XLSX workbooks alongside their releases.  We
use those files when free API credentials are unavailable, retain workbook
hashes, and keep estimate/revision labels in observation metadata.  Nothing in
this module treats retrieval time as the economic-data vintage.
"""

from __future__ import annotations

import hashlib
import re
import zipfile
from datetime import datetime
from decimal import Decimal, InvalidOperation
from html.parser import HTMLParser
from io import BytesIO
from typing import Any
from urllib.parse import urljoin, urlparse

from openpyxl import load_workbook

from .providers import HTTPProvider, ProviderResult
from .raw_evidence import (
    EvidenceResponse,
    build_evidence_bundle,
    parse_evidence_bundle,
)

BEA_GDP_PAGE = "https://www.bea.gov/data/gdp/gross-domestic-product"
BEA_PIO_PAGE = "https://www.bea.gov/data/income-saving/personal-income"
BEA_PIO_SECTION2_WORKBOOK = (
    "https://apps.bea.gov/national/Release/XLS/Survey/Section2All_xls.xlsx"
)
BEA_VINTAGE_WORKBOOK = (
    "https://apps.bea.gov/national/xls/gdp-gdi-vintage-history.xlsx"
)
CENSUS_MARTS_INDEX = "https://www2.census.gov/retail/releases/historical/marts/"
CENSUS_MARTS_RELEASE_PAGE = "https://www.census.gov/retail/sales.html"
CENSUS_MARTS_CURRENT_WORKBOOK = (
    "https://www.census.gov/retail/marts/www/marts_current.xlsx"
)
XLSX_CONTENT_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


class _LinkParser(HTMLParser):
    def __init__(self) -> None:
        super().__init__(convert_charrefs=True)
        self.links: list[tuple[str, str]] = []
        self._href: str | None = None
        self._text: list[str] = []

    def handle_starttag(self, tag: str, attrs: list[tuple[str, str | None]]) -> None:
        if tag.lower() == "a":
            self._href = dict(attrs).get("href")
            self._text = []

    def handle_data(self, data: str) -> None:
        if self._href is not None:
            self._text.append(data)

    def handle_endtag(self, tag: str) -> None:
        if tag.lower() == "a" and self._href is not None:
            self.links.append((self._href, " ".join("".join(self._text).split())))
            self._href = None
            self._text = []


def _decimal(value: Any) -> Decimal | None:
    if value is None:
        return None
    text = str(value).strip().replace(",", "")
    if not text or set(text) <= {"."} or text.upper() in {"NA", "N/A", "(NA)", "(*)"}:
        return None
    try:
        parsed = Decimal(text)
    except (InvalidOperation, ValueError):
        return None
    return parsed if parsed.is_finite() else None


def _quarter_start(period: str) -> str | None:
    match = re.fullmatch(r"(\d{4})Q([1-4])", str(period).strip().upper())
    if not match:
        return None
    year, quarter = (int(value) for value in match.groups())
    return f"{year:04d}-{(quarter - 1) * 3 + 1:02d}-01"


def _month_start(value: Any) -> str | None:
    if isinstance(value, datetime):
        return value.date().replace(day=1).isoformat()
    text = str(value or "").strip()
    if match := re.fullmatch(r"(\d{4})M(\d{1,2})", text, re.IGNORECASE):
        year, month = (int(part) for part in match.groups())
    elif match := re.search(r"([A-Z][a-z]{2,8})\s+(\d{4})", text):
        try:
            month = datetime.strptime(match.group(1), "%B").month
        except ValueError:
            try:
                month = datetime.strptime(match.group(1), "%b").month
            except ValueError:
                return None
        year = int(match.group(2))
    else:
        return None
    if not 1 <= month <= 12:
        return None
    return f"{year:04d}-{month:02d}-01"


def _previous_month(value_date: str) -> str:
    year, month, _ = (int(part) for part in value_date.split("-"))
    if month == 1:
        return f"{year - 1:04d}-12-01"
    return f"{year:04d}-{month - 1:02d}-01"


def _previous_quarter(period: str) -> str:
    year, quarter = (int(value) for value in re.fullmatch(r"(\d{4})Q([1-4])", period).groups())
    if quarter == 1:
        return f"{year - 1}Q4"
    return f"{year}Q{quarter - 1}"


def _release_date(value: Any) -> str | None:
    if isinstance(value, datetime):
        return value.date().isoformat()
    text = str(value or "").strip()
    match = re.search(r"[A-Z][a-z]{2,8}\s+\d{1,2},\s+\d{4}", text)
    if not match:
        return None
    for date_format in ("%b %d, %Y", "%B %d, %Y"):
        try:
            return datetime.strptime(match.group(0), date_format).date().isoformat()
        except ValueError:
            continue
    return None


def _estimate_round(value: Any) -> str | None:
    match = re.search(
        r"\((Advance|Second|Third|Revised)\s+Estimate\)",
        str(value or ""),
        re.IGNORECASE,
    )
    return match.group(1).title() if match else None


def _xlsx_bytes(content: bytes, *, max_expanded_bytes: int) -> bytes:
    if not content.startswith(b"PK"):
        raise ValueError("response is not an XLSX ZIP archive")
    with zipfile.ZipFile(BytesIO(content)) as archive:
        total = sum(item.file_size for item in archive.infolist())
        if total > max_expanded_bytes:
            raise ValueError("XLSX exceeded configured expanded-size limit")
        if "[Content_Types].xml" not in archive.namelist():
            raise ValueError("XLSX content-types manifest is missing")
    return content


def _artifact(url: str, content: bytes, content_type: str) -> dict[str, Any]:
    return {
        "url": url,
        "sha256": hashlib.sha256(content).hexdigest(),
        "size": len(content),
        "content_type": content_type,
    }


class _ReleaseWorkbookProvider(HTTPProvider):
    max_html_bytes = 5 * 1024 * 1024
    max_workbook_bytes = 4 * 1024 * 1024
    max_expanded_bytes = 48 * 1024 * 1024
    allowed_hosts: frozenset[str] = frozenset()

    def _download(self, url: str, *, expected: str) -> tuple[bytes, str, str]:
        limit = self.max_html_bytes if expected == "html" else self.max_workbook_bytes
        chunks: list[bytes] = []
        size = 0
        with self.client.stream("GET", url) as response:
            response.raise_for_status()
            final_url = urlparse(str(response.url))
            if (
                final_url.scheme != "https"
                or not final_url.hostname
                or final_url.hostname.lower() not in self.allowed_hosts
            ):
                raise ValueError("download redirected outside the official host allowlist")
            for chunk in response.iter_bytes():
                size += len(chunk)
                if size > limit:
                    raise ValueError(f"{expected} response exceeded configured size limit")
                chunks.append(chunk)
            content_type = response.headers.get("content-type", "").split(";", 1)[0].lower()
            last_modified = response.headers.get("last-modified", "")
        content = b"".join(chunks)
        if expected == "html":
            if content_type and "html" not in content_type:
                raise ValueError(f"unexpected HTML content type: {content_type}")
        else:
            _xlsx_bytes(content, max_expanded_bytes=self.max_expanded_bytes)
            if content_type and content_type not in {
                XLSX_CONTENT_TYPE,
                "application/octet-stream",
                "application/zip",
            }:
                raise ValueError(f"unexpected XLSX content type: {content_type}")
        return (
            content,
            content_type or ("text/html" if expected == "html" else XLSX_CONTENT_TYPE),
            last_modified,
        )


class BEAGDPReleaseProvider(_ReleaseWorkbookProvider):
    """Parse BEA's current GDP release and official vintage-history workbook."""

    key = "bea-release"
    base_url = "https://www.bea.gov"
    allowed_hosts = frozenset({"bea.gov", "www.bea.gov", "apps.bea.gov"})

    GROWTH_SERIES = {
        "Personal consumption expenditures": "BEA-DPCERL",
        "Gross private domestic investment": "BEA-GPDI-GROWTH",
        "Fixed investment": "BEA-FIXED-INVESTMENT-GROWTH",
        "Exports": "BEA-EXPORTS-GROWTH",
        "Imports": "BEA-IMPORTS-GROWTH",
        "Government consumption expenditures and gross investment": "BEA-GOVERNMENT-GROWTH",
    }
    CONTRIBUTION_SERIES = {
        "Personal consumption expenditures": "BEA-PCE-CONTRIBUTION",
        "Gross private domestic investment": "BEA-GPDI-CONTRIBUTION",
        "Net exports of goods and services": "BEA-NET-EXPORTS-CONTRIBUTION",
        "Government consumption expenditures and gross investment": "BEA-GOVERNMENT-CONTRIBUTION",
    }

    def gdp_pce(self) -> ProviderResult:
        dataset = "gdp-release-workbooks"
        try:
            page, page_type, page_modified = self._download(
                BEA_GDP_PAGE,
                expected="html",
            )
            comparison_url = self._comparison_url(page)
            vintage, vintage_type, vintage_modified = self._download(
                BEA_VINTAGE_WORKBOOK, expected="xlsx"
            )
            comparison, comparison_type, comparison_modified = self._download(
                comparison_url, expected="xlsx"
            )
            records, vintage_records, release_metadata = self._parse_vintage(vintage)
            component_records, component_metadata = self._parse_components(comparison)
            records.extend(component_records)
            raw_bundle, raw_bundle_metadata = build_evidence_bundle(
                provider=self.key,
                dataset=dataset,
                responses=(
                    EvidenceResponse(
                        role="release-page",
                        url=BEA_GDP_PAGE,
                        content_type=page_type,
                        raw_bytes=page,
                        response_witness={"last_modified": page_modified},
                    ),
                    EvidenceResponse(
                        role="vintage-workbook",
                        url=BEA_VINTAGE_WORKBOOK,
                        content_type=vintage_type,
                        raw_bytes=vintage,
                        response_witness={"last_modified": vintage_modified},
                    ),
                    EvidenceResponse(
                        role="comparison-workbook",
                        url=comparison_url,
                        content_type=comparison_type,
                        raw_bytes=comparison,
                        request_witness={"discovered_from": "release-page"},
                        response_witness={"last_modified": comparison_modified},
                    ),
                ),
            )
        except Exception as exc:
            return ProviderResult.failure(self.key, dataset, f"{type(exc).__name__}: {exc}")

        artifacts = [
            _artifact(BEA_GDP_PAGE, page, page_type),
            _artifact(BEA_VINTAGE_WORKBOOK, vintage, vintage_type),
            _artifact(comparison_url, comparison, comparison_type),
        ]
        return ProviderResult(
            provider=self.key,
            dataset=dataset,
            records=records,
            supplemental_records={"release_vintages": vintage_records},
            raw_bytes=raw_bundle,
            metadata={
                **raw_bundle_metadata,
                "source_url": BEA_GDP_PAGE,
                "vintage_workbook_url": BEA_VINTAGE_WORKBOOK,
                "comparison_workbook_url": comparison_url,
                "vintage_last_modified": vintage_modified,
                "comparison_last_modified": comparison_modified,
                "artifacts": artifacts,
                "vintage_policy": (
                    "latest published vintage for normalized observations; every valid "
                    "release vintage is retained in the independent vintage store"
                ),
                "unit_policy": "source workbook values retained",
                "attribution": "U.S. Bureau of Economic Analysis",
                **release_metadata,
                **component_metadata,
            },
        )

    @classmethod
    def replay_evidence_bundle(
        cls,
        raw_bytes: bytes,
    ) -> tuple[list[dict[str, Any]], dict[str, list[dict[str, Any]]], dict[str, Any]]:
        evidence = parse_evidence_bundle(
            raw_bytes,
            expected_provider=cls.key,
            expected_dataset="gdp-release-workbooks",
        )
        if set(evidence.responses) != {
            "release-page",
            "vintage-workbook",
            "comparison-workbook",
        }:
            raise ValueError("BEA GDP evidence roles are incomplete")
        urls = {
            item["role"]: item["url"]
            for item in evidence.manifest["responses"]
        }
        page = evidence.responses["release-page"]
        comparison_url = cls._comparison_url(page)
        if (
            urls["release-page"] != BEA_GDP_PAGE
            or urls["vintage-workbook"] != BEA_VINTAGE_WORKBOOK
            or urls["comparison-workbook"] != comparison_url
        ):
            raise ValueError("BEA GDP evidence URLs do not match official discovery")
        vintage_bytes = _xlsx_bytes(
            evidence.responses["vintage-workbook"],
            max_expanded_bytes=cls.max_expanded_bytes,
        )
        comparison_bytes = _xlsx_bytes(
            evidence.responses["comparison-workbook"],
            max_expanded_bytes=cls.max_expanded_bytes,
        )
        records, vintage_records, release_metadata = cls._parse_vintage(
            vintage_bytes
        )
        validator = cls()
        try:
            component_records, component_metadata = validator._parse_components(
                comparison_bytes
            )
        finally:
            validator.close()
        records.extend(component_records)
        return (
            records,
            {"release_vintages": vintage_records},
            {
                "source_url": BEA_GDP_PAGE,
                "vintage_workbook_url": BEA_VINTAGE_WORKBOOK,
                "comparison_workbook_url": comparison_url,
                **release_metadata,
                **component_metadata,
            },
        )

    @staticmethod
    def _comparison_url(page: bytes) -> str:
        parser = _LinkParser()
        parser.feed(page.decode("utf-8", errors="replace"))
        candidates = []
        for href, label in parser.links:
            absolute = urljoin(BEA_GDP_PAGE, href)
            parsed = urlparse(absolute)
            if (
                parsed.scheme == "https"
                and parsed.hostname in {"bea.gov", "www.bea.gov"}
                and parsed.path.lower().endswith(".xlsx")
                and ("Historical Comparisons" in label or re.search(r"/hist[^/]+\.xlsx$", parsed.path, re.I))
            ):
                candidates.append(absolute)
        if len(set(candidates)) != 1:
            raise ValueError("BEA page must expose exactly one historical-comparisons workbook")
        return candidates[0]

    @staticmethod
    def _parse_vintage(
        content: bytes,
    ) -> tuple[list[dict[str, Any]], list[dict[str, Any]], dict[str, Any]]:
        workbook = load_workbook(BytesIO(content), read_only=True, data_only=True)
        try:
            if "Vintage History" not in workbook.sheetnames:
                raise ValueError("BEA vintage workbook is missing the expected sheet")
            sheet = workbook["Vintage History"]
            updated = ""
            vintage_records: list[dict[str, Any]] = []
            current_period: str | None = None
            for row in sheet.iter_rows(values_only=True):
                values = list(row) + [None] * 7
                if (
                    not updated
                    and isinstance(values[0], str)
                    and values[0].startswith("Last Updated")
                ):
                    updated = values[0].removeprefix("Last Updated").strip()
                if isinstance(values[0], str) and _quarter_start(values[0]):
                    current_period = values[0].strip().upper()
                    continue
                if current_period is None:
                    continue
                nominal_gdp = _decimal(values[2])
                real_gdp = _decimal(values[4])
                release_date = _release_date(values[6])
                if (
                    not values[1]
                    or nominal_gdp is None
                    or real_gdp is None
                    or release_date is None
                ):
                    continue
                value_date = _quarter_start(current_period)
                vintage_label = str(values[1]).strip()
                metadata = {
                    "estimate_quarter": current_period,
                    "vintage_label": vintage_label,
                    "estimate_round": vintage_label,
                    "source_revision_date": release_date,
                    "source_revision_text": str(values[6]).strip(),
                    "workbook_last_updated": updated,
                    "seasonal_adjustment": "SAAR",
                }
                period_records = [
                    {
                        "series_id": "BEA-A191RL",
                        "date": value_date,
                        "value": real_gdp,
                        "release_date": release_date,
                        "estimate_round": vintage_label,
                        "vintage_label": vintage_label,
                        "metadata": {
                            **metadata,
                            "unit": "percent change from preceding period",
                        },
                    },
                    {
                        "series_id": "BEA-GDP-NOMINAL-SAAR",
                        "date": value_date,
                        "value": nominal_gdp,
                        "release_date": release_date,
                        "estimate_round": vintage_label,
                        "vintage_label": vintage_label,
                        "metadata": {
                            **metadata,
                            "unit": "USD billions, current dollars",
                        },
                    },
                ]
                nominal_gdi = _decimal(values[3])
                real_gdi = _decimal(values[5])
                if nominal_gdi is not None:
                    period_records.append(
                        {
                            "series_id": "BEA-GDI-NOMINAL-SAAR",
                            "date": value_date,
                            "value": nominal_gdi,
                            "release_date": release_date,
                            "estimate_round": vintage_label,
                            "vintage_label": vintage_label,
                            "metadata": {
                                **metadata,
                                "unit": "USD billions, current dollars",
                            },
                        }
                    )
                if real_gdi is not None:
                    period_records.append(
                        {
                            "series_id": "BEA-GDI-REAL-GROWTH-SAAR",
                            "date": value_date,
                            "value": real_gdi,
                            "release_date": release_date,
                            "estimate_round": vintage_label,
                            "vintage_label": vintage_label,
                            "metadata": {
                                **metadata,
                                "unit": "percent change from preceding period",
                            },
                        }
                    )
                vintage_records.extend(period_records)
            identities = [
                (
                    item["series_id"],
                    item["date"],
                    item["release_date"],
                    item["estimate_round"],
                )
                for item in vintage_records
            ]
            if len(identities) != len(set(identities)):
                raise ValueError("BEA vintage workbook duplicated a release identity")
            periods = {item["date"] for item in vintage_records}
            current_records: list[dict[str, Any]] = []
            for period in sorted(periods):
                period_records = [
                    item for item in vintage_records if item["date"] == period
                ]
                latest_release = max(item["release_date"] for item in period_records)
                latest_rounds = {
                    item["estimate_round"]
                    for item in period_records
                    if item["release_date"] == latest_release
                }
                if len(latest_rounds) != 1:
                    raise ValueError(
                        "BEA vintage workbook has ambiguous latest estimate rounds"
                    )
                current_records.extend(
                    item
                    for item in period_records
                    if item["release_date"] == latest_release
                )
            if not current_records or not vintage_records:
                raise ValueError("BEA vintage workbook yielded no GDP observations")
            return current_records, vintage_records, {
                "workbook_last_updated": updated,
                "quarter_count": len(periods),
                "vintage_release_count": len(
                    {
                        (
                            item["date"],
                            item["release_date"],
                            item["estimate_round"],
                        )
                        for item in vintage_records
                    }
                ),
                "vintage_observation_count": len(vintage_records),
            }
        finally:
            workbook.close()

    def _parse_components(
        self, content: bytes
    ) -> tuple[list[dict[str, Any]], dict[str, Any]]:
        workbook = load_workbook(BytesIO(content), read_only=True, data_only=True)
        if "GDPhistQ" not in workbook.sheetnames:
            raise ValueError("BEA comparisons workbook is missing GDPhistQ")
        sheet = workbook["GDPhistQ"]
        rows = [list(row) + [None] * 8 for row in sheet.iter_rows(values_only=True)]
        growth_index = next(
            (
                index
                for index, values in enumerate(rows)
                if "Comparisons -- Percent Change from Preceding Period"
                in str(values[0] or "").replace("\n", " ")
            ),
            None,
        )
        contribution_index = next(
            (
                index
                for index, values in enumerate(rows)
                if "Comparisons -- Contributions to Percent Change"
                in str(values[0] or "").replace("\n", " ")
            ),
            None,
        )
        if growth_index is None or contribution_index is None:
            raise ValueError("BEA comparisons workbook is missing growth or contribution sections")
        growth_title = str(rows[growth_index][0] or "")
        match = re.search(r"(\d{4}Q[1-4])", growth_title)
        if not match:
            raise ValueError("BEA comparisons workbook has no estimate quarter")
        period = match.group(1)
        previous_period = _previous_quarter(period)
        release_date = _release_date(sheet.cell(1, 1).value)
        estimate_round = _estimate_round(growth_title)

        def parse_section(
            section_rows: list[list[Any]],
            series: dict[str, str],
            *,
            unit: str,
            section_name: str,
        ) -> tuple[list[dict[str, Any]], set[str]]:
            records: list[dict[str, Any]] = []
            found: set[str] = set()
            context = ""
            for values in section_rows:
                label = str(values[0] or "").strip()
                if label in {
                    "Personal consumption expenditures",
                    "Gross private domestic investment",
                    "Exports",
                    "Imports",
                    "Government consumption expenditures and gross investment",
                    "Net exports of goods and services",
                }:
                    context = label
                series_id = series.get(label)
                if section_name == "growth" and context == "Personal consumption expenditures":
                    if label == "Goods":
                        series_id = "BEA-PCE-GOODS-GROWTH"
                    elif label == "Services":
                        series_id = "BEA-PCE-SERVICES-GROWTH"
                current_value = _decimal(values[1])
                if not series_id or current_value is None:
                    continue
                metadata = {
                    "estimate_quarter": period,
                    "estimate_round": estimate_round,
                    "source_revision_date": release_date,
                    "seasonal_adjustment": "SAAR",
                    "unit": unit,
                    "component_label": label,
                    "comparison_section": section_name,
                }
                records.append(
                    {
                        "series_id": series_id,
                        "date": _quarter_start(period),
                        "value": current_value,
                        "metadata": metadata,
                    }
                )
                for period_index, value_index in ((2, 3), (4, 5), (6, 7)):
                    if str(values[period_index] or "").strip().upper() == previous_period:
                        previous_value = _decimal(values[value_index])
                        if previous_value is not None:
                            records.append(
                                {
                                    "series_id": series_id,
                                    "date": _quarter_start(previous_period),
                                    "value": previous_value,
                                    "metadata": {
                                        **metadata,
                                        "estimate_quarter": previous_period,
                                    },
                                }
                            )
                        break
                found.add(label)
            return records, found

        growth_records, growth_found = parse_section(
            rows[growth_index + 1 : contribution_index],
            self.GROWTH_SERIES,
            unit="percent change from preceding period",
            section_name="growth",
        )
        contribution_end = next(
            (
                index
                for index in range(contribution_index + 1, len(rows))
                if "Comparisons --" in str(rows[index][0] or "").replace("\n", " ")
            ),
            len(rows),
        )
        contribution_records, contribution_found = parse_section(
            rows[contribution_index + 1 : contribution_end],
            self.CONTRIBUTION_SERIES,
            unit="percentage points contribution to real GDP growth",
            section_name="contribution",
        )
        required_growth = {"Personal consumption expenditures", "Goods", "Services"}
        required_contributions = set(self.CONTRIBUTION_SERIES)
        if not required_growth <= growth_found:
            raise ValueError("BEA comparisons workbook is missing required PCE growth components")
        if not required_contributions <= contribution_found:
            raise ValueError("BEA comparisons workbook is missing required GDP contributions")
        return [*growth_records, *contribution_records], {
            "comparison_quarter": period,
            "comparison_release_date": release_date,
            "comparison_estimate_round": estimate_round,
        }


class BEAPIOReleaseProvider(_ReleaseWorkbookProvider):
    """Parse BEA PIO history and cross-check it against the current release summary."""

    key = "bea-pio-release"
    base_url = "https://www.bea.gov"
    allowed_hosts = frozenset({"bea.gov", "www.bea.gov", "apps.bea.gov"})
    max_workbook_bytes = 8 * 1024 * 1024

    SERIES = {
        "BEA-REAL-PCE-MOM": {
            "sheet": "T20801-M",
            "code": "DPCERAM",
            "label": "Real personal consumption expenditures",
            "unit": "percent change from preceding month",
            "seasonal_adjustment": "seasonally adjusted monthly percent change",
        },
        "BEA-REAL-DPI-MOM": {
            "sheet": "T20600-M",
            "code": "A067RM",
            "label": "Real disposable personal income",
            "unit": "percent change from preceding month",
            "seasonal_adjustment": "seasonally adjusted monthly percent change",
            "requires_reference_year": True,
        },
        "BEA-PERSONAL-SAVING-RATE": {
            "sheet": "T20600-M",
            "code": "A072RC",
            "label": "Personal saving as a percentage of disposable personal income",
            "unit": "percent of disposable personal income",
            "seasonal_adjustment": "seasonally adjusted",
        },
        "BEA-DPI-NOMINAL-SAAR": {
            "sheet": "T20600-M",
            "code": "A067RC",
            "label": "Disposable personal income",
            "unit": "USD millions SAAR",
            "seasonal_adjustment": "seasonally adjusted at annual rates",
        },
        "BEA-DPI-REAL-SAAR": {
            "sheet": "T20600-M",
            "code": "A067RX",
            "label": "Real disposable personal income",
            "unit": "millions of chained dollars SAAR",
            "seasonal_adjustment": "seasonally adjusted at annual rates",
            "requires_reference_year": True,
        },
        "BEA-DPI-NOMINAL-MOM": {
            "sheet": "T20600-M",
            "code": "A067RCM",
            "label": "Disposable personal income",
            "unit": "percent change from preceding month",
            "seasonal_adjustment": "seasonally adjusted monthly percent change",
        },
        "BEA-REAL-PCE-SAAR": {
            "sheet": "T20806-M",
            "code": "DPCERX",
            "label": "Real personal consumption expenditures",
            "unit": "millions of chained dollars SAAR",
            "seasonal_adjustment": "seasonally adjusted at annual rates",
            "requires_reference_year": True,
        },
        "BEA-PCE-PRICE-INDEX": {
            "sheet": "T20804-M",
            "code": "DPCERG",
            "label": "Personal consumption expenditures price index",
            "unit": "chain-type price index",
            "seasonal_adjustment": "seasonally adjusted",
            "requires_reference_year": True,
        },
        "BEA-CORE-PCE-PRICE-INDEX": {
            "sheet": "T20804-M",
            "code": "DPCCRG",
            "label": "PCE price index excluding food and energy",
            "unit": "chain-type price index",
            "seasonal_adjustment": "seasonally adjusted",
            "requires_reference_year": True,
        },
    }
    SHEET_CONTRACTS = {
        "T20600-M": (
            "personal income and its disposition",
            "millions of dollars",
            "seasonally adjusted at annual rates",
        ),
        "T20801-M": (
            "percent change from preceding period in real personal consumption expenditures",
            "[percent]",
        ),
        "T20806-M": (
            "real personal consumption expenditures",
            "millions of chained",
            "seasonally adjusted at annual rates",
        ),
        "T20804-M": (
            "price indexes for personal consumption expenditures",
            "index numbers",
        ),
    }
    LATEST_ACCEPTABLE_HISTORY_START = {
        "T20600-M": "1959-01-01",
        "T20801-M": "1959-02-01",
        "T20806-M": "2007-01-01",
        "T20804-M": "1959-01-01",
    }
    SUMMARY_SERIES = frozenset(
        {
            "BEA-REAL-PCE-MOM",
            "BEA-REAL-DPI-MOM",
            "BEA-PERSONAL-SAVING-RATE",
        }
    )

    def personal_income_outlays(self) -> ProviderResult:
        dataset = "personal-income-outlays-release"
        try:
            page, page_type, page_modified = self._download(
                BEA_PIO_PAGE,
                expected="html",
            )
            summary_url = self._workbook_url(page)
            summary, summary_type, summary_modified = self._download(
                summary_url, expected="xlsx"
            )
            section2, section2_type, section2_modified = self._download(
                BEA_PIO_SECTION2_WORKBOOK, expected="xlsx"
            )
            summary_values, summary_metadata = self._parse_summary_workbook(summary)
            records, history_metadata = self._parse_section2_workbook(section2)
            self._cross_check(summary_values, summary_metadata, records, history_metadata)
            raw_bundle, raw_bundle_metadata = build_evidence_bundle(
                provider=self.key,
                dataset=dataset,
                responses=(
                    EvidenceResponse(
                        role="release-page",
                        url=BEA_PIO_PAGE,
                        content_type=page_type,
                        raw_bytes=page,
                        response_witness={"last_modified": page_modified},
                    ),
                    EvidenceResponse(
                        role="summary-workbook",
                        url=summary_url,
                        content_type=summary_type,
                        raw_bytes=summary,
                        request_witness={"discovered_from": "release-page"},
                        response_witness={"last_modified": summary_modified},
                    ),
                    EvidenceResponse(
                        role="section2-workbook",
                        url=BEA_PIO_SECTION2_WORKBOOK,
                        content_type=section2_type,
                        raw_bytes=section2,
                        response_witness={"last_modified": section2_modified},
                    ),
                ),
            )
        except Exception as exc:
            return ProviderResult.failure(self.key, dataset, f"{type(exc).__name__}: {exc}")
        return ProviderResult(
            provider=self.key,
            dataset=dataset,
            records=records,
            raw_bytes=raw_bundle,
            metadata={
                **raw_bundle_metadata,
                "source_url": BEA_PIO_PAGE,
                "summary_workbook_url": summary_url,
                "section2_workbook_url": BEA_PIO_SECTION2_WORKBOOK,
                "summary_workbook_last_modified": summary_modified,
                "section2_workbook_last_modified": section2_modified,
                "artifacts": [
                    _artifact(BEA_PIO_PAGE, page, page_type),
                    _artifact(summary_url, summary, summary_type),
                    _artifact(BEA_PIO_SECTION2_WORKBOOK, section2, section2_type),
                ],
                "vintage_policy": (
                    "full history as revised in the latest official Section 2 release; "
                    "normalized observations contain the current release vintage only"
                ),
                "revision_storage": (
                    "raw release artifact hash retained; queryable cross-vintage history "
                    "is not yet implemented"
                ),
                "attribution": "U.S. Bureau of Economic Analysis",
                **history_metadata,
                "summary_workbook_title": summary_metadata["workbook_title"],
                "summary_cross_check": "passed",
            },
        )

    @classmethod
    def replay_evidence_bundle(
        cls,
        raw_bytes: bytes,
    ) -> tuple[list[dict[str, Any]], dict[str, Any]]:
        evidence = parse_evidence_bundle(
            raw_bytes,
            expected_provider=cls.key,
            expected_dataset="personal-income-outlays-release",
        )
        if set(evidence.responses) != {
            "release-page",
            "summary-workbook",
            "section2-workbook",
        }:
            raise ValueError("BEA PIO evidence roles are incomplete")
        urls = {
            item["role"]: item["url"]
            for item in evidence.manifest["responses"]
        }
        page = evidence.responses["release-page"]
        summary_url = cls._workbook_url(page)
        if (
            urls["release-page"] != BEA_PIO_PAGE
            or urls["summary-workbook"] != summary_url
            or urls["section2-workbook"] != BEA_PIO_SECTION2_WORKBOOK
        ):
            raise ValueError("BEA PIO evidence URLs do not match official discovery")
        summary_bytes = _xlsx_bytes(
            evidence.responses["summary-workbook"],
            max_expanded_bytes=cls.max_expanded_bytes,
        )
        section2_bytes = _xlsx_bytes(
            evidence.responses["section2-workbook"],
            max_expanded_bytes=cls.max_expanded_bytes,
        )
        summary_values, summary_metadata = cls._parse_summary_workbook(
            summary_bytes
        )
        records, history_metadata = cls._parse_section2_workbook(
            section2_bytes
        )
        cls._cross_check(
            summary_values,
            summary_metadata,
            records,
            history_metadata,
        )
        return records, {
            "source_url": BEA_PIO_PAGE,
            "summary_workbook_url": summary_url,
            "section2_workbook_url": BEA_PIO_SECTION2_WORKBOOK,
            **history_metadata,
            "summary_workbook_title": summary_metadata["workbook_title"],
            "summary_cross_check": "passed",
        }

    @staticmethod
    def _workbook_url(page: bytes) -> str:
        parser = _LinkParser()
        parser.feed(page.decode("utf-8", errors="replace"))
        candidates = []
        for href, label in parser.links:
            absolute = urljoin(BEA_PIO_PAGE, href)
            parsed = urlparse(absolute)
            if (
                parsed.scheme == "https"
                and parsed.hostname in {"bea.gov", "www.bea.gov"}
                and parsed.path.lower().endswith(".xlsx")
                and (
                    "Historical Comparisons" in label
                    or re.search(r"/pi\d{4}-hist\.xlsx$", parsed.path, re.IGNORECASE)
                )
            ):
                candidates.append(absolute)
        if len(set(candidates)) != 1:
            raise ValueError("BEA PIO page must expose exactly one historical-comparisons workbook")
        return candidates[0]

    @classmethod
    def _parse_summary_workbook(
        cls, content: bytes
    ) -> tuple[dict[str, Decimal], dict[str, Any]]:
        workbook = load_workbook(BytesIO(content), read_only=True, data_only=True)
        try:
            if "PIOhist_M" not in workbook.sheetnames:
                raise ValueError("BEA PIO summary workbook is missing PIOhist_M")
            sheet = workbook["PIOhist_M"]
            rows = [list(row) + [None] * 8 for row in sheet.iter_rows(values_only=True)]
            title = str(rows[1][0] or "") if len(rows) > 1 else ""
            if len(rows) < 6 or "Personal Income and Outlays" not in title:
                raise ValueError("BEA PIO summary workbook has no release title")
            if "Historical Comparisons" not in str(rows[2][0] or ""):
                raise ValueError("BEA PIO summary workbook has no historical-comparisons marker")
            current_date = _month_start(rows[4][1]) or _month_start(title)
            title_date = _month_start(title)
            release_date = _release_date(rows[0][6])
            if not current_date or current_date != title_date or not release_date:
                raise ValueError("BEA PIO release month or release date is inconsistent")
            if datetime.fromisoformat(release_date) <= datetime.fromisoformat(current_date):
                raise ValueError("BEA PIO release date must follow its observation month")

            target_rows: dict[str, list[Any]] = {}
            section = ""
            subsection = ""

            def capture(series_id: str, values: list[Any]) -> None:
                if series_id in target_rows:
                    raise ValueError(f"BEA PIO summary duplicated required series {series_id}")
                target_rows[series_id] = values

            for values in rows:
                label = str(values[0] or "").strip()
                if label in {
                    "Current dollars",
                    "Chained dollars",
                    "Chain-type price indexes",
                    "Personal saving as a percentage of DPI",
                }:
                    section = label
                    subsection = ""
                    continue
                if label.endswith("change from preceding month:"):
                    subsection = "month-over-month"
                    continue
                if section == "Chained dollars" and subsection == "month-over-month":
                    if label == "DPI":
                        capture("BEA-REAL-DPI-MOM", values)
                    elif label == "PCE":
                        capture("BEA-REAL-PCE-MOM", values)
                elif (
                    section == "Personal saving as a percentage of DPI"
                    and label == "Personal saving rate"
                ):
                    capture("BEA-PERSONAL-SAVING-RATE", values)

            current_values: dict[str, Decimal] = {}
            for series_id in cls.SUMMARY_SERIES:
                values = target_rows.get(series_id)
                value = _decimal(values[1]) if values else None
                if value is None:
                    raise ValueError(f"BEA PIO summary is missing current value for {series_id}")
                current_values[series_id] = value
            return current_values, {
                "latest_value_date": current_date,
                "source_revision_date": release_date,
                "workbook_title": title,
            }
        finally:
            workbook.close()

    @classmethod
    def _parse_section2_workbook(
        cls, content: bytes
    ) -> tuple[list[dict[str, Any]], dict[str, Any]]:
        workbook = load_workbook(BytesIO(content), read_only=True, data_only=True)
        records: list[dict[str, Any]] = []
        release_dates: set[str] = set()
        latest_dates: set[str] = set()
        series_counts: dict[str, int] = {}
        try:
            definitions_by_sheet: dict[str, dict[str, tuple[str, dict[str, Any]]]] = {}
            for series_id, definition in cls.SERIES.items():
                sheet_definitions = definitions_by_sheet.setdefault(definition["sheet"], {})
                code = definition["code"]
                if code in sheet_definitions:
                    raise ValueError(f"duplicate BEA NIPA code configured: {code}")
                sheet_definitions[code] = (series_id, definition)

            for sheet_name, sheet_definitions in definitions_by_sheet.items():
                if sheet_name not in workbook.sheetnames:
                    raise ValueError(f"BEA Section 2 workbook is missing {sheet_name}")
                sheet = workbook[sheet_name]
                sheet_contract = " ".join(
                    str(sheet.cell(row=row, column=1).value or "")
                    for row in (1, 2)
                ).casefold()
                if any(
                    marker not in sheet_contract
                    for marker in cls.SHEET_CONTRACTS[sheet_name]
                ):
                    raise ValueError(
                        f"BEA Section 2 sheet {sheet_name} title/unit contract changed"
                    )
                coverage_text = str(sheet.cell(row=3, column=1).value or "")
                coverage_match = re.search(
                    r"Monthly data from\s+(\d{4}M\d{2})\s+to\s+(\d{4}M\d{2})",
                    coverage_text,
                    re.IGNORECASE,
                )
                if not coverage_match:
                    raise ValueError(
                        f"BEA Section 2 sheet {sheet_name} has no declared monthly coverage"
                    )
                declared_start = _month_start(coverage_match.group(1))
                declared_end = _month_start(coverage_match.group(2))
                release_date = _release_date(sheet.cell(row=5, column=1).value)
                if not release_date:
                    raise ValueError(f"BEA Section 2 sheet {sheet_name} has no release date")
                release_dates.add(release_date)

                header = next(
                    sheet.iter_rows(min_row=8, max_row=8, values_only=True),
                    None,
                )
                if not header:
                    raise ValueError(f"BEA Section 2 sheet {sheet_name} has no period header")
                periods: list[tuple[int, str]] = []
                for index, raw_period in enumerate(header[3:], start=3):
                    if raw_period in (None, ""):
                        continue
                    period = _month_start(raw_period)
                    if not period:
                        raise ValueError(
                            f"BEA Section 2 sheet {sheet_name} has invalid period {raw_period!r}"
                        )
                    periods.append((index, period))
                period_dates = [period for _, period in periods]
                if len(period_dates) < 2 or period_dates != sorted(set(period_dates)):
                    raise ValueError(
                        f"BEA Section 2 sheet {sheet_name} periods are not unique and ascending"
                    )
                if any(
                    _previous_month(current) != previous
                    for previous, current in zip(period_dates, period_dates[1:], strict=False)
                ):
                    raise ValueError(f"BEA Section 2 sheet {sheet_name} periods are not contiguous")
                if (declared_start, declared_end) != (
                    period_dates[0],
                    period_dates[-1],
                ):
                    raise ValueError(
                        f"BEA Section 2 sheet {sheet_name} coverage does not match headers"
                    )
                if period_dates[0] > cls.LATEST_ACCEPTABLE_HISTORY_START[sheet_name]:
                    raise ValueError(
                        f"BEA Section 2 sheet {sheet_name} historical coverage is truncated"
                    )
                latest_date = period_dates[-1]
                latest_dates.add(latest_date)

                matched_rows: dict[str, tuple[Any, ...]] = {}
                for row in sheet.iter_rows(min_row=9, values_only=True):
                    code = str(row[2] or "").strip() if len(row) > 2 else ""
                    if code not in sheet_definitions:
                        continue
                    if code in matched_rows:
                        raise ValueError(
                            f"BEA Section 2 sheet {sheet_name} duplicated NIPA code {code}"
                        )
                    matched_rows[code] = row
                missing_codes = set(sheet_definitions) - set(matched_rows)
                if missing_codes:
                    raise ValueError(
                        f"BEA Section 2 sheet {sheet_name} is missing NIPA codes "
                        f"{', '.join(sorted(missing_codes))}"
                    )

                for code, (series_id, definition) in sheet_definitions.items():
                    row = matched_rows[code]
                    reference_year = None
                    if definition.get("requires_reference_year"):
                        reference_match = re.search(
                            r"chained\s*\((\d{4})\)\s*dollars",
                            f"{sheet_contract} {str(row[1] or '').casefold()}",
                        )
                        if not reference_match:
                            reference_match = re.search(
                                r"\b(\d{4})\s*=\s*100\b",
                                f"{sheet_contract} {str(row[1] or '').casefold()}",
                            )
                        if not reference_match:
                            raise ValueError(
                                f"BEA Section 2 series {series_id} has no reference year"
                            )
                        reference_year = int(reference_match.group(1))
                    count = 0
                    latest_value: Decimal | None = None
                    numeric_history_started = False
                    for index, period in periods:
                        value = _decimal(row[index] if index < len(row) else None)
                        if value is None:
                            if numeric_history_started:
                                raise ValueError(
                                    f"BEA Section 2 series {series_id} has a missing/non-numeric "
                                    f"value after history began at {period}"
                                )
                            continue
                        numeric_history_started = True
                        metadata = {
                            "source_revision_date": release_date,
                            "release_freshness_days": 45,
                            "vintage_status": "current_release_vintage",
                            "official_series_code": code,
                            "source_table": sheet_name,
                            "component_label": definition["label"],
                            "unit": definition["unit"],
                            "seasonal_adjustment": definition["seasonal_adjustment"],
                            "estimate_month": period[:7],
                        }
                        if reference_year:
                            metadata["reference_year"] = reference_year
                            if "chained dollars" in metadata["unit"]:
                                metadata["unit"] = metadata["unit"].replace(
                                    "chained dollars",
                                    f"chained {reference_year} dollars",
                                )
                        records.append(
                            {
                                "series_id": series_id,
                                "date": period,
                                "value": value,
                                "metadata": metadata,
                            }
                        )
                        count += 1
                        if period == latest_date:
                            latest_value = value
                    if not count or latest_value is None:
                        raise ValueError(
                            f"BEA Section 2 series {series_id} has no current numeric value"
                        )
                    series_counts[series_id] = count

            if len(release_dates) != 1 or len(latest_dates) != 1:
                raise ValueError("BEA Section 2 sheets do not share one release date and month")
            release_date = next(iter(release_dates))
            latest_date = next(iter(latest_dates))
            if datetime.fromisoformat(release_date) <= datetime.fromisoformat(latest_date):
                raise ValueError("BEA Section 2 release date must follow its latest month")
            if set(series_counts) != set(cls.SERIES):
                raise ValueError("BEA Section 2 workbook did not yield every configured series")
            return records, {
                "latest_value_date": latest_date,
                "source_revision_date": release_date,
                "series_counts": series_counts,
                "record_count": len(records),
                "precision_policy": "retain workbook values without additional rounding",
            }
        finally:
            workbook.close()

    @classmethod
    def _cross_check(
        cls,
        summary_values: dict[str, Decimal],
        summary_metadata: dict[str, Any],
        records: list[dict[str, Any]],
        history_metadata: dict[str, Any],
    ) -> None:
        if summary_metadata["latest_value_date"] != history_metadata["latest_value_date"]:
            raise ValueError("BEA PIO summary and Section 2 latest months do not match")
        if summary_metadata["source_revision_date"] != history_metadata["source_revision_date"]:
            raise ValueError("BEA PIO summary and Section 2 release dates do not match")
        latest_date = history_metadata["latest_value_date"]
        latest_values = {
            item["series_id"]: item["value"]
            for item in records
            if item["date"] == latest_date and item["series_id"] in cls.SUMMARY_SERIES
        }
        if set(latest_values) != cls.SUMMARY_SERIES:
            raise ValueError("BEA Section 2 current cross-check series are incomplete")
        for series_id, summary_value in summary_values.items():
            if latest_values[series_id] != summary_value:
                raise ValueError(
                    f"BEA PIO summary and Section 2 disagree for {series_id}: "
                    f"{summary_value} != {latest_values[series_id]}"
                )


class CensusMARTSReleaseProvider(_ReleaseWorkbookProvider):
    """Parse Census Monthly Retail Trade official release workbooks without an API key."""

    key = "census-release"
    base_url = "https://www2.census.gov"
    allowed_hosts = frozenset({"www2.census.gov", "www.census.gov"})

    def monthly_retail_sales(self) -> ProviderResult:
        dataset = "marts:retail-food-services"
        current_error: Exception | None = None
        try:
            workbook_url = CENSUS_MARTS_CURRENT_WORKBOOK
            content, content_type, last_modified = self._download(
                workbook_url, expected="xlsx"
            )
            artifacts = [_artifact(workbook_url, content, content_type)]
            responses = (
                EvidenceResponse(
                    role="current-workbook",
                    url=workbook_url,
                    content_type=content_type,
                    raw_bytes=content,
                    request_witness={"method": "GET", "scope": "current"},
                    response_witness={"last_modified": last_modified},
                ),
            )
        except Exception as exc:
            current_error = exc
            try:
                index, index_type, index_modified = self._download(
                    CENSUS_MARTS_INDEX, expected="html"
                )
                workbook_url = self._latest_workbook_url(index)
                content, content_type, last_modified = self._download(
                    workbook_url, expected="xlsx"
                )
                artifacts = [
                    _artifact(CENSUS_MARTS_INDEX, index, index_type),
                    _artifact(workbook_url, content, content_type),
                ]
                responses = (
                    EvidenceResponse(
                        role="archive-index",
                        url=CENSUS_MARTS_INDEX,
                        content_type=index_type,
                        raw_bytes=index,
                        request_witness={
                            "method": "GET",
                            "scope": "historical_archive",
                        },
                        response_witness={"last_modified": index_modified},
                    ),
                    EvidenceResponse(
                        role="archive-workbook",
                        url=workbook_url,
                        content_type=content_type,
                        raw_bytes=content,
                        request_witness={
                            "method": "GET",
                            "scope": "historical_archive",
                            "discovered_from": "archive-index",
                        },
                        response_witness={"last_modified": last_modified},
                    ),
                )
            except Exception as fallback_exc:
                reason = (
                    f"current {type(current_error).__name__}: {current_error}; "
                    f"archive {type(fallback_exc).__name__}: {fallback_exc}"
                )
                return ProviderResult.failure(self.key, dataset, reason)
        try:
            raw_bundle, bundle_metadata = build_evidence_bundle(
                provider=self.key,
                dataset=dataset,
                responses=responses,
            )
            records, metadata = self.replay_evidence_bundle(raw_bundle)
        except Exception as exc:
            return ProviderResult.failure(self.key, dataset, f"{type(exc).__name__}: {exc}")
        return ProviderResult(
            provider=self.key,
            dataset=dataset,
            records=records,
            raw_bytes=raw_bundle,
            metadata={
                **bundle_metadata,
                **metadata,
                "artifacts": artifacts,
            },
        )

    def historical_monthly_retail_sales(self) -> ProviderResult:
        """Parse the latest archived MARTS workbook; used only for diagnostics/tests."""
        dataset = "marts:retail-food-services"
        try:
            index, index_type, index_modified = self._download(
                CENSUS_MARTS_INDEX, expected="html"
            )
            workbook_url = self._latest_workbook_url(index)
            content, content_type, last_modified = self._download(workbook_url, expected="xlsx")
            responses = (
                EvidenceResponse(
                    role="archive-index",
                    url=CENSUS_MARTS_INDEX,
                    content_type=index_type,
                    raw_bytes=index,
                    request_witness={
                        "method": "GET",
                        "scope": "historical_archive",
                    },
                    response_witness={"last_modified": index_modified},
                ),
                EvidenceResponse(
                    role="archive-workbook",
                    url=workbook_url,
                    content_type=content_type,
                    raw_bytes=content,
                    request_witness={
                        "method": "GET",
                        "scope": "historical_archive",
                        "discovered_from": "archive-index",
                    },
                    response_witness={"last_modified": last_modified},
                ),
            )
            raw_bundle, bundle_metadata = build_evidence_bundle(
                provider=self.key,
                dataset=dataset,
                responses=responses,
            )
            records, metadata = self.replay_evidence_bundle(raw_bundle)
        except Exception as exc:
            return ProviderResult.failure(self.key, dataset, f"{type(exc).__name__}: {exc}")
        artifacts = [
            _artifact(CENSUS_MARTS_INDEX, index, index_type),
            _artifact(workbook_url, content, content_type),
        ]
        return ProviderResult(
            provider=self.key,
            dataset=dataset,
            records=records,
            raw_bytes=raw_bundle,
            metadata={
                **bundle_metadata,
                **metadata,
                "artifacts": artifacts,
            },
        )

    @classmethod
    def replay_evidence_bundle(
        cls,
        raw_bytes: bytes,
    ) -> tuple[list[dict[str, Any]], dict[str, Any]]:
        evidence = parse_evidence_bundle(
            raw_bytes,
            expected_provider=cls.key,
            expected_dataset="marts:retail-food-services",
        )
        roles = set(evidence.responses)
        entries = {
            item["role"]: item for item in evidence.manifest["responses"]
        }
        allowed_workbook_types = {
            XLSX_CONTENT_TYPE,
            "application/octet-stream",
            "application/zip",
        }
        if roles == {"current-workbook"}:
            workbook_entry = entries["current-workbook"]
            if (
                workbook_entry["url"] != CENSUS_MARTS_CURRENT_WORKBOOK
                or workbook_entry["content_type"] not in allowed_workbook_types
                or workbook_entry["request_witness"]
                != {"method": "GET", "scope": "current"}
                or set(workbook_entry["response_witness"])
                != {"last_modified"}
            ):
                raise ValueError("Census current-workbook evidence contract is invalid")
            workbook_scope = "current"
            source_url = CENSUS_MARTS_RELEASE_PAGE
            workbook_url = CENSUS_MARTS_CURRENT_WORKBOOK
            workbook_bytes = evidence.responses["current-workbook"]
        elif roles == {"archive-index", "archive-workbook"}:
            index_entry = entries["archive-index"]
            workbook_entry = entries["archive-workbook"]
            if (
                index_entry["url"] != CENSUS_MARTS_INDEX
                or index_entry["content_type"] != "text/html"
                or index_entry["request_witness"]
                != {"method": "GET", "scope": "historical_archive"}
                or set(index_entry["response_witness"]) != {"last_modified"}
                or workbook_entry["content_type"] not in allowed_workbook_types
                or workbook_entry["request_witness"]
                != {
                    "method": "GET",
                    "scope": "historical_archive",
                    "discovered_from": "archive-index",
                }
                or set(workbook_entry["response_witness"])
                != {"last_modified"}
            ):
                raise ValueError("Census archive evidence contract is invalid")
            workbook_url = cls._latest_workbook_url(
                evidence.responses["archive-index"]
            )
            if workbook_entry["url"] != workbook_url:
                raise ValueError("Census archive workbook URL does not replay from index")
            workbook_scope = "historical_archive"
            source_url = CENSUS_MARTS_INDEX
            workbook_bytes = evidence.responses["archive-workbook"]
        else:
            raise ValueError("Census MARTS evidence roles are incomplete or mixed")
        workbook_bytes = _xlsx_bytes(
            workbook_bytes,
            max_expanded_bytes=cls.max_expanded_bytes,
        )
        records, parser_metadata = cls._parse_workbook(workbook_bytes)
        return records, {
            "source_url": source_url,
            "workbook_url": workbook_url,
            "workbook_last_modified": workbook_entry["response_witness"][
                "last_modified"
            ],
            "workbook_scope": workbook_scope,
            "unit": "USD millions",
            "vintage_policy": (
                "latest official advance/preliminary/revised release workbook"
            ),
            "attribution": "U.S. Census Bureau",
            **parser_metadata,
        }

    @staticmethod
    def _latest_workbook_url(index: bytes) -> str:
        parser = _LinkParser()
        parser.feed(index.decode("latin-1", errors="replace"))
        candidates: list[tuple[int, int, str]] = []
        for href, _ in parser.links:
            match = re.fullmatch(r"rs(\d{2})(\d{2})\.xlsx", href.strip(), re.I)
            if not match:
                continue
            short_year, month = (int(value) for value in match.groups())
            year = 2000 + short_year if short_year < 80 else 1900 + short_year
            if 1 <= month <= 12:
                candidates.append((year, month, href))
        if not candidates:
            raise ValueError("Census MARTS directory exposed no release workbook")
        _, _, href = max(candidates)
        return urljoin(CENSUS_MARTS_INDEX, href)

    @staticmethod
    def _month_date(year: int, label: Any) -> str | None:
        text = re.sub(r"[^A-Za-z]", "", str(label or ""))[:3].title()
        try:
            month = datetime.strptime(text, "%b").month
        except ValueError:
            return None
        return f"{year:04d}-{month:02d}-01"

    @staticmethod
    def _month_from_text(value: Any) -> str | None:
        match = re.search(r"([A-Z][a-z]{2})\.?\s+(\d{4})", str(value or ""))
        if not match:
            return None
        try:
            month = datetime.strptime(match.group(1), "%b").month
        except ValueError:
            return None
        return f"{int(match.group(2)):04d}-{month:02d}-01"

    @classmethod
    def _parse_workbook(cls, content: bytes) -> tuple[list[dict[str, Any]], dict[str, Any]]:
        workbook = load_workbook(BytesIO(content), read_only=True, data_only=True)
        if not {"Table 1.", "Table 2."} <= set(workbook.sheetnames):
            raise ValueError("Census MARTS workbook is missing required tables")
        sales_sheet = workbook["Table 1."]
        rows = [list(row) for row in sales_sheet.iter_rows(values_only=True)]
        sales_declaration = " ".join(
            str(value or "")
            for row in rows[:5]
            for value in row
        )
        sales_declaration = " ".join(sales_declaration.casefold().split())
        if "total sales estimates are shown in millions of dollars" not in (
            sales_declaration
        ):
            raise ValueError("Census MARTS sales table declared unit is invalid")
        adjusted_column = None
        for row in rows[:15]:
            for index, value in enumerate(row):
                if isinstance(value, str) and value.strip().lower().startswith("adjusted"):
                    adjusted_column = index
                    break
            if adjusted_column is not None:
                break
        if adjusted_column is None:
            raise ValueError("Census MARTS adjusted-sales columns were not found")

        year_by_column: dict[int, int] = {}
        month_by_column: dict[int, Any] = {}
        status_by_column: dict[int, str] = {}
        active_year: int | None = None
        for row in rows[:15]:
            for index in range(adjusted_column, len(row)):
                value = row[index]
                if isinstance(value, int) and 1900 <= value <= 2100:
                    active_year = value
                if active_year is not None:
                    year_by_column.setdefault(index, active_year)
                if isinstance(value, str) and re.search(r"[A-Za-z]{3}", value):
                    if cls._month_date(active_year or 0, value):
                        month_by_column[index] = value
                if isinstance(value, str) and re.fullmatch(r"\([apr]\)", value.strip(), re.I):
                    status_by_column[index] = value.strip().lower()

        total_row = None
        previous_label = ""
        for row in rows:
            label = str(row[1] or "") if len(row) > 1 else ""
            if "Retail & food services" in previous_label and label.strip().lower().startswith("total"):
                total_row = row
                break
            if label.strip():
                previous_label = label
        if total_row is None:
            raise ValueError("Census MARTS total sales row was not found")

        records = []
        dates: list[str] = []
        status_by_date: dict[str, str] = {}
        for index in range(adjusted_column, min(adjusted_column + 3, len(total_row))):
            value = _decimal(total_row[index])
            year = year_by_column.get(index)
            value_date = cls._month_date(year or 0, month_by_column.get(index))
            if value is None or value_date is None:
                continue
            dates.append(value_date)
            status_by_date[value_date] = status_by_column.get(index, "")
            records.append(
                {
                    "series_id": "CENSUS-MRTS-44X72-SM-SA",
                    "date": value_date,
                    "value": value,
                    "metadata": {
                        "unit": "USD millions",
                        "seasonally_adjusted": True,
                        "estimate_status": status_by_column.get(index, ""),
                        "category": "Retail and food services, total",
                    },
                }
            )
        if len(records) < 2:
            raise ValueError("Census MARTS workbook yielded insufficient monthly sales")
        if dates != sorted(set(dates), reverse=True):
            raise ValueError("Census MARTS adjusted-sales months are not unique and descending")

        change_sheet = workbook["Table 2."]
        change_rows = [list(row) for row in change_sheet.iter_rows(values_only=True)]
        change_declaration = " ".join(
            str(value or "")
            for row in change_rows[:5]
            for value in row
        )
        change_declaration = " ".join(change_declaration.casefold().split())
        if "estimates are shown as percents" not in change_declaration:
            raise ValueError("Census MARTS change table declared unit is invalid")
        change_total = None
        previous_label = ""
        for row in change_rows:
            label = str(row[1] or "") if len(row) > 1 else ""
            if "Retail & food services" in previous_label and label.strip().lower().startswith("total"):
                change_total = row
                break
            if label.strip():
                previous_label = label
        if change_total is None:
            raise ValueError("Census MARTS change row was not found")
        change_columns: dict[str, int] = {}
        for row in change_rows[:15]:
            for index, value in enumerate(row):
                value_date = cls._month_from_text(value)
                if value_date and re.search(
                    r"\b(Advance|Preliminary|Revised)\b", str(value), re.IGNORECASE
                ):
                    change_columns.setdefault(value_date, index)
        if not all(value_date in change_columns for value_date in dates[:2]):
            raise ValueError("Census MARTS change-table month headers do not match sales table")

        for position, value_date in enumerate(dates[:2]):
            mom_index = change_columns[value_date]
            yoy_index = mom_index + 1
            comparison_row = next(
                (
                    row
                    for row in change_rows[:15]
                    if mom_index < len(row)
                    and cls._month_from_text(row[mom_index]) not in {None, value_date}
                ),
                None,
            )
            if comparison_row is None or yoy_index >= len(comparison_row):
                raise ValueError("Census MARTS change-table comparison headers are incomplete")
            prior_month = cls._month_from_text(comparison_row[mom_index])
            year_ago = cls._month_from_text(comparison_row[yoy_index])
            expected_prior = dates[position + 1]
            expected_year_ago = f"{int(value_date[:4]) - 1:04d}{value_date[4:]}"
            if prior_month != expected_prior:
                raise ValueError("Census MARTS current month-over-month header is inconsistent")
            if year_ago != expected_year_ago:
                raise ValueError("Census MARTS year-over-year header is inconsistent")
            for series_id, value_index in (
                ("CENSUS-MRTS-44X72-SM-SA-MOM", mom_index),
                ("CENSUS-MRTS-44X72-SM-SA-YOY", yoy_index),
            ):
                value = _decimal(change_total[value_index])
                if value is not None:
                    records.append(
                        {
                            "series_id": series_id,
                            "date": value_date,
                            "value": value,
                            "metadata": {
                                "unit": "percent",
                                "seasonally_adjusted": True,
                                "estimate_status": status_by_date.get(value_date, ""),
                                "category": "Retail and food services, total",
                            },
                        }
                    )
        return records, {
            "latest_value_date": dates[0],
            "months_in_workbook": dates,
            "precision_policy": "retain workbook values without additional rounding",
        }
