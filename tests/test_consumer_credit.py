from __future__ import annotations

import csv
import hashlib
import io
from datetime import date
from decimal import Decimal
from pathlib import Path

import httpx
import pytest
from openpyxl import Workbook, load_workbook

import research.official_data as official_data
from research.consumer_contract import _validate_run
from research.consumer_credit import (
    G19_SERIES,
    FederalReserveG19Provider,
    NYFedHouseholdDebtProvider,
)
from research.models import IngestionRun, Observation, RawArtifact
from research.official_data import _store_consumer_credit_observations_v2
from research.raw_evidence import EvidenceResponse, build_evidence_bundle, parse_evidence_bundle
from research.services import record_provider_result


def _g19_csv() -> bytes:
    descriptions = list(G19_SERIES)
    identifiers = [f"G19/TEST/{index}.M" for index in range(len(descriptions))]
    rows = [
        ["Series Description", *descriptions],
        ["Unit:", *("Percent" if "Percent change" in item else "Currency" for item in descriptions)],
        ["Multiplier:", *("1" if "Percent change" in item else "1000000" for item in descriptions)],
        ["Currency:", *("USD" for _ in descriptions)],
        ["Unique Identifier:", *identifiers],
        ["Time Period", *(f"SERIES-{index}" for index in range(len(descriptions)))],
        ["2026-04", "4.87", "10.36", "2.93", "5154721.31", "1349505.63", "3805215.68", "20822.88", "11546.30", "9276.58"],
        ["2026-05", "-0.04", "-4.71", "1.61", "5154538.86", "1344207.79", "3810331.07", "-182.45", "-5297.84", "5115.39"],
    ]
    output = io.StringIO()
    csv.writer(output, lineterminator="\n").writerows(rows)
    return output.getvalue().encode()


def _g19_page() -> bytes:
    package = (
        "rel=G19&amp;series=test-package&amp;lastObs=&amp;from=&amp;to=&amp;filetype=csv&amp;"
        "label=include&amp;layout=seriescolumn&amp;type=package"
    )
    return (
        '<span id="ReleaseLabel">G.19 - last released Wednesday, July 8, 2026</span>'
        f'<input name="FreqRequest" value="{package}">'
        '<label>Consumer Credit Outstanding (S.A.) [csv]</label>'
    ).encode()


def _g19_client(
    csv_payload: bytes | None = None,
    *,
    page_payload: bytes | None = None,
) -> httpx.Client:
    payload = csv_payload or _g19_csv()
    page = page_payload or _g19_page()

    def handler(request: httpx.Request) -> httpx.Response:
        if request.url.path.endswith("Choose.aspx"):
            assert request.url.params["rel"] == "G19"
            return httpx.Response(200, content=page, headers={"content-type": "text/html"})
        assert request.url.path.endswith("Output.aspx")
        assert request.url.params["series"] == "test-package"
        return httpx.Response(200, content=payload, headers={"content-type": "text/csv"})

    return httpx.Client(
        base_url="https://www.federalreserve.gov",
        transport=httpx.MockTransport(handler),
    )


def _hhdc_workbook(
    *,
    latest_quarter: str = "26:Q1",
    balance_unit: str = "Trillions of $",
    delinquency_unit: str = "Percent",
) -> bytes:
    workbook = Workbook()
    contents = workbook.active
    contents.title = "TABLE OF CONTENTS"
    contents.cell(2, 2, "QUARTERLY REPORT ON HOUSEHOLD DEBT AND CREDIT")

    balances = workbook.create_sheet("Page 3 Data")
    balances.append(["Total Debt Balance and Its Composition"])
    balances.append([balance_unit])
    balances.append(["Return to Table of Contents"])
    balances.append(
        [None, "Mortgage", "HE Revolving", "Auto Loan", "Credit Card", "Student Loan", "Other", "Total"]
    )
    balances.append(["25:Q4", 13.17, 0.4336, 1.667, 1.277, 1.664, 0.5641, 18.7757])
    balances.append([latest_quarter, 13.191, 0.446, 1.685, 1.252, 1.658, 0.562, 18.794])

    delinquencies = workbook.create_sheet("Page 12 Data")
    delinquencies.append(["Percent of Balance 90+ Days Delinquent by Loan Type"])
    delinquencies.append([delinquency_unit])
    delinquencies.append(["Return to Table of Contents"])
    delinquencies.append(
        [None, "MORTGAGE", "HELOC", "AUTO", "CC", "STUDENT LOAN", "OTHER", "ALL"]
    )
    delinquencies.append(["25:Q4", 0.92, 0.82, 5.21, 12.70, 9.57, 9.52, 3.12])
    delinquencies.append([latest_quarter, 1.09, 0.95, 5.6, 13.12, 10.34, 9.76, 3.36])

    output = io.BytesIO()
    workbook.save(output)
    return output.getvalue()


def _mutate_hhdc_workbook(payload: bytes, mutation: str) -> bytes:
    workbook = load_workbook(io.BytesIO(payload))
    if mutation == "duplicate-header":
        workbook["Page 3 Data"].cell(4, 9, "Mortgage")
    elif mutation == "invalid-value":
        workbook["Page 3 Data"].cell(5, 2, "not-a-number")
    elif mutation == "later-start":
        workbook["Page 3 Data"].cell(5, 2).value = None
    elif mutation == "started-gap":
        sheet = workbook["Page 3 Data"]
        sheet.insert_rows(5)
        for column, value in enumerate(
            ["25:Q3", 13.0, 0.42, 1.65, 1.25, 1.64, 0.55, 18.56],
            start=1,
        ):
            sheet.cell(5, column, value)
        sheet.cell(6, 2).value = None
    elif mutation == "entirely-missing":
        workbook["Page 3 Data"].cell(6, 2).value = None
        workbook["Page 3 Data"].cell(5, 2).value = None
    elif mutation == "duplicate-quarter":
        workbook["Page 12 Data"].cell(6, 1, "25:Q4")
    elif mutation == "quarter-gap":
        workbook["Page 12 Data"].cell(5, 1, "25:Q3")
    else:
        raise AssertionError(f"unknown HHDC mutation: {mutation}")
    output = io.BytesIO()
    workbook.save(output)
    workbook.close()
    return output.getvalue()


def _retime_evidence_bundle(
    raw_bytes: bytes,
    *,
    provider: str,
    dataset: str,
    retrieved_at: str,
    payload_overrides: dict[str, bytes] | None = None,
) -> bytes:
    evidence = parse_evidence_bundle(
        raw_bytes,
        expected_provider=provider,
        expected_dataset=dataset,
    )
    overrides = payload_overrides or {}
    responses = []
    for entry in evidence.manifest["responses"]:
        role = entry["role"]
        responses.append(
            EvidenceResponse(
                role=role,
                url=entry["url"],
                content_type=entry["content_type"],
                raw_bytes=overrides.get(role, evidence.responses[role]),
                request_witness=entry["request_witness"],
                response_witness={"retrieved_at": retrieved_at},
            )
        )
    return build_evidence_bundle(
        provider=provider,
        dataset=dataset,
        responses=tuple(responses),
    )[0]


def _hhdc_client(workbook: bytes, *, filename_quarter: str = "2026q1") -> httpx.Client:
    page = (
        '<a href="/medialibrary/interactives/householdcredit/data/xls/'
        f'hhd_c_report_{filename_quarter}.xlsx">Data Underlying Report</a>'
    ).encode()

    def handler(request: httpx.Request) -> httpx.Response:
        if request.url.path.endswith("databank.html"):
            return httpx.Response(200, content=page, headers={"content-type": "text/html"})
        return httpx.Response(
            200,
            content=workbook,
            headers={
                "content-type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            },
        )

    return httpx.Client(
        base_url="https://www.newyorkfed.org",
        transport=httpx.MockTransport(handler),
    )


def test_g19_provider_discovers_package_and_normalizes_full_history():
    payload = _g19_csv()
    result = FederalReserveG19Provider(client=_g19_client(payload)).consumer_credit()

    assert result.ok
    assert result.row_count == 18
    assert result.metadata["release_date"] == "2026-07-08"
    assert result.metadata["latest_value_date"] == "2026-05-01"
    assert result.metadata["artifacts"][1]["sha256"] == hashlib.sha256(payload).hexdigest()
    evidence = parse_evidence_bundle(
        result.raw_bytes,
        expected_provider="federal-reserve-g19",
        expected_dataset="consumer-credit",
    )
    assert set(evidence.responses) == {"choose-page", "output-csv"}
    assert evidence.responses["choose-page"] == _g19_page()
    assert evidence.responses["output-csv"] == payload
    replay_records, replay_metadata = FederalReserveG19Provider.replay_evidence_bundle(
        result.raw_bytes
    )
    assert replay_records == result.records
    assert replay_metadata["release_date"] == "2026-07-08"
    latest = {
        item["series_id"]: item
        for item in result.records
        if item["date"] == "2026-05-01"
    }
    assert latest["G19-CONSUMER-CREDIT-OUTSTANDING-SA"]["value"] == Decimal(
        "5154538.86"
    )
    assert latest["G19-REVOLVING-CREDIT-GROWTH-SAAR"]["value"] == Decimal("-4.71")


def test_g19_provider_fails_closed_when_required_column_is_missing():
    payload = _g19_csv().replace(
        b"Percent change of total revolving consumer credit", b"Removed revolving credit"
    )
    result = FederalReserveG19Provider(client=_g19_client(payload)).consumer_credit()

    assert not result.ok
    assert "required columns missing" in result.error


@pytest.mark.parametrize(
    ("original", "replacement"),
    [
        (b"Unit:,Percent", b"Unit:,Currency"),
        (b"Multiplier:,1,1,1", b"Multiplier:,2,1,1"),
    ],
)
def test_g19_provider_rejects_wrong_declared_unit_or_multiplier(
    original,
    replacement,
):
    payload = _g19_csv().replace(original, replacement, 1)

    result = FederalReserveG19Provider(
        client=_g19_client(payload)
    ).consumer_credit()

    assert not result.ok
    assert "source unit or multiplier is invalid" in result.error


def test_g19_provider_rejects_noncanonical_package_parameters():
    page = _g19_page().replace(b"lastObs=", b"lastObs=1", 1)

    result = FederalReserveG19Provider(
        client=_g19_client(page_payload=page)
    ).consumer_credit()

    assert not result.ok
    assert "package parameters are incomplete" in result.error


@pytest.mark.parametrize(
    ("violation", "message"),
    [
        ("header", "metadata headers are invalid"),
        ("row-width", "row width"),
        ("currency", "source currency is invalid"),
        ("started-tail-gap", "missing month after starting"),
        ("started-gap", "missing month after starting"),
        ("entirely-missing", "entirely missing required series"),
        ("non-month-data", "non-month row contains required values"),
        ("duplicate-month", "duplicated a monthly period"),
        ("month-gap", "missing month"),
    ],
)
def test_g19_provider_rejects_structural_and_history_fallbacks(violation, message):
    rows = list(csv.reader(io.StringIO(_g19_csv().decode())))
    if violation == "header":
        rows[3][0] = "Currency"
    elif violation == "row-width":
        rows[-1].pop()
    elif violation == "currency":
        rows[3][1] = "EUR"
    elif violation == "started-tail-gap":
        rows[-1][1] = ""
    elif violation == "started-gap":
        earlier = ["2026-03", *rows[6][1:]]
        rows.insert(6, earlier)
        rows[7][1] = ""
    elif violation == "entirely-missing":
        for row in rows[6:]:
            row[1] = ""
    elif violation == "non-month-data":
        rows[-1][0] = "source note"
    elif violation == "duplicate-month":
        rows[-1][0] = rows[-2][0]
    else:
        rows[6][0] = "2026-03"
    output = io.StringIO()
    csv.writer(output, lineterminator="\n").writerows(rows)

    result = FederalReserveG19Provider(
        client=_g19_client(output.getvalue().encode())
    ).consumer_credit()

    assert not result.ok
    assert message in result.error


def test_g19_provider_allows_required_series_to_start_later():
    rows = list(csv.reader(io.StringIO(_g19_csv().decode())))
    rows[6][1] = ""
    output = io.StringIO()
    csv.writer(output, lineterminator="\n").writerows(rows)

    result = FederalReserveG19Provider(
        client=_g19_client(output.getvalue().encode())
    ).consumer_credit()

    assert result.ok
    assert result.row_count == 17
    total_growth = [
        record
        for record in result.records
        if record["series_id"] == "G19-CONSUMER-CREDIT-GROWTH-SAAR"
    ]
    assert [record["date"] for record in total_growth] == ["2026-05-01"]


def test_g19_replay_rejects_observation_month_after_retrieval_date():
    result = FederalReserveG19Provider(client=_g19_client()).consumer_credit()
    assert result.ok
    evidence = parse_evidence_bundle(
        result.raw_bytes,
        expected_provider="federal-reserve-g19",
        expected_dataset="consumer-credit",
    )
    earlier_release_page = evidence.responses["choose-page"].replace(
        b"July 8, 2026",
        b"April 8, 2026",
    )
    replay_bundle = _retime_evidence_bundle(
        result.raw_bytes,
        provider="federal-reserve-g19",
        dataset="consumer-credit",
        retrieved_at="2026-04-30T12:00:00+00:00",
        payload_overrides={"choose-page": earlier_release_page},
    )

    with pytest.raises(ValueError, match="observation period is in the future"):
        FederalReserveG19Provider.replay_evidence_bundle(replay_bundle)


def test_nyfed_provider_discovers_latest_workbook_and_preserves_attribution():
    payload = _hhdc_workbook()
    result = NYFedHouseholdDebtProvider(
        client=_hhdc_client(payload)
    ).household_debt()

    assert result.ok
    assert result.row_count == 28
    assert result.metadata["latest_value_date"] == "2026-03-31"
    assert result.metadata["attribution"] == "New York Fed Consumer Credit Panel / Equifax"
    assert result.metadata["artifacts"][1]["sha256"] == hashlib.sha256(payload).hexdigest()
    evidence = parse_evidence_bundle(
        result.raw_bytes,
        expected_provider="ny-fed-household-credit",
        expected_dataset="household-debt-credit",
    )
    assert set(evidence.responses) == {
        "databank-page",
        "household-debt-workbook",
    }
    assert evidence.responses["household-debt-workbook"] == payload
    replay_records, replay_metadata = NYFedHouseholdDebtProvider.replay_evidence_bundle(
        result.raw_bytes
    )
    assert replay_records == result.records
    assert replay_metadata["latest_value_date"] == "2026-03-31"
    latest = {
        item["series_id"]: item
        for item in result.records
        if item["date"] == "2026-03-31"
    }
    assert latest["HHDC-TOTAL-DEBT-BALANCE"]["value"] == Decimal("18.794")
    assert latest["HHDC-CREDIT-CARD-90D-DELINQUENT"]["value"] == Decimal("13.12")


def test_nyfed_provider_rejects_filename_and_workbook_period_mismatch():
    result = NYFedHouseholdDebtProvider(
        client=_hhdc_client(_hhdc_workbook(), filename_quarter="2025q4")
    ).household_debt()

    assert not result.ok
    assert "does not match filename" in result.error


@pytest.mark.parametrize(
    "workbook",
    [
        _hhdc_workbook(balance_unit="USD millions"),
        _hhdc_workbook(delinquency_unit="Basis points"),
    ],
)
def test_nyfed_provider_rejects_wrong_declared_workbook_units(workbook):
    result = NYFedHouseholdDebtProvider(
        client=_hhdc_client(workbook)
    ).household_debt()

    assert not result.ok
    assert "declared unit is invalid" in result.error


@pytest.mark.parametrize(
    ("mutation", "message"),
    [
        ("duplicate-header", "duplicated a column header"),
        ("invalid-value", "required value is invalid"),
        ("started-gap", "missing quarter after starting"),
        ("entirely-missing", "entirely missing required series"),
        ("duplicate-quarter", "duplicated a quarterly period"),
        ("quarter-gap", "missing quarter"),
    ],
)
def test_nyfed_provider_rejects_structural_and_quarter_history_fallbacks(
    mutation,
    message,
):
    payload = _mutate_hhdc_workbook(_hhdc_workbook(), mutation)

    result = NYFedHouseholdDebtProvider(
        client=_hhdc_client(payload)
    ).household_debt()

    assert not result.ok
    assert message in result.error


def test_nyfed_provider_allows_required_series_to_start_later():
    payload = _mutate_hhdc_workbook(_hhdc_workbook(), "later-start")

    result = NYFedHouseholdDebtProvider(
        client=_hhdc_client(payload)
    ).household_debt()

    assert result.ok
    assert result.row_count == 27
    mortgage = [
        record
        for record in result.records
        if record["series_id"] == "HHDC-MORTGAGE-BALANCE"
    ]
    assert [record["date"] for record in mortgage] == ["2026-03-31"]


def test_nyfed_replay_rejects_quarter_end_after_retrieval_date():
    result = NYFedHouseholdDebtProvider(
        client=_hhdc_client(_hhdc_workbook())
    ).household_debt()
    assert result.ok
    replay_bundle = _retime_evidence_bundle(
        result.raw_bytes,
        provider="ny-fed-household-credit",
        dataset="household-debt-credit",
        retrieved_at="2026-03-01T12:00:00+00:00",
    )

    with pytest.raises(ValueError, match="observation period is in the future"):
        NYFedHouseholdDebtProvider.replay_evidence_bundle(replay_bundle)


@pytest.mark.django_db
def test_consumer_credit_v2_persists_private_append_only_evidence(
    settings,
    tmp_path,
):
    settings.RAW_ARTIFACT_ROOT = tmp_path / "raw-artifacts"
    g19 = FederalReserveG19Provider(client=_g19_client()).consumer_credit()
    household = NYFedHouseholdDebtProvider(
        client=_hhdc_client(_hhdc_workbook())
    ).household_debt()

    g19_run = record_provider_result(
        g19,
        persist=_store_consumer_credit_observations_v2,
    )
    household_run = record_provider_result(
        household,
        persist=_store_consumer_credit_observations_v2,
    )

    assert g19_run.status == IngestionRun.Status.SUCCESS
    assert household_run.status == IngestionRun.Status.SUCCESS
    assert RawArtifact.objects.filter(run=g19_run).count() == 1
    assert RawArtifact.objects.filter(run=household_run).count() == 1
    for run, result in ((g19_run, g19), (household_run, household)):
        artifact = RawArtifact.objects.get(run=run)
        assert artifact.uri.startswith(f"private://{run.source.key}/")
        path = (
            Path(settings.RAW_ARTIFACT_ROOT)
            / artifact.sha256[:2]
            / f"{artifact.sha256}.bin"
        )
        assert path.read_bytes() == result.raw_bytes
        assert Observation.objects.filter(batch_id=run.batch_id).count() == len(
            result.records
        )

    first_g19_rows = list(
        Observation.objects.filter(batch_id=g19_run.batch_id)
        .order_by("pk")
        .values_list("pk", "value", "updated_at")
    )
    repeated_result = FederalReserveG19Provider(
        client=_g19_client()
    ).consumer_credit()
    repeated = record_provider_result(
        repeated_result,
        persist=_store_consumer_credit_observations_v2,
    )
    assert repeated.status == IngestionRun.Status.SUCCESS
    assert repeated.batch_id != g19_run.batch_id
    assert Observation.objects.filter(batch_id=repeated.batch_id).count() == len(
        repeated_result.records
    )
    assert list(
        Observation.objects.filter(batch_id=g19_run.batch_id)
        .order_by("pk")
        .values_list("pk", "value", "updated_at")
    ) == first_g19_rows


@pytest.mark.django_db
def test_g19_signed_zero_persists_and_replays_without_masking_nonzero_tamper(
    settings,
    tmp_path,
):
    settings.RAW_ARTIFACT_ROOT = tmp_path / "raw-artifacts"
    signed_zero_csv = _g19_csv().replace(
        b"2026-04,4.87,",
        b"2026-04,-0.00000000,",
        1,
    )
    result = FederalReserveG19Provider(
        client=_g19_client(signed_zero_csv)
    ).consumer_credit()
    expected = next(
        record
        for record in result.records
        if record["series_id"] == "G19-CONSUMER-CREDIT-GROWTH-SAAR"
        and record["date"] == "2026-04-01"
    )
    assert expected["value"].is_zero() and expected["value"].is_signed()

    run = record_provider_result(
        result,
        persist=_store_consumer_credit_observations_v2,
    )
    assert run.status == IngestionRun.Status.SUCCESS, run.error
    stored = Observation.objects.get(
        batch_id=run.batch_id,
        series__key="g19-consumer-credit-growth-saar",
        value_date__date=date(2026, 4, 1),
    )
    assert stored.value == Decimal("0")
    assert _validate_run("g19", run).run.pk == run.pk

    stored.value = Decimal("0.00000001")
    stored.save(update_fields=["value", "updated_at"])
    with pytest.raises(ValueError, match="normalized rows do not match exact evidence"):
        _validate_run("g19", run)


@pytest.mark.django_db
def test_g19_signed_zero_persistence_rejects_one_database_quantum_tamper(
    monkeypatch,
    settings,
    tmp_path,
):
    settings.RAW_ARTIFACT_ROOT = tmp_path / "raw-artifacts"
    signed_zero_csv = _g19_csv().replace(
        b"2026-04,4.87,",
        b"2026-04,-0.00000000,",
        1,
    )
    original_store = official_data.store_series_observations

    def tampering_store(result, source, run, **kwargs):
        row_count = original_store(result, source, run, **kwargs)
        Observation.objects.filter(
            batch_id=run.batch_id,
            series__key="g19-consumer-credit-growth-saar",
            value_date__date=date(2026, 4, 1),
        ).update(value=Decimal("0.00000001"))
        return row_count

    monkeypatch.setattr(official_data, "store_series_observations", tampering_store)
    result = FederalReserveG19Provider(
        client=_g19_client(signed_zero_csv)
    ).consumer_credit()
    run = record_provider_result(
        result,
        persist=_store_consumer_credit_observations_v2,
    )

    assert run.status == IngestionRun.Status.FAILED
    assert "persistence postcondition failed" in run.error
    assert not Observation.objects.filter(batch_id=run.batch_id).exists()
    assert not RawArtifact.objects.filter(run=run).exists()


@pytest.mark.django_db
def test_consumer_credit_v2_rejects_normalized_and_metadata_tamper(
    settings,
    tmp_path,
):
    settings.RAW_ARTIFACT_ROOT = tmp_path / "raw-artifacts"
    g19 = FederalReserveG19Provider(client=_g19_client()).consumer_credit()
    g19.records[0]["value"] = Decimal("999")
    g19_run = record_provider_result(
        g19,
        persist=_store_consumer_credit_observations_v2,
    )
    assert g19_run.status == IngestionRun.Status.FAILED
    assert "normalized observations" in g19_run.error

    household = NYFedHouseholdDebtProvider(
        client=_hhdc_client(_hhdc_workbook())
    ).household_debt()
    household.metadata["workbook_url"] = "https://www.newyorkfed.org/forged.xlsx"
    household_run = record_provider_result(
        household,
        persist=_store_consumer_credit_observations_v2,
    )
    assert household_run.status == IngestionRun.Status.FAILED
    assert "replay metadata" in household_run.error

    assert not Observation.objects.filter(
        batch_id__in=(g19_run.batch_id, household_run.batch_id)
    )
    assert not RawArtifact.objects.filter(run__in=(g19_run, household_run))
