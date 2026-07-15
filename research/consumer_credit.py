"""Official U.S. consumer-credit and household-debt source adapters.

The Board's G.19 Data Download Program supplies monthly consumer-credit
balances and growth rates.  The New York Fed publishes the quarterly
Household Debt and Credit workbook, based on its Consumer Credit Panel /
Equifax data.  Both adapters preserve the official file fingerprint and fail
closed when the expected release structure or latest period is inconsistent.
"""

from __future__ import annotations

import calendar
import csv
import hashlib
import html
import io
import re
import zipfile
from datetime import UTC, date, datetime, timedelta
from decimal import Decimal, InvalidOperation
from pathlib import PurePosixPath
from typing import Any
from urllib.parse import parse_qsl, urlencode, urljoin, urlparse

import httpx
from openpyxl import load_workbook

from .providers import HTTPProvider, ProviderResult
from .raw_evidence import EvidenceResponse, build_evidence_bundle, parse_evidence_bundle

G19_CHOOSE_PATH = "/datadownload/Choose.aspx"
G19_OUTPUT_PATH = "/datadownload/Output.aspx"
G19_SOURCE_PAGE = "https://www.federalreserve.gov/releases/g19/current/"
G19_CHOOSE_URL = "https://www.federalreserve.gov/datadownload/Choose.aspx?rel=G19"

NYFED_DATABANK_PATH = "/microeconomics/databank.html"
NYFED_DATABANK_URL = "https://www.newyorkfed.org/microeconomics/databank.html"
NYFED_HHDC_PAGE = "https://www.newyorkfed.org/microeconomics/hhdc"

G19_SERIES = {
    "Percent change of total consumer credit, seasonally adjusted at an annual rate": (
        "G19-CONSUMER-CREDIT-GROWTH-SAAR",
        "% annual rate",
    ),
    "Percent change of total revolving consumer credit, seasonally adjusted at an annual rate": (
        "G19-REVOLVING-CREDIT-GROWTH-SAAR",
        "% annual rate",
    ),
    "Percent change of total nonrevolving consumer credit, seasonally adjusted at an annual rate": (
        "G19-NONREVOLVING-CREDIT-GROWTH-SAAR",
        "% annual rate",
    ),
    "Total consumer credit owned and securitized, seasonally adjusted level": (
        "G19-CONSUMER-CREDIT-OUTSTANDING-SA",
        "USD millions",
    ),
    "Revolving consumer credit owned and securitized, seasonally adjusted level": (
        "G19-REVOLVING-CREDIT-OUTSTANDING-SA",
        "USD millions",
    ),
    "Nonrevolving consumer credit owned and securitized, seasonally adjusted level": (
        "G19-NONREVOLVING-CREDIT-OUTSTANDING-SA",
        "USD millions",
    ),
    "Total consumer credit owned and securitized, seasonally adjusted flow, monthly rate": (
        "G19-CONSUMER-CREDIT-FLOW-SA",
        "USD millions per month",
    ),
    "Revolving consumer credit owned and securitized, seasonally adjusted flow, monthly rate": (
        "G19-REVOLVING-CREDIT-FLOW-SA",
        "USD millions per month",
    ),
    "Nonrevolving consumer credit owned and securitized, seasonally adjusted flow, monthly rate": (
        "G19-NONREVOLVING-CREDIT-FLOW-SA",
        "USD millions per month",
    ),
}

HHDC_BALANCE_SERIES = {
    "Mortgage": "HHDC-MORTGAGE-BALANCE",
    "HE Revolving": "HHDC-HELOC-BALANCE",
    "Auto Loan": "HHDC-AUTO-LOAN-BALANCE",
    "Credit Card": "HHDC-CREDIT-CARD-BALANCE",
    "Student Loan": "HHDC-STUDENT-LOAN-BALANCE",
    "Other": "HHDC-OTHER-BALANCE",
    "Total": "HHDC-TOTAL-DEBT-BALANCE",
}

HHDC_DELINQUENCY_SERIES = {
    "MORTGAGE": "HHDC-MORTGAGE-90D-DELINQUENT",
    "HELOC": "HHDC-HELOC-90D-DELINQUENT",
    "AUTO": "HHDC-AUTO-90D-DELINQUENT",
    "CC": "HHDC-CREDIT-CARD-90D-DELINQUENT",
    "STUDENT LOAN": "HHDC-STUDENT-LOAN-90D-DELINQUENT",
    "OTHER": "HHDC-OTHER-90D-DELINQUENT",
    "ALL": "HHDC-ALL-90D-DELINQUENT",
}


def _decimal(value: Any) -> Decimal | None:
    if value in (None, "", ".", "n.a."):
        return None
    try:
        parsed = Decimal(str(value))
    except (InvalidOperation, ValueError):
        return None
    return parsed if parsed.is_finite() else None


def _artifact(url: str, content: bytes, content_type: str) -> dict[str, Any]:
    return {
        "url": url,
        "sha256": hashlib.sha256(content).hexdigest(),
        "size": len(content),
        "content_type": content_type,
    }


def _retrieved_at(entry: dict[str, Any], *, context: str) -> datetime:
    witness = entry.get("response_witness")
    if not isinstance(witness, dict) or set(witness) != {"retrieved_at"}:
        raise ValueError(f"{context} retrieval witness is invalid")
    try:
        retrieved_at = datetime.fromisoformat(
            str(witness["retrieved_at"]).replace("Z", "+00:00")
        )
    except (TypeError, ValueError) as exc:
        raise ValueError(f"{context} retrieval time is invalid") from exc
    if retrieved_at.tzinfo is None:
        raise ValueError(f"{context} retrieval time lacks a timezone")
    retrieved_at = retrieved_at.astimezone(UTC)
    if retrieved_at > datetime.now(UTC) + timedelta(minutes=5):
        raise ValueError(f"{context} retrieval time is in the future")
    return retrieved_at


class FederalReserveG19Provider(HTTPProvider):
    """Monthly seasonally adjusted G.19 history from the Board DDP CSV."""

    key = "federal-reserve-g19"
    base_url = "https://www.federalreserve.gov"

    def __init__(
        self,
        *,
        client: httpx.Client | None = None,
        timeout: float = 60.0,
        max_html_bytes: int = 2 * 1024 * 1024,
        max_csv_bytes: int = 4 * 1024 * 1024,
    ) -> None:
        self.max_html_bytes = max_html_bytes
        self.max_csv_bytes = max_csv_bytes
        super().__init__(
            client=client,
            timeout=timeout,
            headers={
                "Accept": "text/html,text/csv",
                "User-Agent": "AtlasMacro/0.1 G19 data downloader",
            },
        )

    def consumer_credit(self) -> ProviderResult:
        dataset = "consumer-credit"
        try:
            choose = self.client.get(G19_CHOOSE_PATH, params={"rel": "G19"})
            choose.raise_for_status()
            choose_retrieved_at = datetime.now(UTC)
            choose_bytes = choose.content
            if not choose_bytes or len(choose_bytes) > self.max_html_bytes:
                raise ValueError("G.19 package page exceeded configured size or was empty")
            choose_url = str(choose.url)
            choose_type = choose.headers.get("content-type", "").split(";", 1)[0].lower()
            if choose_url != G19_CHOOSE_URL or choose_type != "text/html":
                raise ValueError("G.19 package page response identity is invalid")
            choose_text = choose_bytes.decode("utf-8")
            package_params = self._package_params(choose_text)

            output_url = self._canonical_output_url(package_params)
            response = self.client.get(output_url)
            response.raise_for_status()
            output_retrieved_at = datetime.now(UTC)
            csv_bytes = response.content
            if not csv_bytes or len(csv_bytes) > self.max_csv_bytes:
                raise ValueError("G.19 CSV exceeded configured size or was empty")
            output_type = response.headers.get("content-type", "").split(";", 1)[0].lower()
            if str(response.url) != output_url or output_type not in {
                "text/csv",
                "application/csv",
                "application/octet-stream",
            }:
                raise ValueError("G.19 CSV response identity is invalid")
            raw_bundle, bundle_metadata = build_evidence_bundle(
                provider=self.key,
                dataset=dataset,
                responses=(
                    EvidenceResponse(
                        role="choose-page",
                        url=G19_CHOOSE_URL,
                        content_type=choose_type,
                        raw_bytes=choose_bytes,
                        request_witness={"rel": "G19"},
                        response_witness={
                            "retrieved_at": choose_retrieved_at.isoformat(),
                        },
                    ),
                    EvidenceResponse(
                        role="output-csv",
                        url=output_url,
                        content_type=output_type,
                        raw_bytes=csv_bytes,
                        request_witness={
                            "discovered_from": "choose-page",
                            "params": package_params,
                        },
                        response_witness={
                            "retrieved_at": output_retrieved_at.isoformat(),
                        },
                    ),
                ),
            )
            records, replay_metadata = self.replay_evidence_bundle(raw_bundle)
        except (httpx.HTTPError, UnicodeError, csv.Error, ValueError) as exc:
            return ProviderResult.failure(self.key, dataset, f"{type(exc).__name__}: {exc}")

        return ProviderResult(
            provider=self.key,
            dataset=dataset,
            records=records,
            fetched_at=output_retrieved_at,
            raw_bytes=raw_bundle,
            metadata={
                **bundle_metadata,
                **replay_metadata,
                "artifacts": [
                    _artifact(G19_CHOOSE_URL, choose_bytes, choose_type),
                    _artifact(output_url, csv_bytes, output_type),
                ],
            },
        )

    @staticmethod
    def _package_params(page: str) -> dict[str, str]:
        match = re.search(
            r'name="FreqRequest"\s+value="([^"]+)"[^>]*>'
            r'<label[^>]*>Consumer Credit Outstanding \(S\.A\.\)',
            page,
            flags=re.IGNORECASE,
        )
        if match is None:
            raise ValueError("G.19 seasonally adjusted package link not found")
        pairs = parse_qsl(html.unescape(match.group(1)), keep_blank_values=True)
        expected = {
            "rel",
            "series",
            "lastObs",
            "from",
            "to",
            "filetype",
            "label",
            "layout",
            "type",
        }
        if len(pairs) != len(expected) or {key for key, _value in pairs} != expected:
            raise ValueError("G.19 package parameters are duplicated or incomplete")
        params = dict(pairs)
        if (
            params.get("rel") != "G19"
            or not str(params.get("series") or "").strip()
            or params.get("lastObs") != ""
            or params.get("from") != ""
            or params.get("to") != ""
            or params.get("filetype") != "csv"
            or params.get("label") != "include"
            or params.get("layout") != "seriescolumn"
            or params.get("type") != "package"
        ):
            raise ValueError("G.19 package parameters are incomplete")
        return params

    @staticmethod
    def _canonical_output_url(params: dict[str, str]) -> str:
        return (
            "https://www.federalreserve.gov"
            f"{G19_OUTPUT_PATH}?{urlencode(sorted(params.items()))}"
        )

    @classmethod
    def replay_evidence_bundle(
        cls,
        raw_bytes: bytes,
    ) -> tuple[list[dict[str, Any]], dict[str, Any]]:
        evidence = parse_evidence_bundle(
            raw_bytes,
            expected_provider=cls.key,
            expected_dataset="consumer-credit",
        )
        if set(evidence.responses) != {"choose-page", "output-csv"}:
            raise ValueError("G.19 evidence roles are incomplete")
        entries = {
            item["role"]: item for item in evidence.manifest["responses"]
        }
        choose_entry = entries["choose-page"]
        output_entry = entries["output-csv"]
        if (
            choose_entry["url"] != G19_CHOOSE_URL
            or choose_entry["content_type"] != "text/html"
            or choose_entry["request_witness"] != {"rel": "G19"}
            or output_entry["content_type"]
            not in {"text/csv", "application/csv", "application/octet-stream"}
        ):
            raise ValueError("G.19 evidence response contract is invalid")
        choose_retrieved_at = _retrieved_at(
            choose_entry,
            context="G.19 choose page",
        )
        output_retrieved_at = _retrieved_at(
            output_entry,
            context="G.19 output CSV",
        )
        if output_retrieved_at < choose_retrieved_at:
            raise ValueError("G.19 response retrieval chronology is invalid")
        choose_text = evidence.responses["choose-page"].decode("utf-8")
        package_params = cls._package_params(choose_text)
        release_date = cls._release_date(choose_text)
        output_url = cls._canonical_output_url(package_params)
        if (
            output_entry["url"] != output_url
            or output_entry["request_witness"]
            != {
                "discovered_from": "choose-page",
                "params": package_params,
            }
        ):
            raise ValueError("G.19 output request does not replay from package page")
        records, latest_period = cls._parse_csv(evidence.responses["output-csv"])
        if release_date > output_retrieved_at.date():
            raise ValueError("G.19 official release date is in the future")
        if date.fromisoformat(latest_period) > output_retrieved_at.date():
            raise ValueError("G.19 observation period is in the future")
        for record in records:
            record["metadata"]["source_revision_date"] = release_date.isoformat()
            record["metadata"]["release_freshness_days"] = 45
        return records, {
            "source_page": G19_SOURCE_PAGE,
            "release_date": release_date.isoformat(),
            "latest_value_date": latest_period,
            "frequency": "monthly",
            "seasonal_adjustment": "seasonally adjusted",
            "quality_status": "complete",
            "retrieved_at": output_retrieved_at.isoformat(),
            "response_retrieved_at": {
                "choose-page": choose_retrieved_at.isoformat(),
                "output-csv": output_retrieved_at.isoformat(),
            },
        }

    @staticmethod
    def _release_date(page: str) -> date:
        match = re.search(r"last released\s+([A-Za-z]+,?\s+[A-Za-z]+\s+\d{1,2},\s+\d{4})", page)
        if match is None:
            raise ValueError("G.19 release date not found")
        raw = match.group(1).replace(",", "", 1)
        try:
            return datetime.strptime(raw, "%A %B %d, %Y").date()
        except ValueError as exc:
            raise ValueError("G.19 release date is invalid") from exc

    @staticmethod
    def _parse_csv(payload: bytes) -> tuple[list[dict[str, Any]], str]:
        text = payload.decode("utf-8-sig")
        rows = list(csv.reader(io.StringIO(text)))
        expected_headers = (
            "Series Description",
            "Unit:",
            "Multiplier:",
            "Currency:",
            "Unique Identifier:",
            "Time Period",
        )
        if len(rows) < 7 or any(not row for row in rows[:6]):
            raise ValueError("G.19 CSV header is missing")
        if tuple(row[0] for row in rows[:6]) != expected_headers:
            raise ValueError("G.19 CSV metadata headers are invalid")
        column_count = len(rows[0])
        if column_count < 2 or any(len(row) != column_count for row in rows[:6]):
            raise ValueError("G.19 CSV metadata row width is inconsistent")
        descriptions = rows[0][1:]
        units = rows[1][1:]
        multipliers = rows[2][1:]
        currencies = rows[3][1:]
        identifiers = rows[4][1:]
        series_headers = rows[5][1:]
        metadata_columns = (
            descriptions,
            units,
            multipliers,
            currencies,
            identifiers,
            series_headers,
        )
        if any(
            len(values) != len(descriptions)
            or any(not str(value).strip() for value in values)
            for values in metadata_columns
        ):
            raise ValueError("G.19 CSV metadata columns must be complete and non-empty")
        if any(
            len(values) != len(set(values))
            for values in (descriptions, identifiers, series_headers)
        ):
            raise ValueError("G.19 CSV duplicated a description, identifier, or series header")
        missing_descriptions = sorted(set(G19_SERIES) - set(descriptions))
        if missing_descriptions:
            raise ValueError(f"G.19 required columns missing: {', '.join(missing_descriptions)}")
        for description, (_series_id, normalized_unit) in G19_SERIES.items():
            index = descriptions.index(description)
            expected_source_unit = (
                "percent" if normalized_unit.startswith("%") else "currency"
            )
            expected_multiplier = (
                Decimal("1")
                if expected_source_unit == "percent"
                else Decimal("1000000")
            )
            if (
                str(units[index]).strip().casefold() != expected_source_unit
                or _decimal(multipliers[index]) != expected_multiplier
            ):
                raise ValueError(
                    f"G.19 source unit or multiplier is invalid for {description}"
                )
            if str(currencies[index]).strip().upper() != "USD":
                raise ValueError(f"G.19 source currency is invalid for {description}")

        required_columns = {
            description: (descriptions.index(description) + 1, target)
            for description, target in G19_SERIES.items()
        }
        period_rows: dict[date, list[str]] = {}
        for row in rows[6:]:
            if not row or not re.fullmatch(r"\d{4}-\d{2}", row[0].strip()):
                if any(
                    index < len(row) and str(row[index]).strip()
                    for index, _target in required_columns.values()
                ):
                    raise ValueError("G.19 CSV non-month row contains required values")
                continue
            if len(row) != column_count:
                raise ValueError("G.19 CSV data row width does not match its headers")
            year, month = (int(part) for part in row[0].strip().split("-", 1))
            try:
                period = date(year, month, 1)
            except ValueError as exc:
                raise ValueError("G.19 CSV monthly period is invalid") from exc
            if period in period_rows:
                raise ValueError("G.19 CSV duplicated a monthly period")
            period_rows[period] = row
        ordered_periods = sorted(period_rows)
        for previous, current in zip(ordered_periods, ordered_periods[1:], strict=False):
            expected = (
                date(previous.year + 1, 1, 1)
                if previous.month == 12
                else date(previous.year, previous.month + 1, 1)
            )
            if current != expected:
                raise ValueError("G.19 CSV history contains a missing month")
        records: list[dict[str, Any]] = []
        latest_by_series: dict[str, str] = {}
        started_series: set[str] = set()
        missing_markers = {"", ".", "n.a."}
        for period in ordered_periods:
            value_date = period.isoformat()
            row = period_rows[period]
            for description, (index, target) in required_columns.items():
                raw_value = row[index]
                value = _decimal(raw_value)
                series_id, normalized_unit = target
                if value is None:
                    normalized_missing = str(raw_value).strip().casefold()
                    if normalized_missing not in missing_markers:
                        raise ValueError(
                            f"G.19 required value is invalid for {description} "
                            f"at {value_date}"
                        )
                    if series_id in started_series:
                        raise ValueError(
                            f"G.19 series has a missing month after starting: "
                            f"{description} at {value_date}"
                        )
                    continue
                started_series.add(series_id)
                latest_by_series[series_id] = value_date
                records.append(
                    {
                        "series_id": series_id,
                        "date": value_date,
                        "value": value,
                        "metadata": {
                            "source_series_id": identifiers[index - 1],
                            "description": description,
                            "source_unit": units[index - 1],
                            "source_multiplier": multipliers[index - 1],
                            "source_currency": currencies[index - 1],
                            "unit": normalized_unit,
                            "frequency": "monthly",
                            "seasonal_adjustment": "seasonally adjusted",
                            "official_source_url": G19_SOURCE_PAGE,
                        },
                    }
                )
        entirely_missing = sorted(
            {item[0] for item in G19_SERIES.values()} - started_series
        )
        if entirely_missing:
            raise ValueError(
                "G.19 CSV has entirely missing required series: "
                f"{', '.join(entirely_missing)}"
            )
        if not records or set(latest_by_series) != {item[0] for item in G19_SERIES.values()}:
            raise ValueError("G.19 CSV did not provide complete required history")
        latest_periods = set(latest_by_series.values())
        if len(latest_periods) != 1:
            raise ValueError("G.19 required series have inconsistent latest months")
        return records, latest_periods.pop()


class NYFedHouseholdDebtProvider(HTTPProvider):
    """Quarterly national household-debt balances and delinquency rates."""

    key = "ny-fed-household-credit"
    base_url = "https://www.newyorkfed.org"

    def __init__(
        self,
        *,
        client: httpx.Client | None = None,
        timeout: float = 90.0,
        max_html_bytes: int = 4 * 1024 * 1024,
        max_workbook_bytes: int = 8 * 1024 * 1024,
        max_expanded_bytes: int = 64 * 1024 * 1024,
    ) -> None:
        self.max_html_bytes = max_html_bytes
        self.max_workbook_bytes = max_workbook_bytes
        self.max_expanded_bytes = max_expanded_bytes
        super().__init__(
            client=client,
            timeout=timeout,
            headers={
                "Accept": "text/html,application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                "User-Agent": "AtlasMacro/0.1 NY Fed household debt downloader",
            },
        )

    def household_debt(self) -> ProviderResult:
        dataset = "household-debt-credit"
        try:
            bank = self.client.get(NYFED_DATABANK_PATH)
            bank.raise_for_status()
            bank_retrieved_at = datetime.now(UTC)
            bank_bytes = bank.content
            if not bank_bytes or len(bank_bytes) > self.max_html_bytes:
                raise ValueError("NY Fed data bank page exceeded configured size or was empty")
            bank_url = str(bank.url)
            bank_type = bank.headers.get("content-type", "").split(";", 1)[0].lower()
            if bank_url != NYFED_DATABANK_URL or bank_type != "text/html":
                raise ValueError("NY Fed data bank response identity is invalid")
            bank_text = bank_bytes.decode("utf-8")
            workbook_url, expected_period = self._latest_workbook(bank_text)

            response = self.client.get(workbook_url)
            response.raise_for_status()
            workbook_retrieved_at = datetime.now(UTC)
            workbook_bytes = response.content
            if not workbook_bytes or len(workbook_bytes) > self.max_workbook_bytes:
                raise ValueError("NY Fed household debt workbook exceeded configured size or was empty")
            workbook_type = (
                response.headers.get("content-type", "").split(";", 1)[0].lower()
            )
            if (
                str(response.url) != workbook_url
                or workbook_type
                not in {
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    "application/octet-stream",
                    "application/zip",
                }
            ):
                raise ValueError("NY Fed household debt workbook response identity is invalid")
            raw_bundle, bundle_metadata = build_evidence_bundle(
                provider=self.key,
                dataset=dataset,
                responses=(
                    EvidenceResponse(
                        role="databank-page",
                        url=NYFED_DATABANK_URL,
                        content_type=bank_type,
                        raw_bytes=bank_bytes,
                        response_witness={
                            "retrieved_at": bank_retrieved_at.isoformat(),
                        },
                    ),
                    EvidenceResponse(
                        role="household-debt-workbook",
                        url=workbook_url,
                        content_type=workbook_type,
                        raw_bytes=workbook_bytes,
                        request_witness={
                            "discovered_from": "databank-page",
                            "expected_period": expected_period,
                        },
                        response_witness={
                            "retrieved_at": workbook_retrieved_at.isoformat(),
                        },
                    ),
                ),
            )
            records, replay_metadata = self.replay_evidence_bundle(raw_bundle)
        except (httpx.HTTPError, UnicodeError, OSError, ValueError, zipfile.BadZipFile) as exc:
            return ProviderResult.failure(self.key, dataset, f"{type(exc).__name__}: {exc}")

        return ProviderResult(
            provider=self.key,
            dataset=dataset,
            records=records,
            fetched_at=workbook_retrieved_at,
            raw_bytes=raw_bundle,
            metadata={
                **bundle_metadata,
                **replay_metadata,
                "artifacts": [
                    _artifact(NYFED_DATABANK_URL, bank_bytes, bank_type),
                    _artifact(
                        workbook_url,
                        workbook_bytes,
                        workbook_type,
                    ),
                ],
            },
        )

    @classmethod
    def replay_evidence_bundle(
        cls,
        raw_bytes: bytes,
    ) -> tuple[list[dict[str, Any]], dict[str, Any]]:
        evidence = parse_evidence_bundle(
            raw_bytes,
            expected_provider=cls.key,
            expected_dataset="household-debt-credit",
        )
        if set(evidence.responses) != {
            "databank-page",
            "household-debt-workbook",
        }:
            raise ValueError("NY Fed household debt evidence roles are incomplete")
        entries = {
            item["role"]: item for item in evidence.manifest["responses"]
        }
        bank_entry = entries["databank-page"]
        workbook_entry = entries["household-debt-workbook"]
        allowed_workbook_types = {
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            "application/octet-stream",
            "application/zip",
        }
        if (
            bank_entry["url"] != NYFED_DATABANK_URL
            or bank_entry["content_type"] != "text/html"
            or bank_entry["request_witness"] != {}
            or workbook_entry["content_type"] not in allowed_workbook_types
        ):
            raise ValueError("NY Fed household debt evidence response contract is invalid")
        bank_retrieved_at = _retrieved_at(
            bank_entry,
            context="NY Fed data bank page",
        )
        workbook_retrieved_at = _retrieved_at(
            workbook_entry,
            context="NY Fed household debt workbook",
        )
        if workbook_retrieved_at < bank_retrieved_at:
            raise ValueError("NY Fed household debt retrieval chronology is invalid")
        bank_text = evidence.responses["databank-page"].decode("utf-8")
        workbook_url, expected_period = cls._latest_workbook(bank_text)
        if (
            workbook_entry["url"] != workbook_url
            or workbook_entry["request_witness"]
            != {
                "discovered_from": "databank-page",
                "expected_period": expected_period,
            }
        ):
            raise ValueError("NY Fed workbook request does not replay from data bank")
        validator = cls()
        try:
            workbook_bytes = evidence.responses["household-debt-workbook"]
            validator._validate_archive(workbook_bytes)
            records, latest_period, workbook_title = cls._parse_workbook(workbook_bytes)
        finally:
            validator.close()
        if latest_period != expected_period:
            raise ValueError(
                f"NY Fed workbook latest period {latest_period} does not match filename {expected_period}"
            )
        if date.fromisoformat(latest_period) > workbook_retrieved_at.date():
            raise ValueError("NY Fed household debt observation period is in the future")
        return records, {
            "source_page": NYFED_HHDC_PAGE,
            "data_bank_url": NYFED_DATABANK_URL,
            "workbook_url": workbook_url,
            "workbook_title": workbook_title,
            "latest_value_date": latest_period,
            "frequency": "quarterly",
            "quality_status": "complete",
            "attribution": "New York Fed Consumer Credit Panel / Equifax",
            "retrieved_at": workbook_retrieved_at.isoformat(),
            "response_retrieved_at": {
                "databank-page": bank_retrieved_at.isoformat(),
                "household-debt-workbook": workbook_retrieved_at.isoformat(),
            },
        }

    @staticmethod
    def _latest_workbook(page: str) -> tuple[str, str]:
        candidates: dict[tuple[int, int], set[str]] = {}
        pattern = re.compile(
            r'href=["\']([^"\']*hhd_c_report_(\d{4})q([1-4])\.xlsx)["\']',
            flags=re.IGNORECASE,
        )
        for match in pattern.finditer(page):
            year, quarter = int(match.group(2)), int(match.group(3))
            absolute = urljoin(NYFED_DATABANK_URL, html.unescape(match.group(1)))
            parsed = urlparse(absolute)
            if (
                parsed.scheme != "https"
                or parsed.netloc.lower() != "www.newyorkfed.org"
                or parsed.username
                or parsed.password
                or parsed.query
                or parsed.fragment
                or not re.fullmatch(
                    r"/.*/hhd_c_report_\d{4}q[1-4]\.xlsx",
                    parsed.path,
                    flags=re.IGNORECASE,
                )
            ):
                raise ValueError("NY Fed household debt workbook URL is invalid")
            candidates.setdefault((year, quarter), set()).add(absolute)
        if not candidates:
            raise ValueError("NY Fed household debt workbook link not found")
        year, quarter = max(candidates)
        latest_urls = candidates[(year, quarter)]
        if len(latest_urls) != 1:
            raise ValueError("NY Fed latest household debt workbook link is ambiguous")
        return latest_urls.pop(), _quarter_end(year, quarter).isoformat()

    def _validate_archive(self, payload: bytes) -> None:
        if not payload.startswith(b"PK"):
            raise ValueError("NY Fed household debt response is not an XLSX workbook")
        with zipfile.ZipFile(io.BytesIO(payload)) as archive:
            members = archive.infolist()
            names = [member.filename for member in members]
            if len(members) > 500 or len(names) != len(set(names)):
                raise ValueError("NY Fed workbook contains too many archive members")
            if any(
                member.flag_bits & 0x1
                or member.file_size > self.max_expanded_bytes
                or member.filename.startswith("/")
                or "\\" in member.filename
                or ".." in PurePosixPath(member.filename).parts
                for member in members
            ):
                raise ValueError("NY Fed workbook archive member is unsafe")
            expanded = sum(member.file_size for member in members)
            if expanded > self.max_expanded_bytes:
                raise ValueError("NY Fed workbook exceeds configured expanded-size limit")
            if not {"[Content_Types].xml", "xl/workbook.xml"} <= set(names):
                raise ValueError("NY Fed workbook archive is missing required members")

    @staticmethod
    def _parse_workbook(payload: bytes) -> tuple[list[dict[str, Any]], str, str]:
        try:
            workbook = load_workbook(io.BytesIO(payload), read_only=True, data_only=True)
        except Exception as exc:
            raise ValueError(f"NY Fed workbook could not be parsed: {exc}") from exc
        try:
            required_sheets = {"TABLE OF CONTENTS", "Page 3 Data", "Page 12 Data"}
            missing_sheets = sorted(required_sheets - set(workbook.sheetnames))
            if missing_sheets:
                raise ValueError(f"NY Fed workbook sheets missing: {', '.join(missing_sheets)}")
            title = str(workbook["TABLE OF CONTENTS"].cell(2, 2).value or "").strip()
            if "HOUSEHOLD DEBT AND CREDIT" not in title.upper():
                raise ValueError("NY Fed workbook title is invalid")
            balance_unit = str(workbook["Page 3 Data"].cell(2, 1).value or "").strip()
            delinquency_unit = str(
                workbook["Page 12 Data"].cell(2, 1).value or ""
            ).strip()
            normalized_balance_unit = " ".join(balance_unit.casefold().split())
            normalized_delinquency_unit = " ".join(
                delinquency_unit.casefold().split()
            )
            if re.fullmatch(
                r"trillions? of (?:\$|(?:u\.?s\.? )?dollars?)",
                normalized_balance_unit,
            ) is None:
                raise ValueError("NY Fed balance sheet declared unit is invalid")
            if normalized_delinquency_unit not in {"percent", "percentage"}:
                raise ValueError("NY Fed delinquency sheet declared unit is invalid")
            balance_records, balance_latest = _parse_hhdc_sheet(
                workbook["Page 3 Data"],
                HHDC_BALANCE_SERIES,
                unit="USD trillions",
                chart_name="Total Debt Balance and Its Composition",
            )
            delinquency_records, delinquency_latest = _parse_hhdc_sheet(
                workbook["Page 12 Data"],
                HHDC_DELINQUENCY_SERIES,
                unit="% of balance 90+ days delinquent",
                chart_name="Percent of Balance 90+ Days Delinquent by Loan Type",
            )
        finally:
            workbook.close()
        if balance_latest != delinquency_latest:
            raise ValueError("NY Fed balance and delinquency sheets have different latest periods")
        return [*balance_records, *delinquency_records], balance_latest, title


def _quarter_end(year: int, quarter: int) -> date:
    month = quarter * 3
    return date(year, month, calendar.monthrange(year, month)[1])


def _quarter_label(value: Any) -> tuple[str, str] | None:
    if isinstance(value, datetime):
        quarter = (value.month - 1) // 3 + 1
        end = _quarter_end(value.year, quarter)
        return end.isoformat(), f"{value.year}:Q{quarter}"
    match = re.fullmatch(r"(\d{2,4}):Q([1-4])", str(value or "").strip(), re.IGNORECASE)
    if match is None:
        return None
    raw_year = int(match.group(1))
    year = raw_year if raw_year >= 100 else (1900 + raw_year if raw_year >= 90 else 2000 + raw_year)
    quarter = int(match.group(2))
    return _quarter_end(year, quarter).isoformat(), f"{year}:Q{quarter}"


def _parse_hhdc_sheet(
    sheet: Any,
    series_map: dict[str, str],
    *,
    unit: str,
    chart_name: str,
) -> tuple[list[dict[str, Any]], str]:
    headers = [str(cell.value or "").strip() for cell in sheet[4]]
    named_headers = [header for header in headers[1:] if header]
    duplicate_headers = sorted(
        header for header in set(named_headers) if named_headers.count(header) > 1
    )
    if duplicate_headers:
        raise ValueError(
            f"NY Fed {chart_name} duplicated a column header: {', '.join(duplicate_headers)}"
        )
    found_headers = set(headers)
    missing_headers = sorted(set(series_map) - found_headers)
    if missing_headers:
        raise ValueError(f"NY Fed {chart_name} columns missing: {', '.join(missing_headers)}")
    positions = {name: headers.index(name) for name in series_map}
    period_rows: dict[str, tuple[str, tuple[Any, ...]]] = {}
    for row in sheet.iter_rows(min_row=5, values_only=True):
        period = _quarter_label(row[0] if row else None)
        if period is None:
            continue
        value_date, period_label = period
        if value_date in period_rows:
            raise ValueError(f"NY Fed {chart_name} duplicated a quarterly period")
        period_rows[value_date] = (period_label, row)
    ordered_periods = sorted(date.fromisoformat(value) for value in period_rows)
    for previous, current in zip(ordered_periods, ordered_periods[1:], strict=False):
        previous_quarter = (previous.month - 1) // 3 + 1
        next_year = previous.year + 1 if previous_quarter == 4 else previous.year
        next_quarter = 1 if previous_quarter == 4 else previous_quarter + 1
        if current != _quarter_end(next_year, next_quarter):
            raise ValueError(f"NY Fed {chart_name} history contains a missing quarter")
    records: list[dict[str, Any]] = []
    latest_by_series: dict[str, str] = {}
    started_series: set[str] = set()
    missing_markers = {None, "", ".", "n.a."}
    for period in ordered_periods:
        value_date = period.isoformat()
        period_label, row = period_rows[value_date]
        for name, series_id in series_map.items():
            index = positions[name]
            raw_value = row[index] if index < len(row) else None
            value = _decimal(raw_value)
            if value is None:
                normalized_missing = (
                    raw_value
                    if raw_value is None
                    else str(raw_value).strip().casefold()
                )
                if normalized_missing not in missing_markers:
                    raise ValueError(
                        f"NY Fed {chart_name} required value is invalid for "
                        f"{name} at {period_label}"
                    )
                if series_id in started_series:
                    raise ValueError(
                        f"NY Fed {chart_name} series has a missing quarter after "
                        f"starting: {name} at {period_label}"
                    )
                continue
            started_series.add(series_id)
            latest_by_series[series_id] = value_date
            records.append(
                {
                    "series_id": series_id,
                    "date": value_date,
                    "value": value,
                    "metadata": {
                        "description": name,
                        "chart_name": chart_name,
                        "period_label": period_label,
                        "unit": unit,
                        "frequency": "quarterly",
                        "official_source_url": NYFED_HHDC_PAGE,
                        "attribution": "New York Fed Consumer Credit Panel / Equifax",
                    },
                }
            )
    entirely_missing = sorted(set(series_map.values()) - started_series)
    if entirely_missing:
        raise ValueError(
            f"NY Fed {chart_name} has entirely missing required series: "
            f"{', '.join(entirely_missing)}"
        )
    if not records or set(latest_by_series) != set(series_map.values()):
        raise ValueError(f"NY Fed {chart_name} did not provide complete required history")
    latest_periods = set(latest_by_series.values())
    if len(latest_periods) != 1:
        raise ValueError(f"NY Fed {chart_name} required series have inconsistent latest periods")
    return records, latest_periods.pop()
