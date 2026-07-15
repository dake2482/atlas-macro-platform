from __future__ import annotations

import json
from datetime import UTC, date, datetime
from decimal import Decimal
from io import BytesIO
from pathlib import Path

import httpx
import pytest
from openpyxl import Workbook

from research.macro_official import BEANIPAProvider, CensusMARTSProvider
from research.macro_releases import (
    CENSUS_MARTS_CURRENT_WORKBOOK,
    XLSX_CONTENT_TYPE,
    CensusMARTSReleaseProvider,
)
from research.models import Observation, RawArtifact, SeriesDefinition
from research.official_data import (
    _store_census_marts_observations_v2,
    publish_official_dashboards,
)
from research.providers import ProviderResult
from research.raw_evidence import (
    EvidenceResponse,
    build_evidence_bundle,
    parse_evidence_bundle,
)
from research.services import record_provider_result, store_series_observations


def _client(handler):
    return httpx.Client(base_url="https://example.test", transport=httpx.MockTransport(handler))


def test_bea_missing_key_skips_without_network(monkeypatch):
    monkeypatch.delenv("BEA_API_KEY", raising=False)

    def handler(_request):
        pytest.fail("missing BEA_API_KEY must not make an HTTP request")

    result = BEANIPAProvider(client=_client(handler)).gdp_pce(years=2026)

    assert result.skipped
    assert not result.ok
    assert result.metadata["reason"] == "BEA_API_KEY is not configured"


def test_bea_nipa_gdp_pce_uses_documented_parameters_and_preserves_revision_and_units():
    def handler(request):
        assert request.url.path == "/api/data"
        assert request.url.params["UserID"] == "bea-test-key"
        assert request.url.params["method"] == "GetData"
        assert request.url.params["DataSetName"] == "NIPA"
        assert request.url.params["TableName"] == "T10101"
        assert request.url.params["Frequency"] == "Q"
        assert request.url.params["Year"] == "2025,2026"
        assert request.url.params["ResultFormat"] == "JSON"
        return httpx.Response(
            200,
            json={
                "BEAAPI": {
                    "Results": {
                        "Statistic": "NIPA Table",
                        "UTCProductionTime": "2026-06-25T12:35:00.000",
                        "Data": [
                            {
                                "TableName": "T10101",
                                "SeriesCode": "A191RL",
                                "LineNumber": "1",
                                "LineDescription": "Gross domestic product",
                                "TimePeriod": "2026Q1",
                                "METRIC_NAME": "Percent Change From Preceding Period",
                                "CL_UNIT": "Percent change",
                                "UNIT_MULT": "0",
                                "DataValue": "2.1",
                                "NoteRef": "T10101",
                            },
                            {
                                "TableName": "T10101",
                                "SeriesCode": "DPCERL",
                                "LineNumber": "2",
                                "LineDescription": "Personal consumption expenditures",
                                "TimePeriod": "2026Q1",
                                "METRIC_NAME": "Percent Change From Preceding Period",
                                "CL_UNIT": "Percent change",
                                "UNIT_MULT": "0",
                                "DataValue": "0.5",
                                "NoteRef": "T10101",
                            },
                            {
                                "TableName": "T10101",
                                "SeriesCode": "A006RL",
                                "LineNumber": "3",
                                "LineDescription": "Goods",
                                "TimePeriod": "2026Q1",
                                "METRIC_NAME": "Percent Change From Preceding Period",
                                "CL_UNIT": "Percent change",
                                "UNIT_MULT": "0",
                                "DataValue": "-1.0",
                                "NoteRef": "T10101",
                            },
                        ],
                        "Notes": [
                            {
                                "NoteRef": "T10101",
                                "NoteText": (
                                    "Table 1.1.1. Percent Change From Preceding Period in "
                                    "Real Gross Domestic Product - LastRevised: June 25, 2026"
                                ),
                            }
                        ],
                    }
                }
            },
        )

    result = BEANIPAProvider(api_key="bea-test-key", client=_client(handler)).gdp_pce(
        years=(2025, 2026)
    )

    assert result.ok
    assert result.dataset == "nipa:gdp-pce-growth:Q"
    assert result.row_count == 2
    assert result.records[0]["series_id"] == "BEA-A191RL"
    assert result.records[0]["date"] == "2026-01-01"
    assert result.records[0]["value"] == Decimal("2.1")
    assert result.records[0]["metadata"]["calculation_type"] == "Percent change"
    assert result.records[0]["metadata"]["unit_multiplier"] == "0"
    assert result.records[0]["metadata"]["source_revision_date"] == "2026-06-25"
    assert result.metadata["api_production_time"] == "2026-06-25T12:35:00.000"
    assert result.metadata["vintage_policy"] == "latest-vintage-only"
    assert "not endorsed or certified" in result.metadata["attribution_notice"]


def test_bea_api_error_becomes_provider_failure():
    def handler(_request):
        return httpx.Response(
            200,
            json={
                "BEAAPI": {
                    "Results": {
                        "Error": {
                            "APIErrorCode": "4",
                            "APIErrorDescription": "Invalid Request - Invalid Parameter",
                        }
                    }
                }
            },
        )

    result = BEANIPAProvider(api_key="bad-key", client=_client(handler)).nipa_table(
        "T10101", frequency="Q", years=2026
    )

    assert not result.ok
    assert not result.skipped
    assert "Invalid Parameter" in result.error


def test_census_missing_key_skips_without_network(monkeypatch):
    monkeypatch.delenv("CENSUS_API_KEY", raising=False)

    def handler(_request):
        pytest.fail("missing CENSUS_API_KEY must not make an HTTP request")

    result = CensusMARTSProvider(client=_client(handler)).monthly_retail_sales(
        time="2026-05"
    )

    assert result.skipped
    assert not result.ok
    assert result.metadata["reason"] == "CENSUS_API_KEY is not configured"


def _census_history_payload() -> list[list[str]]:
    headers = [
        "program_code",
        "cell_value",
        "time_slot_id",
        "time_slot_date",
        "time_slot_name",
        "error_data",
        "seasonally_adj",
        "category_code",
        "data_type_code",
        "time",
    ]
    rows = [headers]
    year, month, index = 1992, 1, 0
    overrides = {
        (2025, 5): "714568",
        (2026, 3): "754013",
        (2026, 4): "757036",
        (2026, 5): "763705",
    }
    while (year, month) <= (2026, 5):
        value = overrides.get((year, month), str(100000 + index))
        rows.append(
            [
                "MARTS",
                value,
                f"M{month:02d}",
                f"{year:04d}-{month:02d}",
                f"{year:04d}-{month:02d}",
                "no",
                "yes",
                "44X72",
                "SM",
                f"{year:04d}-{month:02d}",
            ]
        )
        index += 1
        if month == 12:
            year, month = year + 1, 1
        else:
            month += 1
    return rows


def _census_release_workbook() -> bytes:
    workbook = Workbook()
    sales = workbook.active
    sales.title = "Table 1."
    sales.cell(
        4,
        1,
        "(Total sales estimates are shown in millions of dollars and are based on official survey data.)",
    )
    sales.cell(6, 10, "Adjusted2")
    sales.cell(7, 10, 2026)
    sales.cell(7, 13, 2025)
    for column, label in enumerate(
        ("May.3", "Apr.", "Mar.", "May.", "Apr."), start=10
    ):
        sales.cell(8, column, label)
    for column, status in enumerate(
        ("(a)", "(r)", "(r)", "(r)", "(r)"), start=10
    ):
        sales.cell(9, column, status)
    sales.cell(11, 2, "Retail & food services, ")
    sales.cell(12, 2, "  total")
    for column, value in enumerate(
        (763705, 757036, 754013, 714568, 721903), start=10
    ):
        sales.cell(12, column, value)

    changes = workbook.create_sheet("Table 2.")
    changes.cell(
        3,
        1,
        "(Estimates are shown as percents and are based on official survey data.)",
    )
    changes.cell(8, 3, "May. 2026 Advance")
    changes.cell(8, 5, "Apr. 2026 Revised")
    changes.cell(11, 3, "Apr. 2026")
    changes.cell(11, 4, "May. 2025")
    changes.cell(11, 5, "Mar. 2026")
    changes.cell(11, 6, "Apr. 2025")
    changes.cell(14, 2, "Retail & food services, ")
    changes.cell(15, 2, "  total")
    for column, value in enumerate((0.9, 6.9, 0.4, 4.8), start=3):
        changes.cell(15, column, value)

    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def _census_release_result() -> ProviderResult:
    fetched_at = datetime(2026, 6, 17, 12, tzinfo=UTC)
    raw_bytes, bundle_metadata = build_evidence_bundle(
        provider="census-release",
        dataset="marts:retail-food-services",
        responses=(
            EvidenceResponse(
                role="current-workbook",
                url=CENSUS_MARTS_CURRENT_WORKBOOK,
                content_type=XLSX_CONTENT_TYPE,
                raw_bytes=_census_release_workbook(),
                request_witness={"method": "GET", "scope": "current"},
                response_witness={
                    "status_code": 200,
                    "last_modified": "June 17, 2026",
                    "retrieved_at": fetched_at.isoformat(),
                },
            ),
        ),
    )
    records, replay_metadata = CensusMARTSReleaseProvider.replay_evidence_bundle(
        raw_bytes
    )
    return ProviderResult(
        provider="census-release",
        dataset="marts:retail-food-services",
        records=records,
        fetched_at=fetched_at,
        raw_bytes=raw_bytes,
        metadata={**bundle_metadata, **replay_metadata},
    )


def test_census_marts_uses_correct_endpoint_full_history_and_sanitized_artifact():
    def handler(request):
        assert request.url.path == "/data/timeseries/eits/marts"
        assert request.url.params["key"] == "census-test-key"
        assert request.url.params["category_code"] == "44X72"
        assert request.url.params["seasonally_adj"] == "yes"
        assert request.url.params["data_type_code"] == "SM"
        assert request.url.params["time"] == "from 1992"
        assert "cell_value" in request.url.params["get"]
        return httpx.Response(200, json=_census_history_payload())

    result = CensusMARTSProvider(
        api_key="census-test-key", client=_client(handler)
    ).monthly_retail_sales(time="from 1992", require_complete_history=True)

    assert result.ok
    assert result.dataset == "marts:44X72:SM:yes"
    assert result.row_count == 1226
    assert result.metadata["level_count"] == 413
    assert result.metadata["history_start"] == "1992-01-01"
    assert result.metadata["latest_value_date"] == "2026-05-01"
    by_series_date = {
        (item["series_id"], item["date"]): item for item in result.records
    }
    latest = by_series_date[("CENSUS-API-MRTS-44X72-SM-SA", "2026-05-01")]
    assert latest["value"] == Decimal("763705")
    assert latest["metadata"]["estimate_status"] == "advance"
    assert by_series_date[("CENSUS-API-MRTS-44X72-SM-SA-MOM", "2026-05-01")][
        "value"
    ] == Decimal("0.9")
    assert by_series_date[("CENSUS-API-MRTS-44X72-SM-SA-MOM", "2026-04-01")][
        "value"
    ] == Decimal("0.4")
    assert by_series_date[("CENSUS-API-MRTS-44X72-SM-SA-YOY", "2026-05-01")][
        "value"
    ] == Decimal("6.9")
    assert result.metadata["vintage_policy"] == "current-latest-vintage"
    assert result.metadata["source_revision_date"] is None
    assert "not a source vintage" in result.metadata["revision_note"]
    assert result.metadata["license"].startswith("CC0-1.0")
    assert "not endorsed or certified" in result.metadata["attribution_notice"]
    artifact = result.metadata["artifacts"][0]
    assert artifact["url"] == "https://api.census.gov/data/timeseries/eits/marts"
    assert "census-test-key" not in str(result.metadata)
    evidence = parse_evidence_bundle(
        result.raw_bytes,
        expected_provider="census",
        expected_dataset="marts:44X72:SM:yes",
    )
    assert set(evidence.responses) == {"marts-api-response"}
    assert json.loads(evidence.responses["marts-api-response"]) == (
        _census_history_payload()
    )
    assert b"census-test-key" not in result.raw_bytes
    replay_records, replay_metadata = CensusMARTSProvider.replay_evidence_bundle(
        result.raw_bytes,
        expected_dataset=result.dataset,
    )
    assert replay_records == result.records
    assert replay_metadata["requested_time"] == "from 1992"


@pytest.mark.django_db
def test_census_marts_v2_persists_private_append_only_artifact_and_derived_rows(
    settings,
    tmp_path,
):
    settings.RAW_ARTIFACT_ROOT = tmp_path / "raw-artifacts"

    def handler(_request):
        return httpx.Response(200, json=_census_history_payload())

    result = CensusMARTSProvider(
        api_key="census-test-key", client=_client(handler)
    ).monthly_retail_sales(time="from 1992", require_complete_history=True)
    run = record_provider_result(
        result, persist=_store_census_marts_observations_v2
    )

    assert run.status == "success"
    assert run.row_count == 1226
    artifact = RawArtifact.objects.get(run=run)
    assert artifact.uri.startswith("private://census/")
    assert "census-test-key" not in artifact.uri
    path = (
        Path(settings.RAW_ARTIFACT_ROOT)
        / artifact.sha256[:2]
        / f"{artifact.sha256}.bin"
    )
    assert path.read_bytes() == result.raw_bytes
    derived = Observation.objects.get(
        source__key="census",
        series__key="census-api-mrts-44x72-sm-sa-mom",
        value_date=datetime(2026, 5, 1, tzinfo=UTC),
    )
    assert derived.metadata["input_batch_ids"] == [str(run.batch_id)]
    assert [item["value_date"] for item in derived.metadata["input_lineage"]] == [
        "2026-04-01",
        "2026-05-01",
    ]

    first_signature = (derived.pk, derived.value, derived.updated_at)
    repeated_result = CensusMARTSProvider(
        api_key="census-test-key", client=_client(handler)
    ).monthly_retail_sales(time="from 1992", require_complete_history=True)
    repeated = record_provider_result(
        repeated_result,
        persist=_store_census_marts_observations_v2,
    )
    assert repeated.status == "success"
    assert repeated.batch_id != run.batch_id
    derived.refresh_from_db()
    assert (derived.pk, derived.value, derived.updated_at) == first_signature
    assert Observation.objects.filter(batch_id=repeated.batch_id).count() == 1226


@pytest.mark.parametrize(
    "persistence_order",
    [("release", "api"), ("api", "release")],
)
@pytest.mark.django_db
def test_census_release_and_api_persist_with_distinct_series_identities(
    persistence_order,
    settings,
    tmp_path,
):
    settings.RAW_ARTIFACT_ROOT = tmp_path / "raw-artifacts"
    api_result = CensusMARTSProvider(
        api_key="census-test-key",
        client=_client(
            lambda _request: httpx.Response(200, json=_census_history_payload())
        ),
    ).monthly_retail_sales(time="from 1992", require_complete_history=True)
    release_result = _census_release_result()
    assert api_result.ok, api_result.error
    assert release_result.ok, release_result.error
    results = {"api": api_result, "release": release_result}

    runs = {
        identity: record_provider_result(
            results[identity],
            persist=_store_census_marts_observations_v2,
        )
        for identity in persistence_order
    }

    assert {identity: run.status for identity, run in runs.items()} == {
        "api": "success",
        "release": "success",
    }, {identity: run.error for identity, run in runs.items()}
    api_keys = set(
        Observation.objects.filter(source__key="census").values_list(
            "series__key", flat=True
        )
    )
    release_keys = set(
        Observation.objects.filter(source__key="census-release").values_list(
            "series__key", flat=True
        )
    )
    assert api_keys == {
        "census-api-mrts-44x72-sm-sa",
        "census-api-mrts-44x72-sm-sa-mom",
        "census-api-mrts-44x72-sm-sa-yoy",
    }
    assert release_keys == {
        "census-mrts-44x72-sm-sa",
        "census-mrts-44x72-sm-sa-mom",
        "census-mrts-44x72-sm-sa-yoy",
    }
    assert api_keys.isdisjoint(release_keys)
    assert SeriesDefinition.objects.filter(key__in=api_keys | release_keys).count() == 6


def test_census_marts_fails_closed_on_missing_history_month_and_redacts_http_error():
    payload = _census_history_payload()
    missing = [payload[0], *payload[1:20], *payload[21:]]

    result = CensusMARTSProvider(
        api_key="census-test-key",
        client=_client(lambda _request: httpx.Response(200, json=missing)),
    ).monthly_retail_sales(time="from 1992", require_complete_history=True)

    assert not result.ok
    assert "missing month" in result.error

    failed = CensusMARTSProvider(
        api_key="census-test-key",
        client=_client(lambda _request: httpx.Response(403, text="denied")),
    ).monthly_retail_sales(time="from 1992", require_complete_history=True)

    assert not failed.ok
    assert failed.error == "Census MARTS request failed: HTTP 403"
    assert "census-test-key" not in failed.error


@pytest.mark.parametrize(
    ("violation", "message"),
    [
        ("header", "headers do not match"),
        ("row-width", "row width"),
        ("dimension", "violated requested dimensions"),
    ],
)
def test_census_marts_parser_rejects_schema_and_dimension_fallbacks(
    violation,
    message,
):
    payload = _census_history_payload()
    if violation == "header":
        payload[0][0] = "unexpected_program_code"
    elif violation == "row-width":
        payload[1].pop()
    else:
        category_index = payload[0].index("category_code")
        payload[1][category_index] = ""

    with pytest.raises(ValueError, match=message):
        CensusMARTSProvider.parse_response_bytes(
            json.dumps(payload).encode(),
            category_code="44X72",
            seasonally_adjusted=True,
            require_complete_history=True,
            as_of_date=date(2026, 5, 31),
        )


def test_census_marts_derives_month_from_slot_when_response_uses_year_predicate():
    def handler(_request):
        return httpx.Response(
            200,
            json=[
                    [
                        "program_code",
                        "cell_value",
                        "time_slot_id",
                        "time_slot_date",
                        "time_slot_name",
                        "error_data",
                        "seasonally_adj",
                        "category_code",
                        "data_type_code",
                        "time",
                    ],
                    [
                        "MARTS",
                        "100",
                        "M02",
                        "",
                        "February",
                        "no",
                        "no",
                        "44X72",
                        "SM",
                        "2025",
                    ],
            ],
        )

    result = CensusMARTSProvider(
        api_key="census-test-key", client=_client(handler)
    ).monthly_retail_sales(time="2025", seasonally_adjusted=False)

    assert result.ok
    assert result.records[0]["date"] == "2025-02-01"
    assert result.records[0]["series_id"].endswith("-NSA")


@pytest.mark.django_db
def test_legacy_generic_rows_cannot_publish_gdp_or_consumer():
    fetched_at = datetime(2026, 7, 12, tzinfo=UTC)
    bea = ProviderResult(
        provider="bea",
        dataset="bea-fixture",
        fetched_at=fetched_at,
        records=[
            {"series_id": "BEA-A191RL", "date": "2026-04-01", "value": "2.1"},
            {"series_id": "BEA-DPCERL", "date": "2026-04-01", "value": "1.7"},
            {"series_id": "BEA-REAL-PCE-MOM", "date": "2026-06-01", "value": "0.3"},
            {"series_id": "BEA-REAL-DPI-MOM", "date": "2026-06-01", "value": "0.2"},
            {
                "series_id": "BEA-PERSONAL-SAVING-RATE",
                "date": "2026-06-01",
                "value": "3.1",
            },
        ],
    )
    census = ProviderResult(
        provider="census-release",
        dataset="marts:retail-food-services",
        fetched_at=fetched_at,
        records=[
            {
                "series_id": "CENSUS-MRTS-44X72-SM-SA",
                "date": "2026-06-01",
                "value": "763700.0",
            },
            {
                "series_id": "CENSUS-MRTS-44X72-SM-SA-MOM",
                "date": "2026-06-01",
                "value": "0.4",
            },
            {
                "series_id": "CENSUS-MRTS-44X72-SM-SA-YOY",
                "date": "2026-06-01",
                "value": "3.9",
            },
        ],
    )
    record_provider_result(bea, persist=store_series_observations)
    record_provider_result(census, persist=store_series_observations)

    dashboards = {item.key: item for item in publish_official_dashboards()}

    assert "gdp" not in dashboards
    assert "consumer" not in dashboards
